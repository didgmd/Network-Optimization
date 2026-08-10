from openpyxl import Workbook
from PIL import Image
import openpyxl
import os
import time

def save_data_to_excel(list_to_save, list_name:str):
    file = "./" + list_name + ".xlsx"
    # print(f"save {list_name} to {file}")
    # print(f"{os.path.exists(file)}")
    #检查文件是否存在，不存在则创建
    if not os.path.exists(file):
        #print(f"file {file} not exist, create it")
        openpyxl.Workbook().save(file)

    wb = openpyxl.load_workbook(file)  
    sheet_name = wb.sheetnames
    table = wb[sheet_name[0]]
    n_rows = table.max_row
    #print(f"n_rows of {list_name} is {n_rows}")
    n_columns = 1
    for value in list_to_save:
        table.cell(n_rows+1, n_columns).value = value
        n_columns += 1
    wb.save(file)
