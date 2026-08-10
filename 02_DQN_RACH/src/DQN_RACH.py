import os

import sys
from collections import Counter
import numpy as np
import torch
import matplotlib.pyplot as plt
import inspect
from inspect import getframeinfo, stack
import random
import torch.nn as nn
import torch.nn.functional as F
import openpyxl

DEBUG_MODE = False  # 是否打印调试信息


def debug(msg):
    if not DEBUG_MODE:
        return

    caller = getframeinfo(stack()[1][0])
    print("Line %d #" % caller.lineno, end=" ")

    print(msg)


# 定义设备
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(f"device is {device}")

# noinspection DuplicatedCode
N_STATE = 54  # Index: 0, 1, 2, ..., N_STATE-1


# noinspection DuplicatedCode
class Preamble:
    def __init__(self, index):
        self.preamble = index
        self.assignedToList = False
        self.assignedToUser = False


preamble_pool = []
for i in range(N_STATE):
    preamble_pool.append(Preamble(i))

N_STATE_FEATURES = 6

NUM_PREAMBLE_LOW_LIMIT = 0

num_h2h_high = NUM_PREAMBLE_LOW_LIMIT  # 分配给h2h_high的前导码数
nun_h2h_medium = NUM_PREAMBLE_LOW_LIMIT  # 分配给h2h_low的前导码数
num_h2h_low = NUM_PREAMBLE_LOW_LIMIT  # 分配给h2h_low的前导码数
num_m2m_high = NUM_PREAMBLE_LOW_LIMIT  # 分配给m2m_high的前导码数
num_m2m_medium = NUM_PREAMBLE_LOW_LIMIT  # 分配给m2m_low的前导码数
num_m2m_low = NUM_PREAMBLE_LOW_LIMIT  # 分配给m2m_low的前导码数

state_feature = []
num_h2h_medium = 1
# The sequence of input features is: num_m2m_low, num_m2m_high, num_h2h_low, num_h2h_high
for idx_h2h_high in range(0, N_STATE - 5):
    for idx_h2h_medium in range(idx_h2h_high + 1, N_STATE - 4):
        for idx_h2h_low in range(idx_h2h_medium + 1, N_STATE - 3):
            for idx_m2m_high in range(idx_h2h_low + 1, N_STATE - 2):
                for idx_m2m_medium in range(idx_m2m_high + 1, N_STATE - 1):
                    for idx_m2m_low in range(idx_m2m_medium + 1, N_STATE):
                        state_feature.append(
                            [
                                num_h2h_high,
                                num_h2h_medium,
                                num_h2h_low,
                                num_m2m_high,
                                num_m2m_medium,
                                num_m2m_low,
                            ]
                        )
                        num_m2m_low += 1
                    num_m2m_low = 1
                    num_m2m_medium += 1
                num_m2m_medium = 1
                num_m2m_high += 1
            num_m2m_high = 1
            num_h2h_low += 1
        num_h2h_low = 1
        num_h2h_medium += 1
    num_h2h_medium = 1
    num_h2h_high += 1

# 暂时不显示
# debug(f"state_feature is {state_feature}")
# debug(f"len(state_feature) is {len(state_feature)}")
# debug(f"{inspect.currentframe().f_lineno}")


h2h_high_action_range = [-1, 0, 1]
h2h_medium_action_range = [-1, 0, 1]
h2h_low_action_range = [-1, 0, 1]
m2m_high_action_range = [-1, 0, 1]
m2m_medium_action_range = [-1, 0, 1]
m2m_low_action_range = [-1, 0, 1]
action_feature = []

for idx_m2m_low_action in m2m_low_action_range:
    for idx_m2m_high_action in m2m_high_action_range:
        for idx_m2m_medium_action in m2m_medium_action_range:
            for idx_h2h_low_action in h2h_low_action_range:
                for idx_h2h_high_action in h2h_high_action_range:
                    for idx_h2h_medium_action in h2h_medium_action_range:
                        action_feature.append(
                            [
                                idx_h2h_high_action,
                                idx_h2h_medium_action,
                                idx_h2h_low_action,
                                idx_m2m_high_action,
                                idx_m2m_medium_action,
                                idx_m2m_low_action,
                            ]
                        )

N_ACTIONS_FEATURES = len(action_feature)
# 暂时先不显示
# debug(action_feature)
# debug(f"len(action_feature) is {len(action_feature)}")
# debug(N_ACTIONS_FEATURES)
# debug(f"{inspect.currentframe().f_lineno}")

EPSILON = 0.9
ALPHA = 0.1
GAMMA = 0.9
MAX_EPISODES = 10
N_HIDDEN = 2000


class QNet(nn.Module):
    def __init__(self, n_state_features, n_hidden, n_actions_features):
        super(QNet, self).__init__()
        self.fc1 = nn.Linear(n_state_features, n_hidden)
        # self.fc1.weight.data.normal_(0, 0.1)
        self.fc2 = nn.Linear(n_hidden, n_actions_features)
        # self.fc2.weight.data.normal_(0, 0.1)

    def forward(self, x):
        x = self.fc1(x)
        x = F.relu(x)
        actions_value = self.fc2(x)
        return actions_value


# 初始化数组
unserved_user_list = []
unserved_m2m_user_low_list = []
unserved_m2m_user_medium_list = []
unserved_m2m_user_high_list = []
unserved_h2h_user_low_list = []
unserved_h2h_user_medium_list = []
unserved_h2h_user_high_list = []
served_user_list = []
blocked_user_list = []

# 定义用户类型
H2H_HIGH = 0
H2H_MEDIUM = 1
H2H_LOW = 2
M2M_HIGH = 3
M2M_MEDIUM = 4
M2M_LOW = 5

H2H = 6
M2M = 7
ALL = 8


class User:
    def __init__(self, user_type):
        self.user_type = user_type
        self.served = 0
        self.unserved_frame = 0
        self.joined_frame = 0
        self.served_frame = 0
        self.assigned_frame = 0
        self.preamble = -1
        self.delay_frame = 0
        self.collisionDetected = 0
        self.served_frame_minus = False

    def status(self):
        print(
            f"user_type is {self.user_type}, served_frame is {self.served_frame}, "
            f"unserved_frame is {self.unserved_frame}, preamble is {self.preamble}"
        )


class PreambleList:  # 前导码列表
    def __init__(self):
        self.preambleList = []  # 前导码列表
        self.numberOfPreamble = 0  # 前导码数
        self.numOfAvailablePreamble = 0  # 当前可用前导码数

    def update(self, number_of_preamble):  # 更新前导码列表
        debug(
            f"self.numberOfPreamble {self.numberOfPreamble} number_of_preamble {number_of_preamble} len(self.preambleList) {len(self.preambleList)}"
        )

        self.numberOfPreamble = number_of_preamble  # 更新前导码数

        # 如果应有的前导码数大于当前拥有的前导码数，则需要将前导码池中的前导码分配给前导码列表
        if self.numberOfPreamble > len(self.preambleList):
            num_of_preamble_to_add = self.numberOfPreamble - len(self.preambleList)
            debug(f"num_of_preamble_to_add {num_of_preamble_to_add}")

            for i in range(num_of_preamble_to_add):
                # 遍历前导码池
                for candidate_preamble in preamble_pool:  # 遍历前导码池
                    if candidate_preamble.assignedToList:  # 如果前导码已被分配给前导码列表
                        continue
                    elif not candidate_preamble.assignedToList:  # 如果前导码未被分配给前导码列表
                        candidate_preamble.assignedToList = True  # 标记前导码已被分配给前导码列表
                        self.preambleList.append(candidate_preamble)  # 将前导码添加到前导码列表
                        self.numOfAvailablePreamble += 1  # 当前可用前导码数加1
                        break
        elif self.numberOfPreamble < len(self.preambleList):
            # 计算需要还几个前导码到前导码池
            num_of_preamble_to_return = len(self.preambleList) - self.numberOfPreamble
            debug(f"num_of_preamble_to_return {num_of_preamble_to_return}")
            for objPreamble in self.preambleList:
                if objPreamble.assignedToUser:
                    continue
                elif not objPreamble.assignedToUser:
                    objPreamble.assignedToList = False
                    self.preambleList.remove(objPreamble)
                    self.numOfAvailablePreamble -= 1
                    num_of_preamble_to_return -= 1
                    if num_of_preamble_to_return == 0:  # 避免还多了
                        break

    def clear(self):
        self.preambleList.clear()
        self.numberOfPreamble = 0
        self.numOfAvailablePreamble = 0


def update_preamble_lists(curr_state):  # 初始化前导码列表
    # debug(f"{inspect.currentframe().f_lineno}")

    preambleListForH2hHigh.update(curr_state[H2H_HIGH])
    PreambleListForH2hMedium.update(curr_state[H2H_MEDIUM])
    preambleListForH2hLow.update(curr_state[H2H_LOW])
    preambleListForM2mHigh.update(curr_state[M2M_HIGH])
    PreambleListForM2mMedium.update(curr_state[M2M_MEDIUM])
    preambleListForM2mLow.update(curr_state[M2M_LOW])

    # debug(f"{inspect.currentframe().f_lineno}")


def revoke_preamble_from_served_user():
    for served_user in served_user_list:
        served_user.served_frame_minus = False

    # 服务用户被回收前导码并且清除列表
    while True:
        flag_served_user_removed = False
        for served_user in served_user_list:  # 遍历服务用户列表
            if served_user.served_frame_minus:
                continue
            debug(
                f"served_user.user_type {served_user.user_type} served_user.preamble {served_user.preamble} served_user.served_frame {served_user.served_frame}"
            )
            served_user.served_frame -= 1  # 服务帧数减1
            served_user.served_frame_minus = True

            if served_user.served_frame == 0:  # 服务用户的服务帧数为0
                # 回收前的服务用户列表
                debug(
                    f"Before Revoke: len(served_user_list) is {len(served_user_list)}"
                )
                if served_user.user_type == H2H_HIGH:  # 如果用户类型是h2h_high
                    # 回收前的h2h_high前导码列表
                    debug(
                        f"Before Revoke: len(preambleListForH2hHigh.preambleList) is {len(preambleListForH2hHigh.preambleList)}"
                    )
                    for (
                        objPreamble
                    ) in preambleListForH2hHigh.preambleList:  # 遍历h2h_high前导码列表
                        if objPreamble.preamble == served_user.preamble:  # 如果前导码已被分配给用户
                            objPreamble.assignedToUser = False  # 标记前导码未被分配给用户
                            debug(
                                f"After Revoke: len(preambleListForH2hHigh.preambleList) is {len(preambleListForH2hHigh.preambleList)}"
                            )
                            preambleListForH2hHigh.numOfAvailablePreamble += (
                                1  # 当前可用前导码数加1
                            )
                            break
                elif served_user.user_type == H2H_MEDIUM:  # 如果用户类型是h2h_low
                    debug(
                        f"Before Revoke: len(preambleListForH2hMedium.preambleList is {len(PreambleListForH2hMedium.preambleList)}"
                    )
                    for objPreamble in PreambleListForH2hMedium.preambleList:
                        if objPreamble.preamble == served_user.preamble:
                            objPreamble.assignedToUser = False
                            debug(
                                f"After Revoke: len(PreambleListForH2hMedium.preambleList is {len(PreambleListForH2hMedium.preambleList)}"
                            )
                            PreambleListForH2hMedium.numOfAvailablePreamble += 1
                            break

                elif served_user.user_type == H2H_LOW:
                    debug(
                        f"Before Revoke: len(preambleListForH2hLow.preambleList is {len(preambleListForH2hLow.preambleList)}"
                    )
                    for objPreamble in preambleListForH2hLow.preambleList:
                        if objPreamble.preamble == served_user.preamble:
                            objPreamble.assignedToUser = False
                            debug(
                                f"After Revoke: len(preambleListForH2hLow.preambleList is {len(preambleListForH2hLow.preambleList)}"
                            )
                            preambleListForH2hLow.numOfAvailablePreamble += 1
                            break
                elif served_user.user_type == M2M_HIGH:
                    debug(
                        f"Before Revoke: len(preambleListForM2mHigh.preambleList is {len(preambleListForM2mHigh.preambleList)}"
                    )
                    for objPreamble in preambleListForM2mHigh.preambleList:
                        if objPreamble.preamble == served_user.preamble:
                            objPreamble.assignedToUser = False
                            debug(
                                f"After Revoke: len(preambleListForM2mHigh.preambleList is {len(preambleListForM2mHigh.preambleList)}"
                            )
                            preambleListForM2mHigh.numOfAvailablePreamble += 1
                            break
                elif served_user.user_type == M2M_MEDIUM:
                    debug(
                        f"Before Revoke: len(PreambleListForM2mMedium.preambleList is {len(PreambleListForM2mMedium.preambleList)}"
                    )
                    for objPreamble in PreambleListForM2mMedium.preambleList:
                        if objPreamble.preamble == served_user.preamble:
                            objPreamble.assignedToUser = False
                            debug(
                                f"After Revoke: len(preambleListForM2mMedium.preambleList is {len(PreambleListForM2mMedium.preambleList)}"
                            )
                            PreambleListForM2mMedium.numOfAvailablePreamble += 1
                            break

                elif served_user.user_type == M2M_LOW:
                    debug(
                        f"Before Revoke: len(preambleListForM2mLow.preambleList is {len(preambleListForM2mLow.preambleList)}"
                    )
                    for objPreamble in preambleListForM2mLow.preambleList:
                        if objPreamble.preamble == served_user.preamble:
                            objPreamble.assignedToUser = False
                            debug(
                                f"After Revoke: len(preambleListForM2mLow.preambleList is {len(preambleListForM2mLow.preambleList)}"
                            )
                            preambleListForM2mLow.numOfAvailablePreamble += 1
                            break

                served_user_list.remove(served_user)  # 移除用户以避免重复检查
                debug(f"After Revoke: len(served_user_list) is {len(served_user_list)}")
                flag_served_user_removed = True
                break
            else:
                continue

        if not flag_served_user_removed:
            debug(
                f"Preamble revoke finished since there is no served_user with served_frame: 0"
            )
            break


# noinspection DuplicatedCode
def add_new_user_and_preamble_in_every_frame(frame_count: int, state):
    count_of_h2h_high_user = 0  # h2h_high用户数
    count_of_h2h_medium_user = 0  # h2h_medium用户数
    count_of_h2h_low_user = 0  # h2h_low用户数
    count_of_m2m_high_user = 0  # m2m_high用户数
    count_of_m2m_medium_user = 0  # m2m_medium用户数
    count_of_m2m_low_user = 0  # m2m_low用户数

    if globalModeSelector == 1:  # 如果是固定用户模式
        # 添加各种类型的用户
        for i in range(num_of_h2h_user_high_per_frame):
            h2h_user_high = User(H2H_HIGH)
            h2h_user_high.joined_frame = frame_count
            unserved_user_list.append(h2h_user_high)
            unserved_h2h_user_high_list.append(h2h_user_high)

        # debug(unserved_h2h_user_high_list)  # 打印未服务h2h_high用户列表     暂时不显示
        for i in range(num_of_h2h_user_medium_per_frame):
            h2h_user_medium = User(H2H_MEDIUM)
            h2h_user_medium.joined_frame = frame_count
            unserved_user_list.append(h2h_user_medium)
            unserved_h2h_user_medium_list.append(h2h_user_medium)

        for i in range(num_of_h2h_user_low_per_frame):  # 添加h2h_low用户
            h2h_user_low = User(H2H_LOW)  # 创建h2h_low用户
            h2h_user_low.joined_frame = frame_count  # 记录用户加入的帧数
            unserved_user_list.append(h2h_user_low)  # 将用户添加到未服务用户列表
            unserved_h2h_user_low_list.append(h2h_user_low)  # 将用户添加到未服务h2h_low用户列表

        # debug(unserved_h2h_user_low_list)   暂时不显示

        for i in range(num_of_m2m_user_high_per_frame):  # 添加m2m_high用户
            m2m_user_high = User(M2M_HIGH)  # 创建m2m_high用户
            m2m_user_high.joined_frame = frame_count  # 记录用户加入的帧数
            unserved_user_list.append(m2m_user_high)  # 将用户添加到未服务用户列表
            unserved_m2m_user_high_list.append(m2m_user_high)  # 将用户添加到未服务m2m_high用户列表

        # debug(unserved_m2m_user_high_list)    暂时不显示
        for i in range(num_of_m2m_user_medium_per_frame):
            m2m_user_medium = User(M2M_MEDIUM)
            m2m_user_medium.joined_frame = frame_count
            unserved_user_list.append(m2m_user_medium)
            unserved_m2m_user_medium_list.append(m2m_user_medium)

        for i in range(num_of_m2m_user_low_per_frame):  # 添加m2m_low用户
            m2m_user_low = User(M2M_LOW)  # 创建m2m_low用户
            m2m_user_low.joined_frame = frame_count  # 记录用户加入的帧数
            unserved_user_list.append(m2m_user_low)  # 将用户添加到未服务用户列表
            unserved_m2m_user_low_list.append(m2m_user_low)  # 将用户添加到未服务m2m_low用户列表

        # debug(unserved_m2m_user_low_list)    暂时不显示

    elif globalModeSelector == 2:  # 如果是随机用户模式
        for i in range(numOfRandomUserPerFrame):
            objUser = User(
                np.random.choice(
                    [H2H_HIGH, H2H_MEDIUM, H2H_LOW, M2M_HIGH, M2M_MEDIUM, M2M_LOW]
                )
            )
            objUser.joined_frame = frame_count
            unserved_user_list.append(objUser)

            if objUser.user_type == H2H_HIGH:
                count_of_h2h_high_user += 1
                unserved_h2h_user_high_list.append(objUser)
                listForH2hUserHigh.append(objUser)
                debug(f"len(listForH2hUserHigh) is {len(listForH2hUserHigh)}")
            elif objUser.user_type == H2H_MEDIUM:
                count_of_h2h_medium_user += 1
                unserved_h2h_user_medium_list.append(objUser)
                listForH2hUserMedium.append(objUser)
                debug(f"len(listForH2hUserMedium) is {len(listForH2hUserMedium)}")

            elif objUser.user_type == H2H_LOW:
                count_of_h2h_low_user += 1
                unserved_h2h_user_low_list.append(objUser)
                listForH2hUserLow.append(objUser)
                debug(f"len(listForH2hUserLow) is {len(listForH2hUserLow)}")
            elif objUser.user_type == M2M_HIGH:
                count_of_m2m_high_user += 1
                unserved_m2m_user_high_list.append(objUser)
                listForM2mUserHigh.append(objUser)
                debug(f"len(listForM2mUserHigh) is {len(listForM2mUserHigh)}")

            elif objUser.user_type == M2M_MEDIUM:
                count_of_m2m_medium_user += 1
                unserved_m2m_user_medium_list.append(objUser)
                listForM2mUserMedium.append(objUser)
                debug(f"len(listForM2mUserMedium) is {len(listForM2mUserMedium)}")

            elif objUser.user_type == M2M_LOW:
                count_of_m2m_low_user += 1
                unserved_m2m_user_low_list.append(objUser)
                listForM2mUserLow.append(objUser)
                debug(f"len(listForM2mUserLow) is {len(listForM2mUserLow)}")

    # debug(unserved_user_list)   暂时不显示

    # 当6种用户需要的前导码总数不超过可用前导码总数时，全部分配，从而提高前导码利用率，达到平衡资源的效果
    current_count_of_h2h_high = 0
    current_count_of_h2h_medium = 0
    current_count_of_h2h_low = 0
    current_count_of_m2m_high = 0
    current_count_of_m2m_medium = 0
    current_count_of_m2m_low = 0

    # 遍历未被服务用户
    for user in unserved_user_list:
        if user.user_type == H2H_HIGH:
            current_count_of_h2h_high += 1
        elif user.user_type == H2H_MEDIUM:
            current_count_of_h2h_medium += 1
        elif user.user_type == H2H_LOW:
            current_count_of_h2h_low += 1
        elif user.user_type == M2M_HIGH:
            current_count_of_m2m_high += 1
        elif user.user_type == M2M_MEDIUM:
            current_count_of_m2m_medium += 1
        elif user.user_type == M2M_LOW:
            current_count_of_m2m_low += 1

    # 遍历已服务且未归还前导码的用户
    for user in served_user_list:
        if user.user_type == H2H_HIGH:
            current_count_of_h2h_high += 1
        elif user.user_type == H2H_MEDIUM:
            current_count_of_h2h_medium += 1
        elif user.user_type == H2H_LOW:
            current_count_of_h2h_low += 1
        elif user.user_type == M2M_HIGH:
            current_count_of_m2m_high += 1
        elif user.user_type == M2M_MEDIUM:
            current_count_of_m2m_medium += 1
        elif user.user_type == M2M_LOW:
            current_count_of_m2m_low += 1

    # 针对输入用户数判断前导码资源
    # debug(state)
    #
    # # 检查条件
    # if (current_count_of_h2h_high + current_count_of_h2h_medium + current_count_of_h2h_low + current_count_of_m2m_high + current_count_of_m2m_medium + current_count_of_m2m_low) <= sum(state):
    #     state[0] = current_count_of_h2h_high
    #     state[1] = current_count_of_h2h_medium
    #     state[2] = current_count_of_h2h_low
    #     state[3] = current_count_of_m2m_high
    #     state[4] = current_count_of_m2m_medium
    #     state[5] = current_count_of_m2m_low
    #
    # debug(state)
    # update_preamble_lists(state)
    # debug(f"______________________________________")

    flag_h2h_high_no_preamble_or_no_user = True  # 标记h2h_high用户是否没有前导码或者没有用户
    flag_h2h_medium_no_preamble_or_no_user = True
    flag_h2h_low_no_preamble_or_no_user = True  # 标记h2h_low用户是否没有前导码或者没有用户
    flag_m2m_high_no_preamble_or_no_user = True  # 标记m2m_high用户是否没有前导码或者没有用户
    flag_m2m_medium_no_preamble_or_no_user = True
    flag_m2m_low_no_preamble_or_no_user = True  # 标记m2m_low用户是否没有前导码或者没有用户

    for user in unserved_user_list:  # 遍历未服务用户列表
        # debug(f"user.user_type is {user.user_type}")
        if user.user_type == H2H_HIGH:
            flag_h2h_high_no_preamble_or_no_user = False
        elif user.user_type == H2H_MEDIUM:
            flag_h2h_medium_no_preamble_or_no_user = False
        elif user.user_type == H2H_LOW:
            flag_h2h_low_no_preamble_or_no_user = False
        elif user.user_type == M2M_HIGH:
            flag_m2m_high_no_preamble_or_no_user = False
        elif user.user_type == M2M_MEDIUM:
            flag_m2m_medium_no_preamble_or_no_user = False
        elif user.user_type == M2M_LOW:
            flag_m2m_low_no_preamble_or_no_user = False
    debug(f"----------------------------------------")
    # 每帧开始四种用户对应数量进行分配
    debug(f"len(original_unserved_user_list) is {len(unserved_user_list)}")
    debug(
        f"len(original_unserved_h2h_user_high_list) is {len(unserved_h2h_user_high_list)}"
    )
    debug(
        f"len(original_unserved_h2h_user_medium_list) is {len(unserved_h2h_user_medium_list)}"
    )
    debug(
        f"len(original_unserved_h2h_user_low_list) is {len(unserved_h2h_user_low_list)}"
    )
    debug(
        f"len(original_unserved_m2m_user_high_list) is {len(unserved_m2m_user_high_list)}"
    )
    debug(
        f"len(original_unserved_m2m_user_medium_list) is {len(unserved_m2m_user_medium_list)}"
    )
    debug(
        f"len(original_unserved_m2m_user_low_list) is {len(unserved_m2m_user_low_list)}"
    )
    debug(f"==========================================")
    # num_of_h2h_user_high_per_frame.append(count_of_h2h_high_user)
    # num_of_h2h_user_low_per_frame.append(count_of_h2h_low_user)
    # num_of_m2m_user_high_per_frame.append(count_of_m2m_high_user)
    # num_of_m2m_user_low_per_frame.append(count_of_m2m_low_user)
    num1_of_h2h_user_high_per_frame.append(len(unserved_h2h_user_high_list))
    num1_of_h2h_user_medium_per_frame.append(len(unserved_h2h_user_medium_list))
    num1_of_h2h_user_low_per_frame.append(len(unserved_h2h_user_low_list))
    num1_of_m2m_user_high_per_frame.append(len(unserved_m2m_user_high_list))
    num1_of_m2m_user_medium_per_frame.append(len(unserved_m2m_user_medium_list))
    num1_of_m2m_user_low_per_frame.append(len(unserved_m2m_user_low_list))

    debug(f"num1_of_h2h_user_high_per_frame is {num1_of_h2h_user_high_per_frame}")
    debug(f"num1_of_h2h_user_medium_per_frame is {num1_of_h2h_user_medium_per_frame}")
    debug(f"num1_of_h2h_user_low_per_frame is {num1_of_h2h_user_low_per_frame}")
    debug(f"num1_of_m2m_user_high_per_frame is {num1_of_m2m_user_high_per_frame}")
    debug(f"num1_of_m2m_user_medium_per_frame is {num1_of_m2m_user_medium_per_frame}")
    debug(f"num1_of_m2m_user_low_per_frame is {num1_of_m2m_user_low_per_frame}")

    debug(f"collisionDetectionMode is {collisionDetectionMode}")
    if collisionDetectionMode == 1:
        for user in unserved_user_list:
            # 在正式开始分配前导码前，对所有用户的碰撞状态进行初始化
            user.collisionDetected = 0

    debug(f"++++++++++++++++++++++++++++++++++++++++++")

    count_all_user_per_frame_collision = 0
    count_h2h_high_per_frame_collision = 0
    count_h2h_medium_per_frame_collision = 0
    count_h2h_low_per_frame_collision = 0
    count_m2m_high_per_frame_collision = 0
    count_m2m_medium_per_frame_collision = 0
    count_m2m_low_per_frame_collision = 0
    count_h2h_per_frame_collision = 0
    count_m2m_per_frame_collision = 0

    debug(f"len(unserved_user_list) is {len(unserved_user_list)}")
    while len(unserved_user_list) > 0:  # 当未服务用户列表不为空时
        if (
            flag_h2h_high_no_preamble_or_no_user
            and flag_h2h_medium_no_preamble_or_no_user
            and flag_h2h_low_no_preamble_or_no_user
            and flag_m2m_high_no_preamble_or_no_user
            and flag_m2m_medium_no_preamble_or_no_user
            and flag_m2m_low_no_preamble_or_no_user
        ):  # 如果所有用户都没有前导码或者没有用户
            debug(
                "All user have been served or there is no preamble left for any specific user type"
            )  # 所有用户都已服务或者没有前导码可用
            break

        random_user = np.random.choice(unserved_user_list)  # 从未服务用户列表中随机选择一个用户
        debug(f"random_user id is {id(random_user)}")
        # debug(f"random_user.user_type is {random_user.user_type}")
        # unused_preamble_exist = False  # 由于每次只随机选中一个用户，因此可以使用该标志位来判断是否有未被分配的前导码
        if random_user.collisionDetected == 1:
            continue
        if random_user.user_type == H2H_HIGH:  # 如果用户类型是h2h_high
            # h2h_user_high_count = 0
            if not flag_h2h_high_no_preamble_or_no_user:  # 如果没有标记h2h_high用户没有前导码或者没有用户
                if collisionDetectionMode == 1:
                    debug(
                        f"preambleListForH2hHigh.numberOfPreamble is {preambleListForH2hHigh.numberOfPreamble}"
                    )
                    if (
                        preambleListForH2hHigh.numberOfPreamble == 0
                        or len(preambleListForH2hHigh.preambleList) == 0
                    ):
                        flag_h2h_high_no_preamble_or_no_user = True
                        debug(
                            "Set flag_h2h_high_no_preamble_or_no_user to True due to no preamble is available"
                        )
                        continue
                    objPreamble = np.random.choice(preambleListForH2hHigh.preambleList)
                    if objPreamble.assignedToUser:
                        random_user.collisionDetected = 1
                        debug(f"Collision detected")
                        count_all_user_per_frame_collision += 1
                        count_h2h_per_frame_collision += 1
                        count_h2h_high_per_frame_collision += 1
                    elif not objPreamble.assignedToUser:  # 前导码未分配给用户
                        # unused_preamble_exist = True  # 标记有未被分配的前导码
                        random_user.preamble = objPreamble.preamble
                        random_user.served_frame = (
                            1 if np.random.randint(2, 8) <= 6 else 2
                        )  # 随机选择5-10帧作为服务帧数
                        objPreamble.assignedToUser = True  # 标记前导码已被分配给用户
                        debug("Assigned preamble to H2H_HIGH user")
                        preambleListForH2hHigh.numOfAvailablePreamble -= 1

                        random_user.assigned_frame = frame_count
                        random_user.delay_frame = (
                            random_user.assigned_frame - random_user.joined_frame
                        )
                        debug(
                            f"random_user is {random_user.user_type} and random_user.assigned_frame is {random_user.assigned_frame} and random_user.joined_frame is {random_user.joined_frame} and random_user.delay_frame is {random_user.delay_frame} "
                        )

                        unserved_user_list.remove(random_user)  # 从未服务用户列表中移除用户
                        served_user_list.append(random_user)  # 将用户添加到已服务用户列表
                        total_served_user_list.append(random_user)  # 将用户添加到总已服务用户列表
                        unserved_h2h_user_high_list.remove(
                            random_user
                        )  # 将用户移除未服务h2h_high用户列表

                    count_collision_detected = 0
                    if len(unserved_h2h_user_high_list) == 0:  # 如果未服务h2h_high用户列表为空
                        flag_h2h_high_no_preamble_or_no_user = (
                            True  # 标记h2h_high用户没有前导码或者没有用户
                        )
                        debug(
                            "Set flag_h2h_high_no_preamble_or_no_user to True because there is no H2H high user left"
                        )
                    else:
                        for user in unserved_h2h_user_high_list:
                            if user.collisionDetected == 1:
                                count_collision_detected += 1

                        # 如果为服务列表中的用户均为发生过碰撞的用户，则终止while循环
                        if count_collision_detected == len(unserved_h2h_user_high_list):
                            flag_h2h_high_no_preamble_or_no_user = True
                            debug(
                                "Set flag_h2h_high_no_preamble_or_no_user to True because all remain users are collision detected"
                            )

                else:
                    for (
                        objPreamble
                    ) in preambleListForH2hHigh.preambleList:  # 遍历h2h_high前导码列表
                        if objPreamble.assignedToUser:  # 如果前导码已被分配给用户
                            continue
                        elif not objPreamble.assignedToUser:  # 前导码未分配给用户
                            # unused_preamble_exist = True  # 标记有未被分配的前导码
                            random_user.preamble = objPreamble.preamble
                            random_user.served_frame = (
                                1 if np.random.randint(2, 8) <= 6 else 2
                            )  # 随机选择5-10帧作为服务帧数
                            objPreamble.assignedToUser = True  # 标记前导码已被分配给用户
                            debug("Assigned preamble to H2H_HIGH user")
                            preambleListForH2hHigh.numOfAvailablePreamble -= 1
                            random_user.assigned_frame = frame_count
                            random_user.delay_frame = (
                                random_user.assigned_frame - random_user.joined_frame
                            )
                            debug(
                                f"random_user is {random_user.user_type} and random_user.assigned_frame is {random_user.assigned_frame} and random_user.joined_frame is {random_user.joined_frame} and random_user.delay_frame is {random_user.delay_frame} "
                            )
                            # debug(f"{sys._getframe().f_lineno}")

                            unserved_user_list.remove(random_user)  # 从未服务用户列表中移除用户
                            # debug(f"{sys._getframe().f_lineno}")
                            served_user_list.append(random_user)  # 将用户添加到已服务用户列表
                            total_served_user_list.append(random_user)  # 将用户添加到总已服务用户列表
                            unserved_h2h_user_high_list.remove(
                                random_user
                            )  # 将用户移除未服务h2h_high用户列表
                            break

                    if len(unserved_h2h_user_high_list) == 0:  # 如果未服务h2h_high用户列表为空
                        flag_h2h_high_no_preamble_or_no_user = (
                            True  # 标记h2h_high用户没有前导码或者没有用户
                        )
                        debug(
                            "Set flag_h2h_high_no_preamble_or_no_user to True because there is no H2H high user left"
                        )
                # 无论是否为检测碰撞模式，该判断均应进行
                if (
                    preambleListForH2hHigh.numOfAvailablePreamble == 0
                ):  # 如果h2h_high前导码列表中没有未被分配的前导码
                    # 标记h2h_high用户没有前导码或者没有用户
                    flag_h2h_high_no_preamble_or_no_user = True
                    debug(
                        "Set flag_h2h_high_no_preamble_or_no_user to True due to no preamble is available"
                    )

        elif random_user.user_type == H2H_MEDIUM:  # 如果用户类型是h2h_medium
            if not flag_h2h_medium_no_preamble_or_no_user:
                if collisionDetectionMode == 1:
                    debug(
                        f"PreambleListForH2hMedium.numberOfPreamble is {PreambleListForH2hMedium.numberOfPreamble}"
                    )
                    if (
                        PreambleListForH2hMedium.numberOfPreamble == 0
                        or len(PreambleListForH2hMedium.preambleList) == 0
                    ):
                        flag_h2h_medium_no_preamble_or_no_user = True
                        debug(
                            "Set flag_h2h_medium_no_preamble_or_no_user to True due to no preamble is available"
                        )
                        continue
                    objPreamble = np.random.choice(
                        PreambleListForH2hMedium.preambleList
                    )
                    if objPreamble.assignedToUser:  # 如果前导码已被分配给用户
                        random_user.collisionDetected = 1  #
                        debug(f"Collision detected")
                        count_all_user_per_frame_collision += 1
                        count_h2h_per_frame_collision += 1
                        count_h2h_medium_per_frame_collision += 1
                    elif not objPreamble.assignedToUser:  # 前导码未分配给用户
                        # unused_preamble_exist = True  # 标记有未被分配的前导码
                        random_user.preamble = objPreamble.preamble
                        random_user.served_frame = (
                            1 if np.random.randint(2, 8) <= 6 else 2
                        )  # 随机选择5-10帧作为服务帧数
                        objPreamble.assignedToUser = True  # 标记前导码已被分配给用户
                        debug("Assigned preamble to H2H_MEDIUM user")
                        PreambleListForH2hMedium.numOfAvailablePreamble -= 1
                        random_user.assigned_frame = frame_count
                        random_user.delay_frame = (
                            random_user.assigned_frame - random_user.joined_frame
                        )
                        debug(
                            f"random_user is {random_user.user_type} and random_user.assigned_frame is {random_user.assigned_frame} and random_user.joined_frame is {random_user.joined_frame} and random_user.delay_frame is {random_user.delay_frame} "
                        )

                        unserved_user_list.remove(random_user)  # 从未服务用户列表中移除用户
                        served_user_list.append(random_user)  # 将用户添加到已服务用户列表
                        total_served_user_list.append(random_user)  # 将用户添加到总已服务用户列表
                        unserved_h2h_user_medium_list.remove(
                            random_user
                        )  # 将用户移除未服务h2h_high用户列表

                    count_collision_detected = 0
                    if len(unserved_h2h_user_medium_list) == 0:  # 如果未服务h2h_high用户列表为空
                        flag_h2h_medium_no_preamble_or_no_user = (
                            True  # 标记h2h_high用户没有前导码或者没有用户
                        )
                        debug(
                            "Set flag_h2h_medium_no_preamble_or_no_user to True because there is no H2H medium user left"
                        )
                    else:
                        for user in unserved_h2h_user_medium_list:
                            if user.collisionDetected == 1:
                                count_collision_detected += 1

                        if count_collision_detected == len(
                            unserved_h2h_user_medium_list
                        ):
                            flag_h2h_medium_no_preamble_or_no_user = True
                            debug(
                                "Set flag_h2h_medium_no_preamble_or_no_user to True because all remain users are collision detected"
                            )
                else:
                    for (
                        objPreamble
                    ) in PreambleListForH2hMedium.preambleList:  # 遍历h2h_high前导码列表
                        if objPreamble.assignedToUser:  # 如果前导码已被分配给用户
                            continue
                        elif not objPreamble.assignedToUser:  # 前导码未分配给用户
                            # unused_preamble_exist = True  # 标记有未被分配的前导码
                            random_user.preamble = objPreamble.preamble
                            random_user.served_frame = (
                                1 if np.random.randint(2, 8) <= 6 else 2
                            )  # 随机选择5-10帧作为服务帧数
                            objPreamble.assignedToUser = True  # 标记前导码已被分配给用户
                            debug("Assigned preamble to H2H_MEDIUM user")
                            PreambleListForH2hMedium.numOfAvailablePreamble -= 1
                            random_user.assigned_frame = frame_count
                            random_user.delay_frame = (
                                random_user.assigned_frame - random_user.joined_frame
                            )
                            debug(
                                f"random_user is {random_user.user_type} and random_user.assigned_frame is {random_user.assigned_frame} and random_user.joined_frame is {random_user.joined_frame} and random_user.delay_frame is {random_user.delay_frame} "
                            )

                            unserved_user_list.remove(random_user)  # 从未服务用户列表中移除用户
                            served_user_list.append(random_user)  # 将用户添加到已服务用户列表
                            total_served_user_list.append(random_user)  # 将用户添加到总已服务用户列表
                            unserved_h2h_user_medium_list.remove(
                                random_user
                            )  # 将用户移除未服务h2h_high用户列表
                            break

                    if len(unserved_h2h_user_medium_list) == 0:  # 如果未服务h2h_high用户列表为空
                        flag_h2h_medium_no_preamble_or_no_user = (
                            True  # 标记h2h_high用户没有前导码或者没有用户
                        )
                        debug(
                            "Set flag_h2h_medium_no_preamble_or_no_user to True because there is no H2H medium user left"
                        )

                if (
                    PreambleListForH2hMedium.numOfAvailablePreamble == 0
                ):  # 如果h2h_high前导码列表中没有未被分配的前导码
                    # 标记h2h_high用户没有前导码或者没有用户
                    flag_h2h_medium_no_preamble_or_no_user = True
                    debug(
                        "Set flag_h2h_medium_no_preamble_or_no_user to True due to no preamble is available"
                    )

        elif random_user.user_type == H2H_LOW:  # 如果用户类型是h2h_low
            if not flag_h2h_low_no_preamble_or_no_user:  # 如果不是所有h2h_low用户都没有前导码或者没有用户
                if collisionDetectionMode == 1:
                    debug(
                        f"preambleListForH2hLow.numberOfPreamble is {preambleListForH2hLow.numberOfPreamble}"
                    )
                    if (
                        preambleListForH2hLow.numberOfPreamble == 0
                        or len(preambleListForH2hLow.preambleList) == 0
                    ):
                        flag_h2h_low_no_preamble_or_no_user = True
                        debug(
                            "Set flag_h2h_low_no_preamble_or_no_user to True due to no preamble is available"
                        )
                        continue
                    objPreamble = np.random.choice(preambleListForH2hLow.preambleList)
                    if objPreamble.assignedToUser:  # 如果前导码已被分配给用户
                        random_user.collisionDetected = 1
                        debug(f"Collision detected")
                        count_all_user_per_frame_collision += 1
                        count_h2h_per_frame_collision += 1
                        count_h2h_low_per_frame_collision += 1
                    elif not objPreamble.assignedToUser:  # 如果前导码未被分配给用户
                        # unused_preamble_exist = True
                        random_user.preamble = objPreamble.preamble  # 为用户分配前导码
                        random_user.served_frame = (
                            1 if np.random.randint(2, 8) <= 6 else 2
                        )  # 暂定5-10帧，未来可根据情况改
                        objPreamble.assignedToUser = True  # 标记前导码已被分配给用户
                        debug("Assigned preamble to H2H_LOW user")
                        preambleListForH2hLow.numOfAvailablePreamble -= 1
                        random_user.assigned_frame = frame_count
                        random_user.delay_frame = (
                            random_user.assigned_frame - random_user.joined_frame
                        )
                        debug(
                            f"random_user is {random_user.user_type} and random_user.assigned_frame is {random_user.assigned_frame} and random_user.joined_frame is {random_user.joined_frame} and random_user.delay_frame is {random_user.delay_frame} "
                        )

                        unserved_user_list.remove(random_user)
                        served_user_list.append(random_user)
                        total_served_user_list.append(random_user)  # 将用户添加到总已服务用户列表
                        unserved_h2h_user_low_list.remove(random_user)

                    count_collision_detected = 0
                    if len(unserved_h2h_user_low_list) == 0:  # 如果所有h2h_low用户都已服务
                        flag_h2h_low_no_preamble_or_no_user = True
                        debug(
                            "Set flag_h2h_low_no_preamble_or_no_user to True because there is no H2H low user left"
                        )
                    else:
                        for user in unserved_h2h_user_low_list:
                            if user.collisionDetected == 1:
                                count_collision_detected += 1

                        if count_collision_detected == len(unserved_h2h_user_low_list):
                            flag_h2h_low_no_preamble_or_no_user = True
                            debug(
                                "Set flag_h2h_low_no_preamble_or_no_user to True because all remain users are collision detected"
                            )

                else:
                    for (
                        objPreamble
                    ) in preambleListForH2hLow.preambleList:  # 遍历h2h_low前导码列表
                        if objPreamble.assignedToUser:  # 如果前导码已被分配给用户
                            continue
                        elif not objPreamble.assignedToUser:  # 如果前导码未被分配给用户
                            # unused_preamble_exist = True
                            random_user.preamble = objPreamble.preamble  # 为用户分配前导码
                            random_user.served_frame = (
                                1 if np.random.randint(2, 8) <= 6 else 2
                            )  # 暂定5-10帧，未来可根据情况改
                            objPreamble.assignedToUser = True  # 标记前导码已被分配给用户
                            debug("Assigned preamble to H2H_LOW user")
                            preambleListForH2hLow.numOfAvailablePreamble -= 1
                            random_user.assigned_frame = frame_count
                            random_user.delay_frame = (
                                random_user.assigned_frame - random_user.joined_frame
                            )
                            debug(
                                f"random_user is {random_user.user_type} and random_user.assigned_frame is {random_user.assigned_frame} and random_user.joined_frame is {random_user.joined_frame} and random_user.delay_frame is {random_user.delay_frame} "
                            )

                            unserved_user_list.remove(random_user)
                            served_user_list.append(random_user)
                            total_served_user_list.append(random_user)  # 将用户添加到总已服务用户列表
                            unserved_h2h_user_low_list.remove(random_user)
                            break

                    if len(unserved_h2h_user_low_list) == 0:  # 如果所有h2h_low用户都已服务
                        flag_h2h_low_no_preamble_or_no_user = True
                        debug(
                            "Set flag_h2h_low_no_preamble_or_no_user to True because there is no H2H low user left"
                        )

                if (
                    preambleListForH2hLow.numOfAvailablePreamble == 0
                ):  # 如果h2h_low前导码列表中没有未被分配的前导码
                    flag_h2h_low_no_preamble_or_no_user = True
                    debug(
                        "Set flag_h2h_low_no_preamble_or_no_user to True due to no preamble is available"
                    )

        elif random_user.user_type == M2M_HIGH:  # 如果用户类型是m2m_high
            if not flag_m2m_high_no_preamble_or_no_user:
                if collisionDetectionMode == 1:
                    debug(
                        f"preambleListForM2mHigh.numberOfPreamble is {preambleListForM2mHigh.numberOfPreamble}"
                    )
                    if (
                        preambleListForM2mHigh.numberOfPreamble == 0
                        or len(preambleListForM2mHigh.preambleList) == 0
                    ):
                        flag_m2m_high_no_preamble_or_no_user = True
                        debug(
                            "Set flag_m2m_high_no_preamble_or_no_user to True due to no preamble is available"
                        )
                        continue
                    objPreamble = np.random.choice(preambleListForM2mHigh.preambleList)
                    if objPreamble.assignedToUser:
                        random_user.collisionDetected = 1
                        debug(f"Collision detected")
                        count_all_user_per_frame_collision += 1
                        count_m2m_per_frame_collision += 1
                        count_m2m_high_per_frame_collision += 1
                    elif not objPreamble.assignedToUser:
                        # unused_preamble_exist = True
                        random_user.preamble = objPreamble.preamble
                        # random_user.served_frame = 1 if np.random.randint(2, 8) <= 6 else 2
                        # random_user.served_frame = np.random.randint(1, 2)
                        random_user.served_frame = 1
                        objPreamble.assignedToUser = True
                        debug("Assigned preamble to M2M_HIGH user")
                        preambleListForM2mHigh.numOfAvailablePreamble -= 1
                        random_user.assigned_frame = frame_count
                        random_user.delay_frame = (
                            random_user.assigned_frame - random_user.joined_frame
                        )
                        debug(
                            f"random_user is {random_user.user_type} and random_user.assigned_frame is {random_user.assigned_frame} and random_user.joined_frame is {random_user.joined_frame} and random_user.delay_frame is {random_user.delay_frame}"
                        )
                        unserved_user_list.remove(random_user)
                        served_user_list.append(random_user)
                        total_served_user_list.append(random_user)  # 将用户添加到总已服务用户列表
                        unserved_m2m_user_high_list.remove(random_user)

                    count_collision_detected = 0
                    if len(unserved_m2m_user_high_list) == 0:
                        flag_m2m_high_no_preamble_or_no_user = True
                        debug(
                            "Set flag_m2m_high_no_preamble_or_no_user to True because there is no M2M high user left"
                        )
                    else:
                        for user in unserved_m2m_user_high_list:
                            if user.collisionDetected == 1:
                                count_collision_detected += 1

                        if count_collision_detected == len(unserved_m2m_user_high_list):
                            flag_m2m_high_no_preamble_or_no_user = True
                            debug(
                                "Set flag_m2m_high_no_preamble_or_no_user to True because all remain users are collision detected"
                            )
                else:
                    for objPreamble in preambleListForM2mHigh.preambleList:
                        if objPreamble.assignedToUser:
                            continue

                        elif not objPreamble.assignedToUser:
                            # unused_preamble_exist = True
                            random_user.preamble = objPreamble.preamble
                            # random_user.served_frame = np.random.randint(1, 2)
                            # random_user.served_frame = 1 if np.random.randint(2, 8) <= 6 else 2

                            random_user.served_frame = 1
                            objPreamble.assignedToUser = True
                            debug("Assigned preamble to M2M_HIGH user")
                            preambleListForM2mHigh.numOfAvailablePreamble -= 1
                            random_user.assigned_frame = frame_count
                            random_user.delay_frame = (
                                random_user.assigned_frame - random_user.joined_frame
                            )
                            debug(
                                f"random_user is {random_user.user_type} and random_user.assigned_frame is {random_user.assigned_frame} and random_user.joined_frame is {random_user.joined_frame} and random_user.delay_frame is {random_user.delay_frame}"
                            )
                            unserved_user_list.remove(random_user)
                            # debug(f"{sys._getframe().f_lineno}")
                            served_user_list.append(random_user)
                            total_served_user_list.append(random_user)  # 将用户添加到总已服务用户列表
                            unserved_m2m_user_high_list.remove(random_user)
                            break

                    if len(unserved_m2m_user_high_list) == 0:
                        flag_m2m_high_no_preamble_or_no_user = True
                        debug(
                            "Set flag_m2m_high_no_preamble_or_no_user to True because there is no M2M high user left"
                        )

                if preambleListForM2mHigh.numOfAvailablePreamble == 0:
                    flag_m2m_high_no_preamble_or_no_user = True
                    debug(
                        "Set flag_m2m_high_no_preamble_or_no_user to True due to no preamble is available"
                    )

        elif random_user.user_type == M2M_MEDIUM:
            if not flag_m2m_medium_no_preamble_or_no_user:
                if collisionDetectionMode == 1:
                    debug(
                        f"PreambleListForM2mMedium.numberOfPreamble is {PreambleListForM2mMedium.numberOfPreamble}"
                    )
                    if (
                        PreambleListForM2mMedium.numberOfPreamble == 0
                        or len(PreambleListForM2mMedium.preambleList) == 0
                    ):
                        flag_m2m_medium_no_preamble_or_no_user = True
                        debug(
                            "Set flag_m2m_medium_no_preamble_or_no_user to True due to no preamble is available"
                        )
                        continue

                    objPreamble = np.random.choice(
                        PreambleListForM2mMedium.preambleList
                    )
                    if objPreamble.assignedToUser:
                        random_user.collisionDetected = 1
                        debug(f"Collision detected")
                        count_all_user_per_frame_collision += 1
                        count_m2m_per_frame_collision += 1
                        count_m2m_medium_per_frame_collision += 1
                    elif not objPreamble.assignedToUser:
                        # unused_preamble_exist = True
                        random_user.preamble = objPreamble.preamble
                        # random_user.served_frame = 1 if np.random.randint(2, 8) <= 6 else 2
                        # random_user.served_frame = np.random.randint(1, 2)
                        random_user.served_frame = 1
                        objPreamble.assignedToUser = True
                        debug("Assigned preamble to M2M_MEDIUM user")
                        PreambleListForM2mMedium.numOfAvailablePreamble -= 1
                        random_user.assigned_frame = frame_count
                        random_user.delay_frame = (
                            random_user.assigned_frame - random_user.joined_frame
                        )
                        debug(
                            f"random_user is {random_user.user_type} and random_user.assigned_frame is {random_user.assigned_frame} and random_user.joined_frame is {random_user.joined_frame} and random_user.delay_frame is {random_user.delay_frame}"
                        )
                        unserved_user_list.remove(random_user)
                        served_user_list.append(random_user)
                        total_served_user_list.append(random_user)  # 将用户添加到总已服务用户列表
                        unserved_m2m_user_medium_list.remove(random_user)

                    count_collision_detected = 0
                    if len(unserved_m2m_user_medium_list) == 0:
                        flag_m2m_medium_no_preamble_or_no_user = True
                        debug(
                            "Set flag_m2m_medium_no_preamble_or_no_user to True because there is no M2M medium user left"
                        )
                    else:
                        for user in unserved_m2m_user_medium_list:
                            if user.collisionDetected == 1:
                                count_collision_detected += 1

                        if count_collision_detected == len(
                            unserved_m2m_user_medium_list
                        ):
                            flag_m2m_medium_no_preamble_or_no_user = True
                            debug(
                                "Set flag_m2m_medium_no_preamble_or_no_user to True because all remain users are collision detected"
                            )
                else:
                    for objPreamble in PreambleListForM2mMedium.preambleList:
                        if objPreamble.assignedToUser:
                            continue
                        elif not objPreamble.assignedToUser:
                            unused_preamble_exist = True
                            random_user.preamble = objPreamble.preamble
                            # random_user.served_frame = 1 if np.random.randint(2, 8) <= 6 else 2
                            # random_user.served_frame = np.random.randint(1, 2)
                            random_user.served_frame = 1
                            objPreamble.assignedToUser = True
                            debug("Assigned preamble to M2M_MEDIUM user")
                            PreambleListForM2mMedium.numOfAvailablePreamble -= 1
                            random_user.assigned_frame = frame_count
                            random_user.delay_frame = (
                                random_user.assigned_frame - random_user.joined_frame
                            )
                            debug(
                                f"random_user is {random_user.user_type} and random_user.assigned_frame is {random_user.assigned_frame} and random_user.joined_frame is {random_user.joined_frame} and random_user.delay_frame is {random_user.delay_frame}"
                            )
                            unserved_user_list.remove(random_user)
                            # debug(f"{sys._getframe().f_lineno}")
                            served_user_list.append(random_user)
                            total_served_user_list.append(random_user)  # 将用户添加到总已服务用户列表
                            unserved_m2m_user_medium_list.remove(random_user)
                            break

                    if len(unserved_m2m_user_medium_list) == 0:
                        flag_m2m_medium_no_preamble_or_no_user = True
                        debug(
                            "Set flag_m2m_medium_no_preamble_or_no_user to True because there is no M2M medium user left"
                        )
                if PreambleListForM2mMedium.numOfAvailablePreamble == 0:
                    flag_m2m_medium_no_preamble_or_no_user = True
                    debug(
                        "Set flag_m2m_medium_no_preamble_or_no_user to True due to no preamble is available"
                    )

        elif random_user.user_type == M2M_LOW:  # 如果用户类型是m2m_low
            if not flag_m2m_low_no_preamble_or_no_user:
                if collisionDetectionMode == 1:
                    debug(
                        f"preambleListForM2mLow.numberOfPreamble is {preambleListForM2mLow.numberOfPreamble}"
                    )
                    if (
                        preambleListForM2mLow.numberOfPreamble == 0
                        or len(preambleListForM2mLow.preambleList) == 0
                    ):
                        flag_m2m_low_no_preamble_or_no_user = True
                        debug(
                            "Set flag_m2m_low_no_preamble_or_no_user to True due to no preamble is available"
                        )
                        continue
                    objPreamble = np.random.choice(preambleListForM2mLow.preambleList)
                    if objPreamble.assignedToUser:
                        random_user.collisionDetected = 1
                        debug(f"Collision detected")
                        count_all_user_per_frame_collision += 1
                        count_m2m_per_frame_collision += 1
                        count_m2m_low_per_frame_collision += 1
                    elif not objPreamble.assignedToUser:
                        # unused_preamble_exist = True
                        random_user.preamble = objPreamble.preamble
                        # random_user.served_frame = 1 if np.random.randint(2, 8) <= 6 else 2
                        # random_user.served_frame = np.random.randint(1, 2)
                        random_user.served_frame = 1
                        objPreamble.assignedToUser = True
                        debug("Assigned preamble to M2M_LOW user")
                        preambleListForM2mLow.numOfAvailablePreamble -= 1
                        random_user.assigned_frame = frame_count
                        random_user.delay_frame = (
                            random_user.assigned_frame - random_user.joined_frame
                        )
                        debug(
                            f"random_user is {random_user.user_type} and random_user.assigned_frame is {random_user.assigned_frame} and random_user.joined_frame is {random_user.joined_frame} and random_user.delay_frame is {random_user.delay_frame}"
                        )
                        unserved_user_list.remove(random_user)
                        debug(f"{sys._getframe().f_lineno}")
                        served_user_list.append(random_user)
                        total_served_user_list.append(random_user)  # 将用户添加到总已服务用户列表
                        unserved_m2m_user_low_list.remove(random_user)

                    count_collision_detected = 0
                    if len(unserved_m2m_user_low_list) == 0:
                        flag_m2m_low_no_preamble_or_no_user = True
                        debug(
                            "Set flag_m2m_low_no_preamble_or_no_user to True because there is no M2M low user left"
                        )
                    else:
                        for user in unserved_m2m_user_low_list:
                            if user.collisionDetected == 1:
                                count_collision_detected += 1

                        if count_collision_detected == len(unserved_m2m_user_low_list):
                            flag_m2m_low_no_preamble_or_no_user = True
                            debug(
                                "Set flag_m2m_low_no_preamble_or_no_user to True because all remain users are collision detected"
                            )
                else:
                    for objPreamble in preambleListForM2mLow.preambleList:
                        if objPreamble.assignedToUser:
                            continue
                        elif not objPreamble.assignedToUser:
                            # unused_preamble_exist = True
                            random_user.preamble = objPreamble.preamble

                            # random_user.served_frame = np.random.randint(1, 2)
                            random_user.served_frame = 1
                            objPreamble.assignedToUser = True
                            debug("Assigned preamble to M2M_LOW user")
                            preambleListForM2mLow.numOfAvailablePreamble -= 1
                            random_user.assigned_frame = frame_count
                            random_user.delay_frame = (
                                random_user.assigned_frame - random_user.joined_frame
                            )
                            debug(
                                f"random_user is {random_user.user_type} and random_user.assigned_frame is {random_user.assigned_frame} and random_user.joined_frame is {random_user.joined_frame} and random_user.delay_frame is {random_user.delay_frame}"
                            )
                            unserved_user_list.remove(random_user)
                            # debug(f"{sys._getframe().f_lineno}")
                            served_user_list.append(random_user)
                            total_served_user_list.append(random_user)  # 将用户添加到总已服务用户列表
                            unserved_m2m_user_low_list.remove(random_user)
                            break

                    if len(unserved_m2m_user_low_list) == 0:
                        flag_m2m_low_no_preamble_or_no_user = True
                        debug(
                            "Set flag_m2m_low_no_preamble_or_no_user to True because there is no M2M low user left"
                        )
                if preambleListForM2mLow.numOfAvailablePreamble == 0:
                    flag_m2m_low_no_preamble_or_no_user = True
                    debug(
                        "Set flag_m2m_low_no_preamble_or_no_user to True due to no preamble is available"
                    )

        debug(f"len(unserved_user_list) is {len(unserved_user_list)}")

    for user in unserved_user_list:
        user.unserved_frame += 1

    debug(
        f"Average all user per frame collision rate {count_all_user_per_frame_collision} / {num1_of_h2h_user_high_per_frame[frame_count-1]+num1_of_h2h_user_medium_per_frame[frame_count-1]+num1_of_h2h_user_low_per_frame[frame_count-1]+num1_of_m2m_user_high_per_frame[frame_count-1]+num1_of_m2m_user_medium_per_frame[frame_count-1]+num1_of_m2m_user_low_per_frame[frame_count-1]}"
    )

    if (
        num1_of_h2h_user_high_per_frame[frame_count - 1]
        + num1_of_h2h_user_medium_per_frame[frame_count - 1]
        + num1_of_h2h_user_low_per_frame[frame_count - 1]
        + num1_of_m2m_user_high_per_frame[frame_count - 1]
        + num1_of_m2m_user_medium_per_frame[frame_count - 1]
        + num1_of_m2m_user_low_per_frame[frame_count - 1]
    ) != 0:
        avgAllUserPerFrameCollisionRateList.append(
            count_all_user_per_frame_collision
            / (
                num1_of_h2h_user_high_per_frame[frame_count - 1]
                + num1_of_h2h_user_medium_per_frame[frame_count - 1]
                + num1_of_h2h_user_low_per_frame[frame_count - 1]
                + num1_of_m2m_user_high_per_frame[frame_count - 1]
                + num1_of_m2m_user_medium_per_frame[frame_count - 1]
                + num1_of_m2m_user_low_per_frame[frame_count - 1]
            )
        )
    else:
        avgAllUserPerFrameCollisionRateList.append(0.0)

    debug(
        f"Average H2H High per frame collision rate {count_h2h_high_per_frame_collision} / {num1_of_h2h_user_high_per_frame[frame_count-1]}"
    )
    if num1_of_h2h_user_high_per_frame[frame_count - 1] != 0:
        avgH2hHighPerFrameCollisionRateList.append(
            count_h2h_high_per_frame_collision
            / num1_of_h2h_user_high_per_frame[frame_count - 1]
        )
    else:
        avgH2hHighPerFrameCollisionRateList.append(0.0)

    debug(
        f"Average H2H Medium per frame collision rate {count_h2h_medium_per_frame_collision} / {num1_of_h2h_user_medium_per_frame[frame_count-1]}"
    )
    if num1_of_h2h_user_medium_per_frame[frame_count - 1] != 0:
        avgH2hMediumPerFrameCollisionRateList.append(
            count_h2h_medium_per_frame_collision
            / num1_of_h2h_user_medium_per_frame[frame_count - 1]
        )
    else:
        avgH2hMediumPerFrameCollisionRateList.append(0.0)

    debug(
        f"Average H2H Low per frame collision rate {count_h2h_low_per_frame_collision} / {num1_of_h2h_user_low_per_frame[frame_count-1]}"
    )
    if num1_of_h2h_user_low_per_frame[frame_count - 1] != 0:
        avgH2hLowPerFrameCollisionRateList.append(
            count_h2h_low_per_frame_collision
            / num1_of_h2h_user_low_per_frame[frame_count - 1]
        )
    else:
        avgH2hLowPerFrameCollisionRateList.append(0.0)

    debug(
        f"Average M2M High per frame collision rate {count_m2m_high_per_frame_collision} / {num1_of_m2m_user_high_per_frame[frame_count-1]}"
    )
    if num1_of_m2m_user_high_per_frame[frame_count - 1] != 0:
        avgM2mHighPerFrameCollisionRateList.append(
            count_m2m_high_per_frame_collision
            / num1_of_m2m_user_high_per_frame[frame_count - 1]
        )
    else:
        avgM2mHighPerFrameCollisionRateList.append(0.0)

    debug(
        f"Average M2M Medium per frame collision rate {count_m2m_medium_per_frame_collision} / {num1_of_m2m_user_medium_per_frame[frame_count-1]}"
    )
    if num1_of_m2m_user_medium_per_frame[frame_count - 1] != 0:
        avgM2mMediumPerFrameCollisionRateList.append(
            count_m2m_medium_per_frame_collision
            / num1_of_m2m_user_medium_per_frame[frame_count - 1]
        )
    else:
        avgM2mMediumPerFrameCollisionRateList.append(0.0)

    debug(
        f"Average M2M Low per frame collision rate {count_m2m_low_per_frame_collision} / {num1_of_m2m_user_low_per_frame[frame_count-1]}"
    )
    if num1_of_m2m_user_low_per_frame[frame_count - 1] != 0:
        avgM2mLowPerFrameCollisionRateList.append(
            count_m2m_low_per_frame_collision
            / num1_of_m2m_user_low_per_frame[frame_count - 1]
        )
    else:
        avgM2mLowPerFrameCollisionRateList.append(0.0)

    debug(
        f"Average H2H per frame collision rate {count_h2h_per_frame_collision} / {num1_of_m2m_user_high_per_frame[frame_count-1]+num1_of_h2h_user_medium_per_frame[frame_count-1]+num1_of_h2h_user_low_per_frame[frame_count-1]}"
    )
    if (
        num1_of_m2m_user_high_per_frame[frame_count - 1]
        + num1_of_h2h_user_medium_per_frame[frame_count - 1]
        + num1_of_h2h_user_low_per_frame[frame_count - 1]
    ) != 0:
        avgH2hPerFrameCollisionRateList.append(
            count_h2h_per_frame_collision
            / (
                num1_of_m2m_user_high_per_frame[frame_count - 1]
                + num1_of_h2h_user_medium_per_frame[frame_count - 1]
                + num1_of_h2h_user_low_per_frame[frame_count - 1]
            )
        )
    else:
        avgH2hPerFrameCollisionRateList.append(0.0)

    debug(
        f"Average M2M per frame collision rate {count_m2m_per_frame_collision} / {num1_of_m2m_user_high_per_frame[frame_count-1]+num1_of_m2m_user_medium_per_frame[frame_count-1]+num1_of_m2m_user_low_per_frame[frame_count-1]}"
    )
    if (
        num1_of_m2m_user_high_per_frame[frame_count - 1]
        + num1_of_m2m_user_medium_per_frame[frame_count - 1]
        + num1_of_m2m_user_low_per_frame[frame_count - 1]
    ) != 0:
        avgM2mPerFrameCollisionRateList.append(
            count_m2m_per_frame_collision
            / (
                num1_of_m2m_user_high_per_frame[frame_count - 1]
                + num1_of_m2m_user_medium_per_frame[frame_count - 1]
                + num1_of_m2m_user_low_per_frame[frame_count - 1]
            )
        )
    else:
        avgM2mPerFrameCollisionRateList.append(0.0)


def update_blocked_user_list():
    # Use while True to simulation do-while loop
    flag_breaked_from_for_loop = False
    while True:
        for unserved_user in unserved_user_list:  # 遍历未服务用户列表
            # debug(f"len(unserved_user_list) is {len(unserved_user_list)}")

            # 暂时不显示
            # debug(f"user type is {unserved_user.user_type}, joined frame is {unserved_user.joined_frame}, unserved frame is {unserved_user.unserved_frame}")

            # debug(f"len(served_user_list) is {len(served_user_list)}")
            if unserved_user.unserved_frame >= num_of_frame_for_calc:
                blocked_user_list.append(unserved_user)  # 将用户添加到被阻塞用户列表
                debug(f"len(blocked_user_list is {len(blocked_user_list)})")
                # debug(f"{inspect.currentframe().f_lineno}")
                # debug(unserved_user_list)   # 未服务用户列表  暂时不显示
                unserved_user_list.remove(unserved_user)  # 将用户从未服务用户列表中删除
                debug(f"len(unserved_user_list is {len(unserved_user_list)})")

                # 如果未服务用户的未服务帧数大于等于设定计算帧数且用户类型是h2h_high
                if unserved_user.user_type == H2H_HIGH:
                    # 将用户从未服务h2h_high用户列表中删除
                    unserved_h2h_user_high_list.remove(unserved_user)
                    debug(
                        f"len(unserved_h2h_user_high_list) is {len(unserved_h2h_user_high_list)}"
                    )
                    # debug(f"{inspect.currentframe().f_lineno}")
                elif unserved_user.user_type == H2H_MEDIUM:
                    unserved_h2h_user_medium_list.remove(unserved_user)
                    debug(
                        f"len(unserved_h2h_user_medium_list) is {len(unserved_h2h_user_medium_list)}"
                    )

                elif unserved_user.user_type == H2H_LOW:  # 如果用户类型是h2h_low
                    unserved_h2h_user_low_list.remove(unserved_user)
                    debug(
                        f"len(unserved_h2h_user_low_list) is {len(unserved_h2h_user_low_list)}"
                    )
                    # debug(f"{inspect.currentframe().f_lineno}")
                elif unserved_user.user_type == M2M_HIGH:  # 如果用户类型是m2m_high
                    unserved_m2m_user_high_list.remove(unserved_user)
                    debug(
                        f"len(unserved_m2m_user_high_list) is {len(unserved_m2m_user_high_list)}"
                    )
                    # debug(f"{inspect.currentframe().f_lineno}")

                elif unserved_user.user_type == M2M_MEDIUM:
                    unserved_m2m_user_medium_list.remove(unserved_user)
                    debug(
                        f"len(unserved_m2m_user_medium_list) is {len(unserved_m2m_user_medium_list)}"
                    )
                elif unserved_user.user_type == M2M_LOW:  # 如果用户类型是m2m_low
                    unserved_m2m_user_low_list.remove(unserved_user)
                    debug(
                        f"len(unserved_m2m_user_low_list) is {len(unserved_m2m_user_low_list)}"
                    )
                    # debug(f"{inspect.currentframe().f_lineno}")

                flag_breaked_from_for_loop = True
                break

        if flag_breaked_from_for_loop:
            flag_breaked_from_for_loop = False
            continue
        else:
            break


def user_preamble_assign_delay_calculation(frame_count: int):
    sum_all_user_preamble_assign_delay = 0
    sum_h2h_high_preamble_assign_delay = 0
    sum_h2h_medium_preamble_assign_delay = 0
    sum_h2h_low_preamble_assign_delay = 0
    sum_m2m_high_preamble_assign_delay = 0
    sum_m2m_medium_preamble_assign_delay = 0
    sum_m2m_low_preamble_assign_delay = 0
    sum_h2h_preamble_assign_delay = 0
    sum_m2m_preamble_assign_delay = 0

    count_all_user_preamble_assign = 0
    count_h2h_high_preamble_assign = 0
    count_h2h_medium_preamble_assign = 0
    count_h2h_low_preamble_assign = 0
    count_m2m_high_preamble_assign = 0
    count_m2m_medium_preamble_assign = 0
    count_m2m_low_preamble_assign = 0
    count_h2h_preamble_assign = 0
    count_m2m_preamble_assign = 0

    for served_user in total_served_user_list:
        # if (
        #         (frame_count - 2 * (num_of_frame_for_calc - 1))
        #         <= served_user.joined_frame
        #         <= (frame_count - num_of_frame_for_calc + 1)
        # ):
        if (
            (frame_count - num_of_frame_for_calc + 1)
            <= served_user.joined_frame
            <= frame_count
        ):  # 如果用户加入帧数在计算帧数范围内
            sum_all_user_preamble_assign_delay += served_user.delay_frame
            count_all_user_preamble_assign += 1  # 总用户分配preamble次数加1
            if served_user.user_type == H2H_HIGH:
                sum_h2h_preamble_assign_delay += served_user.delay_frame
                count_h2h_preamble_assign += 1
                sum_h2h_high_preamble_assign_delay += served_user.delay_frame
                count_h2h_high_preamble_assign += 1
            elif served_user.user_type == H2H_MEDIUM:
                sum_h2h_preamble_assign_delay += served_user.delay_frame
                count_h2h_preamble_assign += 1
                sum_h2h_medium_preamble_assign_delay += served_user.delay_frame
                count_h2h_medium_preamble_assign += 1
            elif served_user.user_type == H2H_LOW:
                sum_h2h_preamble_assign_delay += served_user.delay_frame
                count_h2h_preamble_assign += 1
                sum_h2h_low_preamble_assign_delay += served_user.delay_frame
                count_h2h_low_preamble_assign += 1
            elif served_user.user_type == M2M_HIGH:
                sum_m2m_high_preamble_assign_delay += served_user.delay_frame
                count_m2m_high_preamble_assign += 1
                sum_m2m_preamble_assign_delay += served_user.delay_frame
                count_m2m_preamble_assign += 1
            elif served_user.user_type == M2M_MEDIUM:
                sum_m2m_medium_preamble_assign_delay += served_user.delay_frame
                count_m2m_medium_preamble_assign += 1
                sum_m2m_preamble_assign_delay += served_user.delay_frame
                count_m2m_preamble_assign += 1
            elif served_user.user_type == M2M_LOW:
                sum_m2m_low_preamble_assign_delay += served_user.delay_frame
                count_m2m_low_preamble_assign += 1
                sum_m2m_preamble_assign_delay += served_user.delay_frame
                count_m2m_preamble_assign += 1

    debug(
        f"Average all user preamble assign delay {sum_all_user_preamble_assign_delay} / {num1_of_h2h_user_high_per_frame[frame_count-1]+num1_of_h2h_user_medium_per_frame[frame_count-1]+num1_of_h2h_user_low_per_frame[frame_count-1]+num1_of_m2m_user_high_per_frame[frame_count-1]+num1_of_m2m_user_medium_per_frame[frame_count-1]+num1_of_m2m_user_low_per_frame[frame_count-1]}"
    )
    if (
        num1_of_h2h_user_high_per_frame[frame_count - 1]
        + num1_of_h2h_user_medium_per_frame[frame_count - 1]
        + num1_of_h2h_user_low_per_frame[frame_count - 1]
        + num1_of_m2m_user_high_per_frame[frame_count - 1]
        + num1_of_m2m_user_medium_per_frame[frame_count - 1]
        + num1_of_m2m_user_low_per_frame[frame_count - 1]
    ) != 0:
        averageAllUserPreambleAssignDelayList.append(sum_all_user_preamble_assign_delay)
    else:
        averageAllUserPreambleAssignDelayList.append(0.0)

    debug(
        f"Average h2h high preamble assign delay {sum_h2h_high_preamble_assign_delay} / {num1_of_h2h_user_high_per_frame[frame_count-1]}"
    )
    if num1_of_h2h_user_high_per_frame[frame_count - 1] != 0:
        averageH2hHighPreambleAssignDelayList.append(sum_h2h_high_preamble_assign_delay)
    else:
        averageH2hHighPreambleAssignDelayList.append(0.0)

    debug(
        f"Average h2h medium preamble assign delay {sum_h2h_medium_preamble_assign_delay} / {num1_of_h2h_user_medium_per_frame[frame_count-1]}"
    )
    if num1_of_h2h_user_medium_per_frame[frame_count - 1] != 0:
        averageH2hMediumPreambleAssignDelayList.append(
            sum_h2h_medium_preamble_assign_delay
        )
    else:
        averageH2hMediumPreambleAssignDelayList.append(0.0)

    debug(
        f"Average h2h low preamble assign delay {sum_h2h_low_preamble_assign_delay} / {num1_of_h2h_user_low_per_frame[frame_count-1]}"
    )
    if num1_of_h2h_user_low_per_frame[frame_count - 1] != 0:
        averageH2hLowPreambleAssignDelayList.append(sum_h2h_low_preamble_assign_delay)
    else:
        averageH2hLowPreambleAssignDelayList.append(0.0)

    debug(
        f"Average m2m high preamble assign delay {sum_m2m_high_preamble_assign_delay} / {num1_of_m2m_user_high_per_frame[frame_count-1]}"
    )
    if num1_of_m2m_user_high_per_frame[frame_count - 1] != 0:
        averageM2mHighPreambleAssignDelayList.append(sum_m2m_high_preamble_assign_delay)
    else:
        averageM2mHighPreambleAssignDelayList.append(0.0)

    debug(
        f"Average m2m medium preamble assign delay {sum_m2m_medium_preamble_assign_delay} / {num1_of_m2m_user_medium_per_frame[frame_count-1]}"
    )
    if num1_of_m2m_user_medium_per_frame[frame_count - 1] != 0:
        averageM2mMediumPreambleAssignDelayList.append(
            sum_m2m_medium_preamble_assign_delay
        )
    else:
        averageM2mMediumPreambleAssignDelayList.append(0.0)

    debug(
        f"Average m2m low preamble assign delay {sum_m2m_low_preamble_assign_delay} / { num1_of_m2m_user_low_per_frame[frame_count-1]}"
    )
    if num1_of_m2m_user_low_per_frame[frame_count - 1] != 0:
        averageM2mLowPreambleAssignDelayList.append(sum_m2m_low_preamble_assign_delay)
    else:
        averageM2mLowPreambleAssignDelayList.append(0.0)

    debug(
        f"Average h2h preamble assign delay {sum_h2h_preamble_assign_delay} / {num1_of_h2h_user_high_per_frame[frame_count-1]+num1_of_h2h_user_medium_per_frame[frame_count-1]+num1_of_h2h_user_low_per_frame[frame_count-1]}"
    )
    if (
        num1_of_h2h_user_high_per_frame[frame_count - 1]
        + num1_of_h2h_user_medium_per_frame[frame_count - 1]
        + num1_of_h2h_user_low_per_frame[frame_count - 1]
    ) != 0:
        averageH2hPreambleAssignDelayList.append(sum_h2h_preamble_assign_delay)
    else:
        averageH2hPreambleAssignDelayList.append(0.0)

    debug(
        f"Average m2m preamble assign delay {sum_m2m_preamble_assign_delay} / {num1_of_m2m_user_high_per_frame[frame_count-1]+num1_of_m2m_user_medium_per_frame[frame_count-1]+num1_of_m2m_user_low_per_frame[frame_count-1]}"
    )
    if (
        num1_of_m2m_user_high_per_frame[frame_count - 1]
        + num1_of_m2m_user_medium_per_frame[frame_count - 1]
        + num1_of_m2m_user_low_per_frame[frame_count - 1]
    ) != 0:
        averageM2mPreambleAssignDelayList.append(sum_m2m_preamble_assign_delay)
    else:
        averageM2mPreambleAssignDelayList.append(0.0)


def user_block_rate_calculation(frame_count: int):
    num_of_h2h_user_high_blocked_for_calc = 0
    num_of_h2h_user_medium_blocked_for_calc = 0
    num_of_h2h_user_low_blocked_for_calc = 0
    num_of_m2m_user_high_blocked_for_calc = 0
    num_of_m2m_user_medium_blocked_for_calc = 0
    num_of_m2m_user_low_blocked_for_calc = 0

    num_of_H2H_user_blocked_for_calc = 0
    num_of_M2M_user_blocked_for_calc = 0

    for user in blocked_user_list:
        if (
            (frame_count - 2 * (num_of_frame_for_calc - 1))
            <= user.joined_frame
            <= (frame_count - num_of_frame_for_calc + 1)
        ):
            if user.user_type == H2H_HIGH:
                num_of_h2h_user_high_blocked_for_calc += 1
            elif user.user_type == H2H_MEDIUM:
                num_of_h2h_user_medium_blocked_for_calc += 1
            elif user.user_type == H2H_LOW:
                num_of_h2h_user_low_blocked_for_calc += 1
            elif user.user_type == M2M_HIGH:
                num_of_m2m_user_high_blocked_for_calc += 1
            elif user.user_type == M2M_MEDIUM:
                num_of_m2m_user_medium_blocked_for_calc += 1
            elif user.user_type == M2M_LOW:
                num_of_m2m_user_low_blocked_for_calc += 1

            elif user.user_type == H2H:
                num_of_H2H_user_blocked_for_calc += 1
            elif user.user_type == M2M:
                num_of_M2M_user_blocked_for_calc += 1

    if globalModeSelector == 1:
        debug(
            f"num_of_h2h_user_high_blocked_for_calc is {num_of_h2h_user_high_blocked_for_calc}"
        )
        block_probs["h2h_high"] = num_of_h2h_user_high_blocked_for_calc / (
            num_of_frame_for_calc * num_of_h2h_user_high_per_frame
        )
        debug(
            f"num_of_h2h_user_medium_blocked_for_calc is {num_of_h2h_user_medium_blocked_for_calc}"
        )
        block_probs["h2h_medium"] = num_of_h2h_user_medium_blocked_for_calc / (
            num_of_frame_for_calc * num_of_h2h_user_medium_per_frame
        )

        debug(
            f"num_of_h2h_user_low_blocked_for_calc is {num_of_h2h_user_low_blocked_for_calc}"
        )
        block_probs["h2h_low"] = num_of_h2h_user_low_blocked_for_calc / (
            num_of_frame_for_calc * num_of_h2h_user_low_per_frame
        )
        debug(
            f"num_of_m2m_user_high_blocked_for_calc is {num_of_m2m_user_high_blocked_for_calc}"
        )
        block_probs["m2m_high"] = num_of_m2m_user_high_blocked_for_calc / (
            num_of_frame_for_calc * num_of_m2m_user_high_per_frame
        )
        debug(
            f"num_of_m2m_user_medium_blocked_for_calc is {num_of_m2m_user_medium_blocked_for_calc}"
        )
        block_probs["m2m_medium"] = num_of_m2m_user_medium_blocked_for_calc / (
            num_of_frame_for_calc * num_of_m2m_user_medium_per_frame
        )
        debug(
            f"num_of_m2m_user_low_blocked_for_calc is {num_of_m2m_user_low_blocked_for_calc}"
        )
        block_probs["m2m_low"] = num_of_m2m_user_low_blocked_for_calc / (
            num_of_frame_for_calc * num_of_m2m_user_low_per_frame
        )

        num_of_H2H_user_blocked_for_calc = (
            num_of_h2h_user_high_blocked_for_calc
            + num_of_h2h_user_medium_blocked_for_calc
            + num_of_h2h_user_low_blocked_for_calc
        )
        debug(f"num_of_H2H_user_blocked_for_calc is {num_of_H2H_user_blocked_for_calc}")
        num_of_M2M_user_blocked_for_calc = (
            num_of_m2m_user_high_blocked_for_calc
            + num_of_m2m_user_medium_blocked_for_calc
            + num_of_m2m_user_low_blocked_for_calc
        )
        debug(
            f"num_of_M2M_user_blocked_for_calc is {num_of_M2M_user_blocked_for_calc }"
        )
        block_probs["H2H"] = num_of_H2H_user_blocked_for_calc / (
            num_of_frame_for_calc
            * (
                num_of_h2h_user_high_per_frame
                + num_of_h2h_user_medium_per_frame
                + num_of_h2h_user_low_per_frame
            )
        )
        block_probs["M2M"] = num_of_M2M_user_blocked_for_calc / (
            num_of_frame_for_calc
            * (
                num_of_m2m_user_high_per_frame
                + num_of_m2m_user_medium_per_frame
                + num_of_m2m_user_low_per_frame
            )
        )
        block_probs["ALL"] = (
            num_of_h2h_user_high_blocked_for_calc
            + num_of_h2h_user_medium_blocked_for_calc
            + num_of_h2h_user_low_blocked_for_calc
            + num_of_m2m_user_high_blocked_for_calc
            + num_of_m2m_user_medium_blocked_for_calc
            + num_of_m2m_user_low_blocked_for_calc
        ) / (
            (num_of_frame_for_calc * num_of_h2h_user_high_per_frame)
            + (num_of_frame_for_calc * num_of_h2h_user_medium_per_frame)
            + (num_of_frame_for_calc * num_of_h2h_user_low_per_frame)
            + num_of_frame_for_calc * num_of_m2m_user_high_per_frame
            + (num_of_frame_for_calc * num_of_m2m_user_medium_per_frame)
            + (num_of_frame_for_calc * num_of_m2m_user_low_per_frame)
        )

    elif globalModeSelector == 2:
        (
            sum_of_h2h_user_high,
            sum_of_h2h_user_medium,
            sum_of_h2h_user_low,
            sum_of_m2m_user_high,
            sum_of_m2m_user_medium,
            sum_of_m2m_user_low,
            sum_of_H2H_user,
            sum_of_M2M_user,
            sum_of_ALL_user,
        ) = (0, 0, 0, 0, 0, 0, 0, 0, 0)

        for user in listForAllUser:
            if (
                (frame_count - 2 * (num_of_frame_for_calc - 1))
                <= user.joined_frame
                <= (frame_count - num_of_frame_for_calc + 1)
            ):
                sum_of_ALL_user += 1

        for user in listForH2hUser:
            if (
                (frame_count - 2 * (num_of_frame_for_calc - 1))
                <= user.joined_frame
                <= (frame_count - num_of_frame_for_calc + 1)
            ):
                sum_of_H2H_user += 1

        for user in listForM2mUser:
            if (
                (frame_count - 2 * (num_of_frame_for_calc - 1))
                <= user.joined_frame
                <= (frame_count - num_of_frame_for_calc + 1)
            ):
                sum_of_M2M_user += 1

        for user in listForH2hUserHigh:
            if (
                (frame_count - 2 * (num_of_frame_for_calc - 1))
                <= user.joined_frame
                <= (frame_count - num_of_frame_for_calc + 1)
            ):
                sum_of_h2h_user_high += 1
        for user in listForH2hUserMedium:
            if (
                (frame_count - 2 * (num_of_frame_for_calc - 1))
                <= user.joined_frame
                <= (frame_count - num_of_frame_for_calc + 1)
            ):
                sum_of_h2h_user_medium += 1

        for user in listForH2hUserLow:
            if (
                (frame_count - 2 * (num_of_frame_for_calc - 1))
                <= user.joined_frame
                <= (frame_count - num_of_frame_for_calc + 1)
            ):
                sum_of_h2h_user_low += 1

        for user in listForM2mUserHigh:
            if (
                (frame_count - 2 * (num_of_frame_for_calc - 1))
                <= user.joined_frame
                <= (frame_count - num_of_frame_for_calc + 1)
            ):
                sum_of_m2m_user_high += 1

        for user in listForM2mUserMedium:
            if (
                (frame_count - 2 * (num_of_frame_for_calc - 1))
                <= user.joined_frame
                <= (frame_count - num_of_frame_for_calc + 1)
            ):
                sum_of_m2m_user_medium += 1

        for user in listForM2mUserLow:
            if (
                (frame_count - 2 * (num_of_frame_for_calc - 1))
                <= user.joined_frame
                <= (frame_count - num_of_frame_for_calc + 1)
            ):
                sum_of_m2m_user_low += 1

        debug(
            f"num_of_h2h_user_high_blocked_for_calc is {num_of_h2h_user_high_blocked_for_calc}"
        )
        debug(f"sum_of_h2h_user_high is {sum_of_h2h_user_high}")
        if sum_of_h2h_user_high != 0:
            block_probs["h2h_high"] = (
                num_of_h2h_user_high_blocked_for_calc / sum_of_h2h_user_high
            )
        debug(
            f"num_of_h2h_user_medium_blocked_for_calc is {num_of_h2h_user_medium_blocked_for_calc}"
        )
        debug(f"sum_of_h2h_user_medium is {sum_of_h2h_user_medium}")

        if sum_of_h2h_user_medium != 0:
            block_probs["h2h_medium"] = (
                num_of_h2h_user_medium_blocked_for_calc / sum_of_h2h_user_medium
            )
        debug(
            f"num_of_h2h_user_low_blocked_for_calc is {num_of_h2h_user_low_blocked_for_calc}"
        )
        debug(f"sum_of_h2h_user_low is {sum_of_h2h_user_low}")

        if sum_of_h2h_user_low != 0:
            block_probs["h2h_low"] = (
                num_of_h2h_user_low_blocked_for_calc / sum_of_h2h_user_low
            )
        debug(
            f"num_of_m2m_user_high_blocked_for_calc is {num_of_m2m_user_high_blocked_for_calc}"
        )
        debug(f"sum_of_m2m_user_high is {sum_of_m2m_user_high}")

        if sum_of_m2m_user_high != 0:
            block_probs["m2m_high"] = (
                num_of_m2m_user_high_blocked_for_calc / sum_of_m2m_user_high
            )
        debug(
            f"num_of_m2m_user_medium_blocked_for_calc is {num_of_m2m_user_medium_blocked_for_calc}"
        )
        debug(f"sum_of_m2m_user_medium is {sum_of_m2m_user_medium}")

        if sum_of_m2m_user_medium != 0:
            block_probs["m2m_medium"] = (
                num_of_m2m_user_medium_blocked_for_calc / sum_of_m2m_user_medium
            )
        debug(
            f"num_of_m2m_user_low_blocked_for_calc is {num_of_m2m_user_low_blocked_for_calc}"
        )
        debug(f"sum_of_m2m_user_low is {sum_of_m2m_user_low}")

        if sum_of_m2m_user_low != 0:
            block_probs["m2m_low"] = (
                num_of_m2m_user_low_blocked_for_calc / sum_of_m2m_user_low
            )
        num_of_M2M_user_blocked_for_calc = (
            num_of_m2m_user_low_blocked_for_calc
            + num_of_m2m_user_medium_blocked_for_calc
            + num_of_m2m_user_high_blocked_for_calc
        )
        debug(f"num_of_M2M_user_blocked_for_calc is {num_of_M2M_user_blocked_for_calc}")
        sum_of_M2M_user = (
            sum_of_m2m_user_low + sum_of_m2m_user_medium + sum_of_m2m_user_high
        )
        debug(f"sum_of_M2M_user is {sum_of_M2M_user}")
        if sum_of_M2M_user != 0:
            block_probs["M2M"] = num_of_M2M_user_blocked_for_calc / sum_of_M2M_user

        num_of_H2H_user_blocked_for_calc = (
            num_of_h2h_user_high_blocked_for_calc
            + num_of_h2h_user_medium_blocked_for_calc
            + num_of_h2h_user_low_blocked_for_calc
        )

        debug(f"num_of_H2H_user_blocked_for_calc is {num_of_H2H_user_blocked_for_calc}")
        sum_of_H2H_user = (
            sum_of_h2h_user_high + sum_of_h2h_user_medium + sum_of_h2h_user_low
        )
        debug(f"sum_of_H2H_user is {sum_of_H2H_user}")
        if sum_of_H2H_user != 0:
            block_probs["H2H"] = num_of_H2H_user_blocked_for_calc / sum_of_H2H_user
        sum_of_ALL_user = (
            sum_of_h2h_user_high
            + sum_of_h2h_user_medium
            + sum_of_h2h_user_low
            + sum_of_m2m_user_high
            + sum_of_m2m_user_medium
            + sum_of_m2m_user_low
        )
        debug(f"sum_of_ALL_user is {sum_of_ALL_user}")
        num_of_ALL_user_blocked_for_calc = (
            num_of_h2h_user_high_blocked_for_calc
            + num_of_h2h_user_medium_blocked_for_calc
            + num_of_h2h_user_low_blocked_for_calc
            + num_of_m2m_user_high_blocked_for_calc
            + num_of_m2m_user_medium_blocked_for_calc
            + num_of_m2m_user_low_blocked_for_calc
        )
        debug(f"num_of_ALL_user_blocked_for_calc is {num_of_ALL_user_blocked_for_calc}")
        if sum_of_ALL_user != 0:
            block_probs["ALL"] = num_of_ALL_user_blocked_for_calc / sum_of_ALL_user


def clean_blocked_user_list(frame_count: int):
    flag_breaked_from_for_loop = False
    while True:
        for user in blocked_user_list:
            if user.joined_frame < (frame_count - 2 * (num_of_frame_for_calc - 1)):
                debug(
                    f"user.user_type {user.user_type} user.joined_frame {user.joined_frame} removed in frame_count {frame_count}"
                )
                blocked_user_list.remove(user)
                flag_breaked_from_for_loop = True
                break

        if flag_breaked_from_for_loop:
            flag_breaked_from_for_loop = False
            continue
        else:
            break


def weighted_reward_calculation():
    # Weight for h2h user: 0.6 0.4
    weightedRateOfH2hUser = (
        user_priority_weights["h2h"]["high"] * block_probs["h2h_high"]
        + user_priority_weights["h2h"]["medium"] * block_probs["h2h_medium"]
        + user_priority_weights["h2h"]["low"] * block_probs["h2h_low"]
    )  # 计算h2h用户的加权阻塞率

    # Weight for m2m user: 0.6 0.4
    weightedRateOfM2mUser = (
        user_priority_weights["m2m"]["high"] * block_probs["m2m_high"]
        + user_priority_weights["m2m"]["medium"] * block_probs["m2m_medium"]
        + user_priority_weights["m2m"]["low"] * block_probs["m2m_low"]
    )  # 计算m2m用户的加权阻塞率

    # Weight for h2h vs m2m: 0.7 0.3
    weightedRateOfAllUser = (
        user_type_weights["h2h"] * weightedRateOfH2hUser
        + user_type_weights["m2m"] * weightedRateOfM2mUser
    )  # 计算所有用户的加权阻塞率

    if weightedRateOfAllUser > 0.5:  # 如果所有用户的加权阻塞率大于0.5
        reward = 0 - weightedRateOfAllUser - 1
    elif weightedRateOfAllUser > 0.2:
        reward = 0 - weightedRateOfAllUser - 0.5
    else:
        reward = 0 - weightedRateOfAllUser  # 如果所有用户的加权阻塞率小于等于0.2，奖励值为加权阻塞率

    return reward


# noinspection DuplicatedCode
def choose_action(curr_state, q_net):
    if np.random.uniform() > EPSILON:
        action = random.choice(action_feature)
        debug(f"random action is {action}")
    else:
        state_tensor = torch.tensor(curr_state).float().to(device)  #
        state_action_values = q_net(state_tensor)  # 获取当前状态对应的所有动作的Q值
        action = action_feature[state_action_values.argmax()]  # 选择Q值最大的动作
        debug(f"action is {action}")

    debug(f"action is {action}")
    if block_probs["h2h_high"] > 0.5:
        action[H2H_HIGH] += 3
    elif block_probs["h2h_high"] > 0.2:
        action[H2H_HIGH] += 1
    else:
        action[H2H_HIGH] -= 1
    if block_probs["h2h_medium"] > 0.5:
        action[H2H_MEDIUM] += 2
    elif block_probs["h2h_medium"] > 0.2:
        action[H2H_MEDIUM] += 1
    else:
        action[H2H_MEDIUM] -= 1

    if block_probs["h2h_low"] > 0.5:
        action[H2H_LOW] += 2
    elif block_probs["h2h_low"] > 0.2:
        action[H2H_LOW] += 1
    else:
        action[H2H_LOW] -= 1
    if block_probs["m2m_high"] > 0.5:
        action[M2M_HIGH] += 2
    elif block_probs["m2m_high"] > 0.2:
        action[M2M_HIGH] += 1
    else:
        action[M2M_HIGH] -= 1
    if block_probs["m2m_medium"] > 0.5:
        action[M2M_MEDIUM] += 1
    elif block_probs["m2m_medium"] > 0.2:
        action[M2M_MEDIUM] += 1
    else:
        action[M2M_MEDIUM] -= 1

    if block_probs["m2m_low"] > 0.5:
        action[M2M_LOW] += 1
    elif block_probs["m2m_low"] > 0.2:
        action[M2M_LOW] += 0
    else:
        action[M2M_LOW] -= 1

    return action


def choose_next_state(curr_state, action, penalty: float):
    # 在阻塞概率对于状态进行调整后得到新的当前状态，选取动作后，将当前状态和动作传入，得到下一状态
    # 对于当前的动作和状态进行显示,为了对下一状态进行计算
    debug(f"curr_state is {curr_state}, action is {action}")

    next_state = [0, 0, 0, 0, 0, 0]

    num_of_preamble_decrease_or_unchange = []  # 前导码数减少或者不变的用户类型对应的索引
    num_of_preamble_increase = []  # 前导码数增加的用户类型对应的索引

    for i in range(len(action)):
        if action[i] > 0:
            num_of_preamble_increase.append(i)
        else:
            num_of_preamble_decrease_or_unchange.append(i)

    state_sum = 0

    # 首先对前导码数减少或者不变的用户类型进行处理
    for index in num_of_preamble_decrease_or_unchange:
        # 前导码数不变的用户类型
        if action[index] == 0:
            next_state[index] = curr_state[index]
        # 前导码数减少的用户类型
        elif curr_state[index] + action[index] < NUM_PREAMBLE_LOW_LIMIT:
            next_state[index] = NUM_PREAMBLE_LOW_LIMIT
        elif curr_state[index] + action[index] >= NUM_PREAMBLE_LOW_LIMIT:
            next_state[index] = curr_state[index] + action[index]

        state_sum += next_state[index]

    # 然后对前导码数增加的用户类型进行处理
    for index in num_of_preamble_increase:
        state_sum += curr_state[index] + action[index]

    if state_sum > 54:
        # 超出范围
        audition_list = []
        for index in num_of_preamble_increase:
            for i in range(action[index]):
                audition_list.append(index)

        debug(f"audition_list is {audition_list}")
        # 假设当前50个，要增加10个，有6个是超出范围的，那么就从audition_list中随机选择6个，然后将这6个对应的action[index]减1
        for i in range(state_sum - 54):
            toDrop = np.random.choice(audition_list)
            action[toDrop] -= 1
            penalty -= 0.1
            audition_list.remove(toDrop)

        for index in num_of_preamble_increase:
            next_state[index] = curr_state[index] + action[index]
    else:
        # 未超出范围
        for index in num_of_preamble_increase:
            next_state[index] = curr_state[index] + action[index]

    debug(f"next_state is {next_state}")

    return next_state


convergence_data_for_plot = []  # 把收敛之前运行的所有帧数以及对应的收敛值存储到一个列表中
convergence_rate_blocked_data_for_plot = []
convergence_data_prev_hundred_frames = []  # 用于存储前100帧的收敛值
ALPHA_q_values_convergence_data_for_plot = []
Access_Success_Probility_test_frame_data_for_plot = []  # 用于存储接入成功率
Access_Success_Probility_frame_data_for_plot = []
Access_Success_Probility_frame_data_h2h_high_for_plot = []
Access_Success_Probility_frame_data_h2h_medium_for_plot = []
Access_Success_Probility_frame_data_h2h_low_for_plot = []
Access_Success_Probility_frame_data_m2m_high_for_plot = []
Access_Success_Probility_frame_data_m2m_medium_for_plot = []
Access_Success_Probility_frame_data_m2m_low_for_plot = []
Preamble_utilization_rate_test_frame_data_for_plot = []
Preamble_utilization_rate_frame_data_for_plot = []
Preamble_utilization_rate_average_data_for_plot = []
Preamble_utilization_rate_h2h_high_data_for_plot = []
Preamble_utilization_rate_h2h_medium_data_for_plot = []
Preamble_utilization_rate_h2h_low_data_for_plot = []
Preamble_utilization_rate_m2m_high_data_for_plot = []
Preamble_utilization_rate_m2m_medium_data_for_plot = []
Preamble_utilization_rate_m2m_low_data_for_plot = []
Preamble_utilization_rate_H2H_user_for_plot = []
Preamble_utilization_rate_M2M_user_for_plot = []
Access_Success_Probility_H2H_for_plot = []
Access_Success_Probility_M2M_for_plot = []
num_of_use_preamble_h2h_high_per_frame_list = []
num_of_use_preamble_h2h_medium_per_frame_list = []
num_of_use_preamble_h2h_low_per_frame_list = []
num_of_use_preamble_m2m_high_per_frame_list = []
num_of_use_preamble_m2m_medium_per_frame_list = []
num_of_use_preamble_m2m_low_per_frame_frame_list = []
new_user_list = []
loss_list_for_plot = []
total_q_value = 0
RMS_list = []
frame_rms_list = []
q_values_list = []


def check_convergence(
    q_predict, q_target, frame_count, total_q_value, q_values_list
):  # 判断收敛情况
    q_convergence_value = ALPHA * abs(q_predict - q_target)  # 计算q_predict和q_target的差值
    q_values = ALPHA * (q_target - q_predict)  # 计算q值
    total_q_value += q_values
    # print(f"q_values is {q_values}")
    q_values_list.append(q_values)
    debug(f"q_values_list is {q_values_list}")
    q_values_list = np.array(torch.tensor(q_values_list).cpu().numpy())
    Q_values = q_values_list.tolist()
    # print(f"Q_values is {Q_values}")
    np.savetxt("Q_values.txt", Q_values, fmt="%f", delimiter=",")

    convergence_data_for_plot.append(
        (frame_count, q_convergence_value)
    )  # 将帧数和收敛值分别存储到一个列表中
    ALPHA_q_values_convergence_data_for_plot.append(
        (frame_count, q_values)
    )  # 将帧数和q值分别存储到一个列表中
    debug(
        f"ALPHA_q_values_convergence_data_for_plot is {ALPHA_q_values_convergence_data_for_plot}"
    )
    convergence_data_prev_hundred_frames.append(
        q_convergence_value
    )  # 将收敛值存储到convergence_data_prev_hundred_frames列表中

    # print(f"convergence_data_prev_hundred_frames is {convergence_data_prev_hundred_frames}")   # 收敛之前的帧数对应的tensor值   是一个列表   暂时不显示
    debug(
        f"convergence_data_for_plot is {convergence_data_for_plot}"
    )  # 把收敛之前运行的所有帧数以及对应的收敛值存储到一个列表中   暂时不显示
    block_rate_list = [(k, v) for k, v in block_probs.items()]  # 把阻塞概率字典转换为列表
    # print(f"block_rate_list is {block_rate_list}")   #    暂时不显示  当前运行帧数对应的阻塞概率（[('m2m_low', 0.0), ('m2m_high', 0.0), ('h2h_low', 0.0), ('h2h_high', 0.0)]）  是一个列表

    convergence_rate_blocked_data_for_plot.append(
        (frame_count, block_rate_list)
    )  # 把收敛之前运行的所有帧数以及对应的阻塞概率 （现在这个阻塞概率是一个词典）存储到一个列表中
    # print(f"convergence_rate_blocked_data_for_plot is {convergence_rate_blocked_data_for_plot}")   #把收敛之前运行的所有帧数以及对应的阻塞概率 （现在这个阻塞概率是一个词典）存储到一个列表中   暂时不显示

    # np.append(convergence_data_prev_hundred_frames, q_convergence_value)
    if frame_count >= 100:  # 从第100帧开始检查收敛
        convergence_data_prev_hundred_frames.pop(0)  # 去掉首元素，为下一个元素留位置
        debug(convergence_data_prev_hundred_frames)
        # debug(convergence_data_prev_hundred_frames[0])
        # debug(convergence_data_prev_hundred_frames[0].item())
        list_for_check_convergence = []  # 用于存储convergence_data_prev_hundred_frames的值
        for i in range(
            len(convergence_data_prev_hundred_frames)
        ):  # 依次遍历convergence_data_prev_hundred_frames的每个元素
            list_for_check_convergence.append(
                convergence_data_prev_hundred_frames[i].item()
            )  # 将tensor转换为numpy，再转换为list
        # list_for_check_convergence = np.array(torch.tensor(convergence_data_prev_hundred_frames).cpu().numpy)
        debug(list_for_check_convergence)
        rms = np.sqrt(np.mean(np.square(list_for_check_convergence)))  # 计算前100帧的收敛值的均方根
        # print(f"rms is {rms}")
        frame_rms_list.append(frame_count)
        np.savetxt("RMS_list.csv", RMS_list, delimiter=",", fmt="%.2f")  # 保存均方根
        np.savetxt(
            "frame_rms_list.csv", frame_rms_list, delimiter=",", fmt="%.2f"
        )  # 保存帧数

        if rms < 0.01:
            debug(
                "!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!! Converged !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!"
            )
            debug(f"frame_count \n{frame_count}")
            # print(f"convergence_data_for_plot is {convergence_data_for_plot}")  # 收敛之前的帧数对应的tensor值   是一个列表  暂时不显示
            # 文件名，列表
            # np.savetxt("convergence_data_for_plot.csv", convergence_data_for_plot, delimiter=",")  # 保存为csv文件 文件名，列表名，分隔符，格式
            # np.savetxt("testResultArray.csv", testResultArray, delimiter=",", fmt="%.2f")
            return True


def double_convergence_check(
    q_predict, q_target, frame_count, total_q_value, q_values_list, loss
):  # 判断收敛情况
    q_convergence_value = ALPHA * abs(q_predict - q_target)  # 计算q_predict和q_target的差值
    q_values = ALPHA * (q_target - q_predict)  # 计算q值
    total_q_value += q_values
    # print(f"q_values is {q_values}")
    q_values_list.append(q_values)
    debug(f"q_values_list is {q_values_list}")
    q_values_list = np.array(torch.tensor(q_values_list).cpu().numpy())
    Q_values = q_values_list.tolist()
    # print(f"Q_values is {Q_values}")
    np.savetxt("Q_values.txt", Q_values, fmt="%f", delimiter=",")

    convergence_data_for_plot.append(
        (frame_count, q_convergence_value)
    )  # 将帧数和收敛值分别存储到一个列表中
    ALPHA_q_values_convergence_data_for_plot.append(
        (frame_count, q_values)
    )  # 将帧数和q值分别存储到一个列表中
    debug(
        f"ALPHA_q_values_convergence_data_for_plot is {ALPHA_q_values_convergence_data_for_plot}"
    )
    convergence_data_prev_hundred_frames.append(
        q_convergence_value
    )  # 将收敛值存储到convergence_data_prev_hundred_frames列表中

    if convergenceCheckMode == 2:
        listLossPrevHundredFrames.append(loss)

    # print(f"convergence_data_prev_hundred_frames is {convergence_data_prev_hundred_frames}")   # 收敛之前的帧数对应的tensor值   是一个列表   暂时不显示
    debug(
        f"convergence_data_for_plot is {convergence_data_for_plot}"
    )  # 把收敛之前运行的所有帧数以及对应的收敛值存储到一个列表中   暂时不显示
    block_rate_list = [(k, v) for k, v in block_probs.items()]  # 把阻塞概率字典转换为列表
    # print(f"block_rate_list is {block_rate_list}")   #    暂时不显示  当前运行帧数对应的阻塞概率（[('m2m_low', 0.0), ('m2m_high', 0.0), ('h2h_low', 0.0), ('h2h_high', 0.0)]）  是一个列表

    convergence_rate_blocked_data_for_plot.append(
        (frame_count, block_rate_list)
    )  # 把收敛之前运行的所有帧数以及对应的阻塞概率 （现在这个阻塞概率是一个词典）存储到一个列表中
    # print(f"convergence_rate_blocked_data_for_plot is {convergence_rate_blocked_data_for_plot}")   #把收敛之前运行的所有帧数以及对应的阻塞概率 （现在这个阻塞概率是一个词典）存储到一个列表中   暂时不显示

    # np.append(convergence_data_prev_hundred_frames, q_convergence_value)
    if frame_count >= 100:  # 从第100帧开始检查收敛
        convergence_data_prev_hundred_frames.pop(0)  # 去掉首元素，为下一个元素留位置
        debug(convergence_data_prev_hundred_frames)
        # debug(convergence_data_prev_hundred_frames[0])
        # debug(convergence_data_prev_hundred_frames[0].item())
        list_for_check_convergence = []  # 用于存储convergence_data_prev_hundred_frames的值
        for i in range(
            len(convergence_data_prev_hundred_frames)
        ):  # 依次遍历convergence_data_prev_hundred_frames的每个元素
            list_for_check_convergence.append(
                convergence_data_prev_hundred_frames[i].item()
            )  # 将tensor转换为numpy，再转换为list
        # list_for_check_convergence = np.array(torch.tensor(convergence_data_prev_hundred_frames).cpu().numpy)
        debug(list_for_check_convergence)
        rms = np.sqrt(np.mean(np.square(list_for_check_convergence)))  # 计算前100帧的收敛值的均方根
        print(f"rms is {rms}")
        frame_rms_list.append(frame_count)
        np.savetxt("RMS_list.csv", RMS_list, delimiter=",", fmt="%.2f")  # 保存均方根
        np.savetxt(
            "frame_rms_list.csv", frame_rms_list, delimiter=",", fmt="%.2f"
        )  # 保存帧数

        if convergenceCheckMode == 2:
            listLossPrevHundredFrames.pop(0)  # 去掉首元素，为下一个元素留位置
            debug(listLossPrevHundredFrames)
            listForCheckLossConvergence = []  # 用于存储listLossPrevHundredFrames的值
            for i in range(
                len(listLossPrevHundredFrames)
            ):  # 依次遍历listLossPrevHundredFrames的每个元素
                listForCheckLossConvergence.append(listLossPrevHundredFrames[i].item())
            debug(listForCheckLossConvergence)
            rms_loss = np.sqrt(
                np.mean(np.square(listForCheckLossConvergence))
            )  # 计算前100帧的loss的均方根
            print(f"rms_loss is {rms_loss}")
            if rms_loss >= 10000:
                return False

        if rms < 10:
            print(
                "!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!! Converged !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!"
            )
            debug(f"frame_count \n{frame_count}")
            # print(f"convergence_data_for_plot is {convergence_data_for_plot}")  # 收敛之前的帧数对应的tensor值   是一个列表  暂时不显示
            # 文件名，列表
            # np.savetxt("convergence_data_for_plot.csv", convergence_data_for_plot, delimiter=",")  # 保存为csv文件 文件名，列表名，分隔符，格式
            # np.savetxt("testResultArray.csv", testResultArray, delimiter=",", fmt="%.2f")
            return True


def plot_convergence():
    # 将帧数和收敛值分别存储到两个列表中
    frames_convergence = [
        data[0] for data in convergence_data_for_plot
    ]  # 收敛之前运行的帧数   是一个列表
    values = [data[1] for data in convergence_data_for_plot]  # 收敛之前tensor的值

    # print(f"frames \n{frames}")  # 收敛之前运行的帧数    作为x轴    暂时不显示
    debug(f"len(frames_convergence) \n{len(frames_convergence)}")
    # print(f"values \n{values}")  # 收敛之前tensor的值   暂时不显示
    # print(f"convergence_data_for_plot \n{convergence_data_for_plot}")  # 收敛之前的帧数对应的tensor值   暂时不显示
    debug("=====================================")
    values = np.array(
        torch.tensor(values).cpu().numpy()
    )  # KEVIN_20230802   表示将tensor转换为numpy，现在是一个收敛值的列表
    # print(f"values \n{values}")   #是一个数组   暂时不显示
    values_list = values.tolist()  # 将收敛值numpy数组转换为list   作为y轴
    debug(f"values_list \n{values_list}")  # 暂时不显示
    np.savetxt(
        "frames_convergence.csv", frames_convergence, delimiter=",", fmt="%.2f"
    )  # 保存帧数
    np.savetxt(
        "values_list.csv", values_list, delimiter=",", fmt="%.2f"
    )  # 保存收敛值  在当前学习率的条件下

    np.savetxt("values_list.csv", values_list, delimiter=",", fmt="%.2f")
    debug(f"len(values_list) \n{len(values_list)}")

    if loopModeSelector == 2:
        convergence_frame_list.extend(frames_convergence)
        convergence_value_list.extend(values_list)
        return

    plt.figure(figsize=(200, 10))
    r = range(len(frames_convergence))
    plt.bar(
        r, values_list, color="k", width=0.25, edgecolor="white"
    )  # color='#FF0088'  亮粉色
    plt.ylabel("convergence value")
    plt.xlabel("Frame count")
    plt.title("Convergence over time")
    # plt.xticks([i for i in range(len(frames))], frames)
    plt.xticks(
        [i for i in range(0, len(frames_convergence), 50)], frames_convergence[::50]
    )  # x轴刻度间隔为30

    # plt.xticks(range(0, len(values_list), 10))
    plt.ylim(0, max(values_list))
    # plt.ylim(0, 0.09)

    plt.show()


def plot_loss(loss_list_for_plot):
    if loopModeSelector == 2:
        return
    frames_loss = [data[0] for data in loss_list_for_plot]  # 收敛之前运行的帧数   是一个列表
    loss_values = [data[1] for data in loss_list_for_plot]  # 收敛之前tensor的值
    debug(f"loss_values is {loss_values}")
    plt.figure(figsize=(110, 10))
    x = frames_loss
    loss_values = np.array(torch.tensor(loss_values).cpu().numpy())
    loss_values_list = loss_values.tolist()

    print(f"loss_values_list is {loss_values_list}")
    np.savetxt(
        "loss_values_list.csv", loss_values_list, delimiter=",", fmt="%.2f"
    )  # 保存损失函数值
    np.savetxt("frames_loss.csv", frames_loss, delimiter=",", fmt="%.2f")  # 保存帧数

    y = loss_values_list
    plt.plot(x, y, color="k", linewidth=1.0, linestyle="-")
    plt.ylabel("loss value")
    plt.xlabel("Frame count")
    plt.title("Loss over time")
    plt.xticks(
        [i for i in range(0, len(frames_loss), 50)], frames_loss[::50]
    )  # x轴刻度间隔为30
    plt.ylim(0, max(loss_values_list))
    plt.show()


def plot_block_probs(frame_count, block_probs):  # 绘制收敛结束后，收敛之前，每帧运行的四种用户类型分别对应的阻塞概率图
    if loopModeSelector == 2:
        return
    # 将帧数和四种用户分别对应的阻塞概率分别存储到两个列表中
    frames = [
        data[0] for data in convergence_rate_blocked_data_for_plot
    ]  # 收敛之前运行的帧数   是一个列表
    rate_blocked_probs = [data[1] for data in convergence_rate_blocked_data_for_plot]

    # rate_blocked_probs_m2m_low = [data[0][1] for data in rate_blocked_probs]  # 获取m2m_low用户类型对应的阻塞概率(包括从输入帧数到收敛帧之前的所有m2m_low用户阻塞概率)
    #
    # rate_blocked_probs_m2m_high = [data[1][1] for data in rate_blocked_probs]
    # rate_blocked_probs_h2h_low = [data[2][1] for data in rate_blocked_probs]
    # rate_blocked_probs_h2h_high = [data[3][1] for data in rate_blocked_probs]  # 分别获取四种用户类型对应的阻塞概率

    rate_blocked_probs_h2h_high = [data[0][1] for data in rate_blocked_probs]
    rate_blocked_probs_h2h_medium = [data[1][1] for data in rate_blocked_probs]
    rate_blocked_probs_h2h_low = [data[2][1] for data in rate_blocked_probs]
    rate_blocked_probs_m2m_high = [data[3][1] for data in rate_blocked_probs]
    rate_blocked_probs_m2m_medium = [data[4][1] for data in rate_blocked_probs]
    rate_blocked_probs_m2m_low = [data[5][1] for data in rate_blocked_probs]
    rate_blocked_probs_H2H = [data[6][1] for data in rate_blocked_probs]
    rate_blocked_probs_M2M = [data[7][1] for data in rate_blocked_probs]
    rate_blocked_probs_ALL = [data[8][1] for data in rate_blocked_probs]

    debug(f"rate_blocked_probs_h2h_high is {rate_blocked_probs_h2h_high}")
    debug(f"rate_blocked_probs_h2h_medium is {rate_blocked_probs_h2h_medium}")
    debug(f"rate_blocked_probs_h2h_low is {rate_blocked_probs_h2h_low}")
    debug(f"rate_blocked_probs_m2m_high is {rate_blocked_probs_m2m_high}")
    debug(f"rate_blocked_probs_m2m_medium is {rate_blocked_probs_m2m_medium}")
    debug(f"rate_blocked_probs_m2m_low is {rate_blocked_probs_m2m_low}")  #
    debug(f"rate_blocked_probs_H2H is {rate_blocked_probs_H2H}")
    debug(f" rate_blocked_probs_M2M is {rate_blocked_probs_M2M}")
    debug(f"rate_blocked_probs_ALL is {rate_blocked_probs_ALL}")

    # print(f"convergence_rate_blocked_data_for_plot is {convergence_rate_blocked_data_for_plot}")  暂时不显示
    debug(f"{inspect.currentframe().f_lineno}")

    debug(f"frames \n{frames}")  # 收敛之前运行的帧数    作为x轴
    debug(f"len(frames) \n{len(frames)}")
    # print(f"rate_blocked_probs \n{rate_blocked_probs}")  # 收敛之前tensor的值   暂时不显示
    # print(f"convergence_rate_blocked_data_for_plot \n{convergence_rate_blocked_data_for_plot}")  #  暂时不显示 收敛之前的帧数对应的tensor值   [(1292, [('m2m_low', 0.0), ('m2m_high', 0.0), ('h2h_low', 0.6), ('h2h_high', 0.75)])]
    debug("=====================================")
    # rate_blocked_probs = np.array(torch.tensor( rate_blocked_probs ).cpu().numpy())  # KEVIN_20230802   表示将tensor转换为numpy，现在是一个收敛值的列表
    # rate_blocked_probs_list = rate_blocked_probs .tolist()  # 将收敛值numpy数组转换为list   作为y轴
    plt.figure(figsize=(110, 10))
    x = frames
    # print(f"x is {x}")   # x轴   收敛之前的帧数   暂时不显示
    y1 = rate_blocked_probs_h2h_high
    y2 = rate_blocked_probs_h2h_medium
    y3 = rate_blocked_probs_h2h_low
    y4 = rate_blocked_probs_m2m_high
    y5 = rate_blocked_probs_m2m_medium
    y6 = rate_blocked_probs_m2m_low
    y7 = rate_blocked_probs_H2H
    y8 = rate_blocked_probs_M2M

    plt.plot(
        x,
        y1,
        label="h2h_high",
        linewidth=1,
        color="r",
        markerfacecolor="red",
        markersize=3,
    )
    plt.plot(
        x,
        y2,
        label="h2h_medium",
        linewidth=1,
        color="g",
        markerfacecolor="green",
        markersize=3,
    )
    plt.plot(
        x,
        y3,
        label="h2h_low",
        linewidth=1,
        color="b",
        markerfacecolor="blue",
        markersize=3,
    )
    plt.plot(
        x,
        y4,
        label="m2m_high",
        linewidth=1,
        color="k",
        markerfacecolor="black",
        markersize=3,
    )
    plt.plot(
        x,
        y5,
        label="m2m_medium",
        linewidth=1,
        color="y",
        markerfacecolor="yellow",
        markersize=3,
    )
    plt.plot(
        x,
        y6,
        label="m2m_low",
        linewidth=1,
        color="c",
        markerfacecolor="cyan",
        markersize=3,
    )
    plt.plot(
        x,
        y7,
        label="H2H",
        linewidth=1,
        color="y",
        markerfacecolor="yellow",
        markersize=3,
    )
    plt.plot(
        x, y8, label="M2M", linewidth=1, color="r", markerfacecolor="red", markersize=3
    )

    plt.title("Block probability over time")
    plt.xlabel("Frame count")
    plt.ylabel("Block probability")
    plt.ylim(0, 1)
    plt.xticks([i for i in range(0, len(frames), 50)], frames[::50])
    plt.legend(prop={"size": 50}, bbox_to_anchor=(10, 10), loc="best")  #
    plt.legend(frameon=False, prop={"family": "simHei", "size": 20})
    plt.legend()
    plt.show()


def plot_test_block_probs(block_probs_list):
    frames = [data[0] for data in block_probs_list]
    block_rate_h2h_high = [data[1] for data in block_probs_list]
    block_rate_h2h_medium = [data[2] for data in block_probs_list]
    block_rate_h2h_low = [data[3] for data in block_probs_list]
    block_rate_m2m_high = [data[4] for data in block_probs_list]
    block_rate_m2m_medium = [data[5] for data in block_probs_list]
    block_rate_m2m_low = [data[6] for data in block_probs_list]
    block_rate_H2H = [data[7] for data in block_probs_list]
    block_rate_M2M = [data[8] for data in block_probs_list]
    block_rate_ALL = [data[9] for data in block_probs_list]

    if loopModeSelector == 2:
        block_rate_h2h_high_list.extend(block_rate_h2h_high)
        block_rate_h2h_medium_list.extend(block_rate_h2h_medium)
        block_rate_h2h_low_list.extend(block_rate_h2h_low)
        block_rate_m2m_high_list.extend(block_rate_m2m_high)
        block_rate_m2m_medium_list.extend(block_rate_m2m_medium)
        block_rate_m2m_low_list.extend(block_rate_m2m_low)
        block_rate_H2H_list.extend(block_rate_H2H)
        block_rate_M2M_list.extend(block_rate_M2M)
        block_rate_ALL_list.extend(block_rate_ALL)
        return

    debug(f"frames is {frames}")
    debug(f"block_rate_h2h_high is {block_rate_h2h_high}")
    debug(f"block_rate_h2h_medium is {block_rate_h2h_medium}")
    debug(f"block_rate_h2h_low is {block_rate_h2h_low}")
    debug(f"block_rate_m2m_high is {block_rate_m2m_high}")
    debug(f"block_rate_m2m_medium is {block_rate_m2m_medium}")
    debug(f"block_rate_m2m_low is {block_rate_m2m_low}")
    debug(f"block_rate_H2H is {block_rate_H2H}")
    debug(f"block_rate_M2M is {block_rate_M2M}")
    debug(f"block_rate_ALL is {block_rate_ALL}")
    debug(f"++++++++++++++++++++++++++++++++++++++++")
    Average_block_rate_h2h_high = sum(block_rate_h2h_high) / len(block_rate_h2h_high)
    Average_block_rate_h2h_medium = sum(block_rate_h2h_medium) / len(
        block_rate_h2h_medium
    )
    Average_block_rate_h2h_low = sum(block_rate_h2h_low) / len(block_rate_h2h_low)

    Average_block_rate_m2m_high = sum(block_rate_m2m_high) / len(block_rate_m2m_high)
    Average_block_rate_m2m_medium = sum(block_rate_m2m_medium) / len(
        block_rate_m2m_medium
    )
    Average_block_rate_m2m_low = sum(block_rate_m2m_low) / len(block_rate_m2m_low)
    Average_block_rate_H2H = sum(block_rate_H2H) / len(block_rate_H2H)
    Average_block_rate_M2M = sum(block_rate_M2M) / len(block_rate_M2M)
    Average_block_rate_ALL = sum(block_rate_ALL) / len(block_rate_ALL)
    debug(f"Average_block_rate_h2h_high is {Average_block_rate_h2h_high}")
    debug(f"Average_block_rate_h2h_medium is {Average_block_rate_h2h_medium}")
    debug(f"Average_block_rate_h2h_low is {Average_block_rate_h2h_low}")
    debug(f"Average_block_rate_m2m_high is {Average_block_rate_m2m_high}")
    debug(f"Average_block_rate_m2m_medium is {Average_block_rate_m2m_medium}")
    debug(f"Average_block_rate_m2m_low is {Average_block_rate_m2m_low}")
    debug(f"Average_block_rate_H2H is {Average_block_rate_H2H}")
    debug(f"Average_block_rate_M2M is {Average_block_rate_M2M}")
    debug(f"Average_block_rate_ALL is {Average_block_rate_ALL}")
    debug(f"++++++++++++++++++++++++++++++++++++++++")

    plt.figure(figsize=(110, 10))
    x = frames
    y1, y2, y3, y4, y5, y6, y7, y8 = (
        block_rate_h2h_high,
        block_rate_h2h_medium,
        block_rate_h2h_low,
        block_rate_m2m_high,
        block_rate_m2m_medium,
        block_rate_m2m_low,
        block_rate_H2H,
        block_rate_M2M,
    )
    plt.plot(
        x,
        y1,
        label="h2h_high",
        linewidth=1,
        color="r",
        markerfacecolor="red",
        markersize=3,
    )
    plt.plot(
        x,
        y2,
        label="h2h_medium",
        linewidth=1,
        color="g",
        markerfacecolor="green",
        markersize=3,
    )
    plt.plot(
        x,
        y3,
        label="h2h_low",
        linewidth=1,
        color="c",
        markerfacecolor="cyan",
        markersize=3,
    )
    plt.plot(
        x,
        y4,
        label="m2m_high",
        linewidth=1,
        color="b",
        markerfacecolor="blue",
        markersize=3,
    )
    plt.plot(
        x,
        y5,
        label="m2m_medium",
        linewidth=1,
        color="m",
        markerfacecolor="magenta",
        markersize=3,
    )
    plt.plot(
        x,
        y6,
        label="m2m_low",
        linewidth=1,
        color="k",
        markerfacecolor="black",
        markersize=3,
    )
    plt.plot(
        x,
        y7,
        label="H2H",
        linewidth=1,
        color="y",
        markerfacecolor="yellow",
        markersize=3,
    )
    plt.plot(
        x, y8, label="M2M", linewidth=1, color="r", markerfacecolor="red", markersize=3
    )
    plt.title("Block probability over time during test")
    plt.xlabel("Frame count")
    plt.ylabel("Block probability")
    plt.ylim(0, 1)
    plt.xticks([i for i in range(0, len(frames), 50)], frames[::50])
    plt.legend(prop={"size": 50}, bbox_to_anchor=(10, 10), loc="best")
    plt.legend(frameon=False, prop={"family": "Times New Roman", "size": 20})
    plt.legend()
    plt.show()


def plot_AHPHA_q_values_convergence(frame_count, total_q_value, q_values):
    if loopModeSelector == 2:
        return

    # 把收敛之前运行的所有帧数以及对应的收敛值存储到一个列表中
    frames_q_values = [data[0] for data in ALPHA_q_values_convergence_data_for_plot]
    Total_q_values = [data[1] for data in ALPHA_q_values_convergence_data_for_plot]
    plt.figure(figsize=(110, 10))
    x = frames_q_values
    Total_q_values = np.array(torch.tensor(Total_q_values).cpu().numpy())
    debug(f" Total_q_values is {Total_q_values}")
    Total_q_value_list = Total_q_values.tolist()
    debug(f"Total_q_value_list is {Total_q_value_list}")

    np.savetxt("Total_q_value_list.csv", Total_q_value_list, delimiter=",", fmt="%.2f")
    np.savetxt("frames_q_values.csv", frames_q_values, delimiter=",", fmt="%.2f")

    y1 = Total_q_value_list

    plt.plot(
        x,
        y1,
        label="ALPHA",
        linewidth=1,
        color="r",
        markerfacecolor="red",
        markersize=3,
    )
    plt.title("Q values over time")
    plt.xlabel("Frame count")
    plt.ylabel("Q values")
    # plt.ylim(0, 1)
    plt.xticks([i for i in range(0, len(frames_q_values), 50)], frames_q_values[::50])
    plt.legend(prop={"size": 50}, bbox_to_anchor=(10, 10), loc="best")
    plt.legend(frameon=False, prop={"family": "Times New Roman", "size": 20})
    plt.legend()
    plt.show()


def plt_Access_Success_Probility(
    test_frame,
    Access_Success_Probility_test_frame_data_for_plot,
    Access_Success_Probility_frame_data_for_plot,
    Access_Success_Probility_frame_data_h2h_high_for_plot,
    Access_Success_Probility_frame_data_h2h_medium_for_plot,
    Access_Success_Probility_frame_data_h2h_low_for_plot,
    Access_Success_Probility_frame_data_m2m_high_for_plot,
    Access_Success_Probility_frame_data_m2m_medium_for_plot,
    Access_Success_Probility_frame_data_m2m_low_for_plot,
):
    if loopModeSelector == 2:
        return

    debug(
        f"Access_Success_Probility_test_frame_data_for_plot is {Access_Success_Probility_test_frame_data_for_plot}"
    )
    debug(
        f"Access_Success_Probility_frame_data_for_plot is {Access_Success_Probility_frame_data_for_plot}"
    )

    debug(
        f"++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
    )

    plt.figure(figsize=(110, 10))
    x_data = Access_Success_Probility_test_frame_data_for_plot
    y_data = Access_Success_Probility_frame_data_for_plot
    y1 = Access_Success_Probility_frame_data_h2h_high_for_plot
    y2 = Access_Success_Probility_frame_data_h2h_medium_for_plot
    y3 = Access_Success_Probility_frame_data_h2h_low_for_plot
    y4 = Access_Success_Probility_frame_data_m2m_high_for_plot
    y5 = Access_Success_Probility_frame_data_m2m_medium_for_plot
    y6 = Access_Success_Probility_frame_data_m2m_low_for_plot
    plt.plot(
        x_data,
        y1,
        label="h2h_high",
        linewidth=1,
        color="r",
        markerfacecolor="red",
        markersize=3,
    )
    plt.plot(
        x_data,
        y2,
        label="h2h_medium",
        linewidth=1,
        color="g",
        markerfacecolor="green",
        markersize=3,
    )
    plt.plot(
        x_data,
        y3,
        label="h2h_low",
        linewidth=1,
        color="c",
        markerfacecolor="cyan",
        markersize=3,
    )
    plt.plot(
        x_data,
        y4,
        label="m2m_high",
        linewidth=1,
        color="b",
        markerfacecolor="blue",
        markersize=3,
    )
    plt.plot(
        x_data,
        y5,
        label="m2m_medium",
        linewidth=1,
        color="k",
        markerfacecolor="black",
        markersize=3,
    )
    plt.plot(
        x_data,
        y6,
        label="m2m_low",
        linewidth=1,
        color="m",
        markerfacecolor="magenta",
        markersize=3,
    )
    plt.plot(
        x_data,
        y_data,
        label="Access_Success_Probility",
        linewidth=1,
        color="y",
        markerfacecolor="yellow",
        markersize=3,
    )

    plt.title("Access_Success_Probility over time")
    plt.xlabel("Frame count")
    plt.ylabel("Access_Success_Probility")
    plt.ylim(0, 1)
    plt.xticks(
        [
            i
            for i in range(
                0, len(Access_Success_Probility_test_frame_data_for_plot), 10
            )
        ],
        Access_Success_Probility_test_frame_data_for_plot[::10],
    )
    plt.legend(prop={"size": 50}, bbox_to_anchor=(10, 10), loc="best")
    plt.legend(frameon=False, prop={"family": "Times New Roman", "size": 20})
    plt.legend()
    plt.show()


def plt_Preamble_utilization_rate(
    test_frame,
    Preamble_utilization_rate_test_frame_data_for_plot,
    Preamble_utilization_rate_h2h_high_data_for_plot,
    Preamble_utilization_rate_h2h_medium_data_for_plot,
    Preamble_utilization_rate_h2h_low_data_for_plot,
    Preamble_utilization_rate_m2m_high_data_for_plot,
    Preamble_utilization_rate_m2m_medium_data_for_plot,
    Preamble_utilization_rate_m2m_low_data_for_plot,
    Preamble_utilization_rate_H2H_user_for_plot,
    Preamble_utilization_rate_M2M_user_for_plot,
):
    if loopModeSelector == 2:
        return

    debug(
        f"Preamble_utilization_rate_test_frame_data_for_plot is {Preamble_utilization_rate_test_frame_data_for_plot}"
    )
    debug(
        f"Preamble_utilization_rate_frame_data_for_plot is {Preamble_utilization_rate_frame_data_for_plot}"
    )

    debug(
        f"Preamble_utilization_rate_h2h_high_data_for_plot is {Preamble_utilization_rate_h2h_high_data_for_plot}"
    )

    debug(
        f" Preamble_utilization_rate_h2h_medium_data_for_plot is {Preamble_utilization_rate_h2h_medium_data_for_plot}"
    )

    debug(
        f"Preamble_utilization_rate_h2h_low_data_for_plot is {Preamble_utilization_rate_h2h_low_data_for_plot}"
    )

    debug(
        f"Preamble_utilization_rate_m2m_high_data_for_plot is {Preamble_utilization_rate_m2m_high_data_for_plot}"
    )

    debug(
        f"Preamble_utilization_rate_m2m_medium_data_for_plot is {Preamble_utilization_rate_m2m_medium_data_for_plot}"
    )

    debug(
        f"Preamble_utilization_rate_m2m_low_data_for_plot is {Preamble_utilization_rate_m2m_low_data_for_plot}"
    )

    debug(
        f"Preamble_utilization_rate_H2H_user_for_plot is {Preamble_utilization_rate_H2H_user_for_plot}"
    )

    debug(
        f" Preamble_utilization_rate_M2M_user_for_plot is {Preamble_utilization_rate_M2M_user_for_plot}"
    )

    debug(
        f"++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
    )

    plt.figure(figsize=(110, 10))
    x_data = Preamble_utilization_rate_test_frame_data_for_plot
    y_data = Preamble_utilization_rate_frame_data_for_plot
    y1 = Preamble_utilization_rate_h2h_high_data_for_plot
    y2 = Preamble_utilization_rate_h2h_medium_data_for_plot
    y3 = Preamble_utilization_rate_h2h_low_data_for_plot
    y4 = Preamble_utilization_rate_m2m_high_data_for_plot
    y5 = Preamble_utilization_rate_m2m_medium_data_for_plot
    y6 = Preamble_utilization_rate_m2m_low_data_for_plot

    plt.plot(
        x_data,
        y_data,
        label="Preamble_utilization_rate",
        linewidth=1,
        color="r",
        markerfacecolor="red",
        markersize=3,
    )
    plt.plot(
        x_data,
        y1,
        label="h2h_high",
        linewidth=1,
        color="g",
        markerfacecolor="green",
        markersize=3,
    )
    plt.plot(
        x_data,
        y2,
        label="h2h_medium",
        linewidth=1,
        color="c",
        markerfacecolor="cyan",
        markersize=3,
    )
    plt.plot(
        x_data,
        y3,
        label="h2h_low",
        linewidth=1,
        color="b",
        markerfacecolor="blue",
        markersize=3,
    )
    plt.plot(
        x_data,
        y4,
        label="m2m_high",
        linewidth=1,
        color="k",
        markerfacecolor="black",
        markersize=3,
    )
    plt.plot(
        x_data,
        y5,
        label="m2m_medium",
        linewidth=1,
        color="m",
        markerfacecolor="magenta",
        markersize=3,
    )
    plt.plot(
        x_data,
        y6,
        label="m2m_low",
        linewidth=1,
        color="y",
        markerfacecolor="yellow",
        markersize=3,
    )
    plt.title("Preamble_utilization_rate over time")
    plt.xlabel("Frame count")
    plt.ylabel("Preamble_utilization_rate")
    plt.ylim(0, 1)
    plt.xticks(
        [
            i
            for i in range(
                0, len(Preamble_utilization_rate_test_frame_data_for_plot), 10
            )
        ],
        Preamble_utilization_rate_test_frame_data_for_plot[::10],
    )
    plt.legend(prop={"size": 50}, bbox_to_anchor=(10, 10), loc="best")
    plt.legend(frameon=False, prop={"family": "Times New Roman", "size": 20})
    plt.legend()
    plt.show()


def rl():
    debug("rl() start")
    q_net = QNet(N_STATE_FEATURES, N_HIDDEN, N_ACTIONS_FEATURES).to(device)
    optimizer = torch.optim.Adam(q_net.parameters(), lr=ALPHA)
    debug("++++++++++++++++++++++++++++++++++++++++++")
    print("******** It is time for model train ... ********")

    # 初始化当前状态  里面的元素是m2m_low,m2m_high,h2h_low,h2h_high   将初始状态[0,0,0,0]换成[8,19,8,19]，是为了方便测试
    # curr_state = state_feature[0]
    # curr_state = [20, 16, 12, 6]
    # curr_state = [1, 1, 1, 1, 1, 1]
    # curr_state = [3, 1, 1, 3, 1, 1]  # 10
    # curr_state = [4, 2, 1, 4, 2, 1]  # 14
    # curr_state = [5, 3, 2, 5, 3, 2]  # 20
    # curr_state = [6, 4, 2, 6, 4, 2]  # 24
    # curr_state = [8, 4, 3, 8, 4, 3]  # 30
    # curr_state = [9, 5, 3, 9, 5, 3]  # 34
    # curr_state = [10, 6, 4, 10, 6, 4]  # 40
    # curr_state = [11, 7, 4, 11, 7, 4]  # 前导码数量  44
    # curr_state = [13, 8, 4, 13, 8, 4]  # 50
    curr_state = [14, 8, 5, 14, 8, 5]  # 54

    debug(f"Initial curr_state is {curr_state}")
    num_preamble_h2h_high = curr_state[H2H_HIGH]  # 当前状态下h2h_high前导码个数
    num_preamble_h2h_medium = curr_state[H2H_MEDIUM]
    num_preamble_h2h_low = curr_state[H2H_LOW]  # 当前状态下h2h_low前导码个数
    num_preamble_m2m_high = curr_state[M2M_HIGH]  # 当前状态下m2m_high前导码个数
    num_preamble_m2m_medium = curr_state[M2M_MEDIUM]
    num_preamble_m2m_low = curr_state[M2M_LOW]  # 当前状态下m2m_low前导码个数

    debug(f"num_preamble_h2h_high is {num_preamble_h2h_high}")
    debug(f"num_preamble_h2h_medium is {num_preamble_h2h_medium}")
    debug(f"num_preamble_h2h_low is {num_preamble_h2h_low}")
    debug(f"num_preamble_m2m_high is {num_preamble_m2m_high}")
    debug(f"num_preamble_m2m_medium is {num_preamble_m2m_medium}")
    debug(f"num_preamble_m2m_low is {num_preamble_m2m_low}")

    frame_count = 0

    flag_initial_run = True  # 用于判断是否是第一次运行

    while True:
        frame_count += 1
        debug(
            f">>>>>>>>>>>>>>>> frame_count: {frame_count} >>>>>>>> curr_state: {curr_state} >>>>>>>>>>>>>>>>"
        )

        revoke_preamble_from_served_user()  # 从已服务用户列表中回收前导码
        debug(f"{sys._getframe().f_lineno}")
        update_preamble_lists(curr_state)  # 更新当前状态下的前导码列表
        debug(f"{sys._getframe().f_lineno}")
        add_new_user_and_preamble_in_every_frame(
            frame_count, curr_state
        )  # 每帧添加新用户并且添加前导码
        debug(f"{sys._getframe().f_lineno}")

        debug(
            "**************** After adding new user and assigning preambles ****************"
        )
        # 超过设定计算帧数来检查一下是否有用户没有被服务
        debug(f"len(unserved_user_list) is {len(unserved_user_list)}")
        debug(f"len(served_user_list) is {len(served_user_list)}")
        debug(f"len(blocked_user_list) is {len(blocked_user_list)}")
        debug("**************** Before updating blocked user list ****************")

        # 更新被阻塞用户列表
        update_blocked_user_list()
        debug("**************** After updating blocked user list ****************")
        debug(f"len(unserved_user_list) is {len(unserved_user_list)}")
        debug(f"len(served_user_list) is {len(served_user_list)}")
        debug(f"len(blocked_user_list) is {len(blocked_user_list)}")
        debug(f"----------------------------------------------")

        # 20230925 用于计算所有用户的平均前导码分配时延 Start
        user_preamble_assign_delay_calculation(frame_count)
        debug(
            f"averageAllUserPreambleAssignDelayList is {averageAllUserPreambleAssignDelayList}"
        )
        debug(
            f"averageH2hHighPreambleAssignDelayList is {averageH2hHighPreambleAssignDelayList}"
        )
        debug(
            f"averageH2hMediumPreambleAssignDelayList is {averageH2hMediumPreambleAssignDelayList}"
        )
        debug(
            f"averageH2hLowPreambleAssignDelayList is {averageH2hLowPreambleAssignDelayList}"
        )
        debug(
            f"averageM2mHighPreambleAssignDelayList is {averageM2mHighPreambleAssignDelayList}"
        )
        debug(
            f"averageM2mMediumPreambleAssignDelayList is {averageM2mMediumPreambleAssignDelayList}"
        )
        debug(
            f"averageM2mLowPreambleAssignDelayList is {averageM2mLowPreambleAssignDelayList}"
        )
        debug(
            f"averageH2hPreambleAssignDelayList is {averageH2hPreambleAssignDelayList}"
        )
        debug(
            f"averageM2mPreambleAssignDelayList is {averageM2mPreambleAssignDelayList}"
        )
        debug(f"+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        # 20230925 用于计算所有用户的平均前导码分配时延 End
        debug(
            f"avgAllUserPerFrameCollisionRateList is {avgAllUserPerFrameCollisionRateList}"
        )
        debug(
            f"avgH2hHighPerFrameCollisionRateList is {avgH2hHighPerFrameCollisionRateList}"
        )
        debug(
            f"avgH2hMediumPerFrameCollisionRateList is {avgH2hMediumPerFrameCollisionRateList}"
        )
        debug(
            f"avgH2hLowPerFrameCollisionRateList is {avgH2hLowPerFrameCollisionRateList}"
        )
        debug(
            f"avgM2mHighPerFrameCollisionRateList is {avgM2mHighPerFrameCollisionRateList}"
        )
        debug(
            f"avgM2mMediumPerFrameCollisionRateList is {avgM2mMediumPerFrameCollisionRateList}"
        )
        debug(
            f"avgM2mLowPerFrameCollisionRateList is {avgM2mLowPerFrameCollisionRateList}"
        )
        debug(f"avgH2hPerFrameCollisionRateList is {avgH2hPerFrameCollisionRateList}")
        debug(f"avgM2mPerFrameCollisionRateList is {avgM2mPerFrameCollisionRateList}")

        debug(f"+++++++++++++++++++++++++++++++++++++++++++++++++")
        debug(f"len(unserved_user_list) is {len(unserved_user_list)}")
        debug(f"len(unserved_h2h_user_high_list) is {len(unserved_h2h_user_high_list)}")
        debug(
            f"len(unserved_h2h_user_medium_list) is {len(unserved_h2h_user_medium_list)}"
        )
        debug(f"len(unserved_h2h_user_low_list) is {len(unserved_h2h_user_low_list)}")
        debug(f"len(unserved_m2m_user_high_list) is {len(unserved_m2m_user_high_list)}")
        debug(
            f"len(unserved_m2m_user_medium_list) is {len(unserved_m2m_user_medium_list)}"
        )
        debug(f"len(unserved_m2m_user_low_list) is {len(unserved_m2m_user_low_list)}")
        debug(f"----------------------------------------------")

        unserved_user_counts_per_frame_dict = {
            "unserved_user_list": len(unserved_user_list),
            "unserved_h2h_user_high_list": len(unserved_h2h_user_high_list),
            "unserved_h2h_user_medium_list": len(unserved_h2h_user_medium_list),
            "unserved_h2h_user_low_list": len(unserved_h2h_user_low_list),
            "unserved_m2m_user_high_list": len(unserved_m2m_user_high_list),
            "unserved_m2m_user_medium_list": len(unserved_m2m_user_medium_list),
            "unserved_m2m_user_low_list": len(unserved_m2m_user_low_list),
        }
        debug(
            f"unserved_user_counts_per_frame_dict is {unserved_user_counts_per_frame_dict}"
        )
        debug(f"+++++++++++++++++++++++++++++++++++++++++++++++")
        num_of_unserved_user_h2h_high_per_frame.append(len(unserved_h2h_user_high_list))
        num_of_unserved_user_h2h_medium_per_frame.append(
            len(unserved_h2h_user_medium_list)
        )
        num_of_unserved_user_h2h_low_per_frame.append(len(unserved_h2h_user_low_list))
        num_of_unserved_user_m2m_high_per_frame.append(len(unserved_m2m_user_high_list))
        num_of_unserved_user_m2m_medium_per_frame.append(
            len(unserved_m2m_user_medium_list)
        )
        num_of_unserved_user_m2m_low_per_frame.append(len(unserved_m2m_user_low_list))
        debug(
            f"num_of_unserved_user_h2h_high_per_frame is {num_of_unserved_user_h2h_high_per_frame}"
        )
        debug(
            f"num_of_unserved_user_h2h_medium_per_frame is {num_of_unserved_user_h2h_medium_per_frame}"
        )
        debug(
            f"num_of_unserved_user_h2h_low_per_frame is {num_of_unserved_user_h2h_low_per_frame}"
        )
        debug(
            f"num_of_unserved_user_m2m_high_per_frame is {num_of_unserved_user_m2m_high_per_frame}"
        )
        debug(
            f"num_of_unserved_user_m2m_medium_per_frame is {num_of_unserved_user_m2m_medium_per_frame}"
        )
        debug(
            f"num_of_unserved_user_m2m_low_per_frame is {num_of_unserved_user_m2m_low_per_frame}"
        )
        debug(f"************************************************")

        if flag_initial_run:
            # 如果不存在被阻塞用户，直接更新当前状态
            if len(blocked_user_list) == 0:
                action = random.choice(action_feature)
                debug(f"random action is {action}")
                next_state = choose_next_state(curr_state, action, 0.0)
                curr_state = next_state
                continue
            else:
                flag_initial_run = False

        # 计算当前状态下的阻塞概率
        user_block_rate_calculation(frame_count)
        debug(f"block_probs is {block_probs}")

        # 清理阻塞列表
        debug(f"len(blocked_user_list) is {len(blocked_user_list)} before clean")
        clean_blocked_user_list(frame_count)
        debug(f"len(blocked_user_list) is {len(blocked_user_list)} after clean")

        # 当阻塞率足够低时，减少该类用户的前导码数量
        if (
            block_probs["h2h_high"] <= 0.01
            and curr_state[H2H_HIGH] > NUM_PREAMBLE_LOW_LIMIT
        ):
            curr_state[H2H_HIGH] -= 1
        if (
            block_probs["h2h_medium"] <= 0.01
            and curr_state[H2H_MEDIUM] > NUM_PREAMBLE_LOW_LIMIT
        ):
            curr_state[H2H_MEDIUM] -= 1
        if (
            block_probs["h2h_low"] <= 0.01
            and curr_state[H2H_LOW] > NUM_PREAMBLE_LOW_LIMIT
        ):
            curr_state[H2H_LOW] -= 1
        if (
            block_probs["m2m_high"] <= 0.01
            and curr_state[M2M_HIGH] > NUM_PREAMBLE_LOW_LIMIT
        ):
            curr_state[M2M_HIGH] -= 1
        if (
            block_probs["m2m_medium"] <= 0.01
            and curr_state[M2M_MEDIUM] > NUM_PREAMBLE_LOW_LIMIT
        ):
            curr_state[M2M_MEDIUM] -= 1
        if (
            block_probs["m2m_low"] <= 0.01
            and curr_state[M2M_LOW] > NUM_PREAMBLE_LOW_LIMIT
        ):
            curr_state[M2M_LOW] -= 1

        reward = weighted_reward_calculation()

        debug(f"reward is {reward}")  # 奖励值

        action = choose_action(curr_state, q_net)  # 根据当前状态和q网络选择动作
        debug(f"action is {action}")

        # 20230925 新策略 Start，根据优先级平衡前导码数量
        if (
            len(unserved_h2h_user_high_list) >= len(unserved_m2m_user_high_list)
            and curr_state[H2H_HIGH] < curr_state[M2M_HIGH]
        ):
            action[M2M_HIGH] -= 1
            action[H2H_HIGH] += 1
        elif (
            len(unserved_m2m_user_high_list) >= len(unserved_h2h_user_medium_list)
            and curr_state[M2M_HIGH] < curr_state[H2H_MEDIUM]
        ):
            action[H2H_MEDIUM] -= 1
            action[M2M_HIGH] += 1
        elif (
            len(unserved_h2h_user_medium_list) >= len(unserved_m2m_user_medium_list)
            and curr_state[H2H_MEDIUM] < curr_state[M2M_MEDIUM]
        ):
            action[M2M_MEDIUM] -= 1
            action[H2H_MEDIUM] += 1
        elif (
            len(unserved_m2m_user_medium_list) >= len(unserved_h2h_user_low_list)
            and curr_state[M2M_MEDIUM] < curr_state[H2H_LOW]
        ):
            action[H2H_LOW] -= 1
            action[M2M_MEDIUM] += 1
        elif (
            len(unserved_h2h_user_low_list) >= len(unserved_m2m_user_low_list)
            and curr_state[H2H_LOW] < curr_state[M2M_LOW]
        ):
            action[M2M_LOW] -= 1
            action[H2H_LOW] += 1
        # 20230925 新策略 End，根据优先级平衡前导码数量

        # 20230925 新策略 Start，根据未使用前导码平衡未服务用户数
        unserved_user_length_list = [
            len(unserved_h2h_user_high_list),
            len(unserved_h2h_user_medium_list),
            len(unserved_h2h_user_low_list),
            len(unserved_m2m_user_high_list),
            len(unserved_m2m_user_medium_list),
            len(unserved_m2m_user_low_list),
        ]

        if len(unserved_h2h_user_high_list) > 0 and curr_state[H2H_HIGH] == 0:
            for i in range(0, len(unserved_user_length_list)):
                user_count = 0
                if unserved_user_length_list[i] == 0:
                    for user in served_user_list:
                        if user.user_type == i:
                            user_count += 1
                    if curr_state[i] > user_count:
                        curr_state[i] -= 1
                        curr_state[H2H_HIGH] += 1
        elif len(unserved_h2h_user_medium_list) > 0 and curr_state[H2H_MEDIUM] == 0:
            for i in range(0, len(unserved_user_length_list)):
                user_count = 0
                if unserved_user_length_list[i] == 0:
                    for user in served_user_list:
                        if user.user_type == i:
                            user_count += 1
                    if curr_state[i] > user_count:
                        curr_state[i] -= 1
                        curr_state[H2H_MEDIUM] += 1
        elif len(unserved_h2h_user_low_list) > 0 and curr_state[H2H_LOW] == 0:
            for i in range(0, len(unserved_user_length_list)):
                user_count = 0
                if unserved_user_length_list[i] == 0:
                    for user in served_user_list:
                        if user.user_type == i:
                            user_count += 1
                    if curr_state[i] > user_count:
                        curr_state[i] -= 1
                        curr_state[H2H_LOW] += 1
        elif len(unserved_m2m_user_high_list) > 0 and curr_state[M2M_HIGH] == 0:
            for i in range(0, len(unserved_user_length_list)):
                user_count = 0
                if unserved_user_length_list[i] == 0:
                    for user in served_user_list:
                        if user.user_type == i:
                            user_count += 1
                    if curr_state[i] > user_count:
                        curr_state[i] -= 1
                        curr_state[M2M_HIGH] += 1
        elif len(unserved_m2m_user_medium_list) > 0 and curr_state[M2M_MEDIUM] == 0:
            for i in range(0, len(unserved_user_length_list)):
                user_count = 0
                if unserved_user_length_list[i] == 0:
                    for user in served_user_list:
                        if user.user_type == i:
                            user_count += 1
                    if curr_state[i] > user_count:
                        curr_state[i] -= 1
                        curr_state[M2M_MEDIUM] += 1
        elif len(unserved_m2m_user_low_list) > 0 and curr_state[M2M_LOW] == 0:
            for i in range(0, len(unserved_user_length_list)):
                user_count = 0
                if unserved_user_length_list[i] == 0:
                    for user in served_user_list:
                        if user.user_type == i:
                            user_count += 1
                    if curr_state[i] > user_count:
                        curr_state[i] -= 1
                        curr_state[M2M_LOW] += 1
        # 20230925 新策略 End，根据未使用前导码平衡未服务用户数

        q_predict = q_net(torch.tensor(curr_state).float().to(device))[
            action_feature.index(action)
        ]
        debug(f"q_predict is {q_predict}")

        penalty1 = 0.0  # 当超出前导码范围时，给予惩罚值
        next_state = choose_next_state(curr_state, action, penalty1)  # 根据当前状态和动作选择下一个状态
        num_of_input_user = (
            num1_of_h2h_user_high_per_frame[frame_count - 1]
            + num1_of_h2h_user_medium_per_frame[frame_count - 1]
            + num1_of_h2h_user_low_per_frame[frame_count - 1]
            + num1_of_m2m_user_high_per_frame[frame_count - 1]
            + num1_of_m2m_user_medium_per_frame[frame_count - 1]
            + num1_of_m2m_user_low_per_frame[frame_count - 1]
        )
        penalty2 = 0.0
        # 惩罚值设置
        if len(unserved_user_list) / num_of_input_user > 0.8:
            penalty2 = -(len(unserved_user_list) / num_of_input_user) * 10
        elif 0.6 < len(unserved_user_list) / num_of_input_user <= 0.8:
            penalty2 = -(len(unserved_user_list) / num_of_input_user) * 10
        elif 0.4 < len(unserved_user_list) / num_of_input_user <= 0.6:
            penalty2 = -(len(unserved_user_list) / num_of_input_user) * 10
        elif 0.2 < len(unserved_user_list) / num_of_input_user <= 0.4:
            penalty2 = -(len(unserved_user_list) / num_of_input_user) * 10
        elif 0.1 < len(unserved_user_list) / num_of_input_user <= 0.2:
            penalty2 = -(len(unserved_user_list) / num_of_input_user) * 10

        reward += penalty1 + penalty2  # 对奖励进行更新
        q_target = (
            reward + GAMMA * q_net(torch.tensor(next_state).float().to(device)).max()
        )
        debug(f"q_target is {q_target}")

        q_values = ALPHA * (q_target - q_predict)
        debug(f"q_values is {q_values}")

        loss = F.mse_loss(q_predict, q_target)
        print(f"loss is {loss}")
        loss_list_for_plot.append((frame_count, loss))
        debug(f"loss_list_for_plot is {loss_list_for_plot}")
        optimizer.zero_grad()
        loss.backward()
        optimizer.step()

        # 状态转换
        curr_state = next_state
        if convergenceCheckMode == 1:
            is_training_over = check_convergence(
                q_predict, q_target, frame_count, total_q_value, q_values_list
            )
        elif convergenceCheckMode == 2:
            is_training_over = double_convergence_check(
                q_predict, q_target, frame_count, total_q_value, q_values_list, loss
            )
        if is_training_over:
            plot_convergence()
            plot_block_probs(
                frame_count, block_probs
            )  # 绘制收敛结束后，收敛之前，每帧运行的四种用户类型分别对应的阻塞概率图
            plot_AHPHA_q_values_convergence(frame_count, total_q_value, q_values)

            plot_loss(loss_list_for_plot)
            # debug(f"convergence_data_for_plot is \n{convergence_data_for_plot}")      #这两行暂时不显示
            # debug(f"convergence_data_prev_hundred_frames is \n{convergence_data_prev_hundred_frames}")

            # block_probs_list = model_test(q_net)

            break

    print("******** It is time for model test... ********")
    # curr_state = state_feature[0]
    # curr_state = [1, 1, 1, 1, 1, 1]
    # curr_state = [4, 2, 1, 4, 2, 1]  # 14
    curr_state = [14, 8, 5, 14, 8, 5]  # 54
    # curr_state = [11, 7, 4, 11, 7, 4]  # 改变前导码数量  44
    # curr_state = [9, 5, 3, 9, 5, 3]  # 34
    # curr_state = [6, 4, 2, 6, 4, 2]  # 24
    # curr_state = [3, 1, 1, 3, 1, 1]  # 10
    # curr_state = [5, 3, 2, 5, 3, 2]  # 20
    # curr_state = [8, 4, 3, 8, 4, 3]  # 30
    # curr_state = [10, 6, 4, 10, 6, 4]  # 40
    # curr_state = [13, 8, 4, 13, 8, 4]  # 50

    test_frame = 0
    block_probs_list = []

    # 清空列表
    preambleListForH2hHigh.clear()
    PreambleListForH2hMedium.clear()
    preambleListForH2hLow.clear()
    preambleListForM2mHigh.clear()
    PreambleListForM2mMedium.clear()
    preambleListForM2mLow.clear()

    unserved_user_list.clear()
    unserved_h2h_user_high_list.clear()
    unserved_h2h_user_low_list.clear()
    unserved_h2h_user_medium_list.clear()
    unserved_m2m_user_high_list.clear()
    unserved_m2m_user_medium_list.clear()
    unserved_m2m_user_low_list.clear()
    served_user_list.clear()

    blocked_user_list.clear()

    listForH2hUserMedium.clear()
    listForH2hUserHigh.clear()
    listForM2mUserLow.clear()
    listForM2mUserMedium.clear()
    listForM2mUserHigh.clear()
    listForH2hUserLow.clear()

    num1_of_h2h_user_high_per_frame.clear()
    num1_of_h2h_user_medium_per_frame.clear()
    num1_of_h2h_user_low_per_frame.clear()
    num1_of_m2m_user_high_per_frame.clear()
    num1_of_m2m_user_medium_per_frame.clear()
    num1_of_m2m_user_low_per_frame.clear()

    num_of_unserved_user_h2h_high_per_frame.clear()
    num_of_unserved_user_h2h_medium_per_frame.clear()
    num_of_unserved_user_h2h_low_per_frame.clear()
    num_of_unserved_user_m2m_high_per_frame.clear()
    num_of_unserved_user_m2m_medium_per_frame.clear()
    num_of_unserved_user_m2m_low_per_frame.clear()

    total_served_user_list.clear()
    averageAllUserPreambleAssignDelayList.clear()
    averageH2hHighPreambleAssignDelayList.clear()
    averageH2hMediumPreambleAssignDelayList.clear()
    averageH2hLowPreambleAssignDelayList.clear()
    averageM2mHighPreambleAssignDelayList.clear()
    averageM2mMediumPreambleAssignDelayList.clear()
    averageM2mLowPreambleAssignDelayList.clear()
    averageH2hPreambleAssignDelayList.clear()
    averageM2mPreambleAssignDelayList.clear()

    avgAllUserPerFrameCollisionRateList.clear()
    avgH2hHighPerFrameCollisionRateList.clear()
    avgH2hMediumPerFrameCollisionRateList.clear()
    avgH2hLowPerFrameCollisionRateList.clear()
    avgM2mHighPerFrameCollisionRateList.clear()
    avgM2mMediumPerFrameCollisionRateList.clear()
    avgM2mLowPerFrameCollisionRateList.clear()
    avgH2hPerFrameCollisionRateList.clear()
    avgM2mPerFrameCollisionRateList.clear()

    debug(f"len(unserved_user_list) is {len(unserved_user_list)}")
    debug(f"len(unserved_h2h_user_high_list) is {len(unserved_h2h_user_high_list)}")
    debug(f"len(unserved_h2h_user_medium_list) is {len(unserved_h2h_user_medium_list)}")
    debug(f"len(unserved_h2h_user_low_list) is {len(unserved_h2h_user_low_list)}")
    debug(f"len(unserved_m2m_user_high_list) is {len(unserved_m2m_user_high_list)}")
    debug(f"len(unserved_m2m_user_medium_list) is {len(unserved_m2m_user_medium_list)}")
    debug(f"len(unserved_m2m_user_low_list) is {len(unserved_m2m_user_low_list)}")
    debug(f"==========================================")

    # 初始化前导码池
    for preamble in preamble_pool:
        preamble.assignedToList = False
        preamble.assignedToUser = False

    flag_initial_run = True

    while test_frame < 100:
        test_frame += 1

        debug(
            f"*************************test_frame******************************* is {test_frame} and curr_state is {curr_state}"
        )

        debug(f"+++++++++++++++++++++++++++++++++++++++++++++++++++++++")

        revoke_preamble_from_served_user()
        debug(f"{sys._getframe().f_lineno}")
        update_preamble_lists(curr_state)
        debug(f"{sys._getframe().f_lineno}")

        # add_new_user_and_preamble_in_every_frame(test_frame)
        add_new_user_and_preamble_in_every_frame(test_frame, curr_state)
        debug(f"___________________________________________")
        # 20230925 用于计算所有用户的平均前导码分配时延 End
        debug(
            f"avgAllUserPerFrameCollisionRateList is {avgAllUserPerFrameCollisionRateList}"
        )
        debug(
            f"avgH2hHighPerFrameCollisionRateList is {avgH2hHighPerFrameCollisionRateList}"
        )
        debug(
            f"avgH2hMediumPerFrameCollisionRateList is {avgH2hMediumPerFrameCollisionRateList}"
        )
        debug(
            f"avgH2hLowPerFrameCollisionRateList is {avgH2hLowPerFrameCollisionRateList}"
        )
        debug(
            f"avgM2mHighPerFrameCollisionRateList is {avgM2mHighPerFrameCollisionRateList}"
        )
        debug(
            f"avgM2mMediumPerFrameCollisionRateList is {avgM2mMediumPerFrameCollisionRateList}"
        )
        debug(
            f"avgM2mLowPerFrameCollisionRateList is {avgM2mLowPerFrameCollisionRateList}"
        )
        debug(f"avgH2hPerFrameCollisionRateList is {avgH2hPerFrameCollisionRateList}")
        debug(f"avgM2mPerFrameCollisionRateList is {avgM2mPerFrameCollisionRateList}")
        Average_All_collision_rate = sum(avgAllUserPerFrameCollisionRateList) / len(
            avgAllUserPerFrameCollisionRateList
        )
        Average_h2h_high_collision_rate = sum(
            avgH2hHighPerFrameCollisionRateList
        ) / len(avgH2hHighPerFrameCollisionRateList)
        Average_h2h_medium_collision_rate = sum(
            avgH2hMediumPerFrameCollisionRateList
        ) / len(avgH2hMediumPerFrameCollisionRateList)
        Average_h2h_low_collision_rate = sum(avgH2hLowPerFrameCollisionRateList) / len(
            avgH2hLowPerFrameCollisionRateList
        )
        Average_m2m_high_collision_rate = sum(
            avgM2mHighPerFrameCollisionRateList
        ) / len(avgM2mHighPerFrameCollisionRateList)
        Average_m2m_medium_collision_rate = sum(
            avgM2mMediumPerFrameCollisionRateList
        ) / len(avgM2mMediumPerFrameCollisionRateList)
        Average_m2m_low_collision_rate = sum(avgM2mLowPerFrameCollisionRateList) / len(
            avgM2mLowPerFrameCollisionRateList
        )
        Average_h2h_collision_rate = sum(avgH2hPerFrameCollisionRateList) / len(
            avgH2hPerFrameCollisionRateList
        )
        Average_m2m_collision_rate = sum(avgM2mPerFrameCollisionRateList) / len(
            avgM2mPerFrameCollisionRateList
        )
        debug(f"Average_All_collision_rate is {Average_All_collision_rate}")
        debug(f"Average_h2h_high_collision_rate is {Average_h2h_high_collision_rate}")
        debug(
            f"Average_h2h_medium_collision_rate is {Average_h2h_medium_collision_rate}"
        )
        debug(f"Average_h2h_low_collision_rate is {Average_h2h_low_collision_rate}")
        debug(f"Average_m2m_high_collision_rate is {Average_m2m_high_collision_rate}")
        debug(
            f"Average_m2m_medium_collision_rate is {Average_m2m_medium_collision_rate}"
        )
        debug(f"Average_m2m_low_collision_rate is {Average_m2m_low_collision_rate}")
        debug(f"Average_h2h_collision_rate is {Average_h2h_collision_rate}")
        debug(f"Average_m2m_collision_rate is {Average_m2m_collision_rate}")

        debug(f"++++++++++++++++++++++++++++++++++++++++++++++")

        debug(f"{sys._getframe().f_lineno}")
        debug(f"len(unserved_user_list) is {len(unserved_user_list)}")
        debug(f"len(unserved_h2h_user_high_list) is {len(unserved_h2h_user_high_list)}")
        debug(
            f"len(unserved_h2h_user_medium_list) is {len(unserved_h2h_user_medium_list)}"
        )
        debug(f"len(unserved_h2h_user_low_list) is {len(unserved_h2h_user_low_list)}")
        debug(f"len(unserved_m2m_user_high_list) is {len(unserved_m2m_user_high_list)}")
        debug(
            f"len(unserved_m2m_user_medium_list) is {len(unserved_m2m_user_medium_list)}"
        )
        debug(f"len(unserved_m2m_user_low_list) is {len(unserved_m2m_user_low_list)}")
        debug(f"len(blocked_user_list) is {len(blocked_user_list)}")
        debug(f"___________________________________________________________________")

        unserved_user_counts_per_test_frame_dict = {
            "unserved_user_list": len(unserved_user_list),
            "unserved_h2h_user_high_list": len(unserved_h2h_user_high_list),
            "unserved_h2h_user_medium_list": len(unserved_h2h_user_medium_list),
            "unserved_h2h_user_low_list": len(unserved_h2h_user_low_list),
            "unserved_m2m_user_high_list": len(unserved_m2m_user_high_list),
            "unserved_m2m_user_medium_list": len(unserved_m2m_user_medium_list),
            "unserved_m2m_user_low_list": len(unserved_m2m_user_low_list),
        }
        debug(
            f"unserved_user_counts_per_test_frame_dict is {unserved_user_counts_per_test_frame_dict}"
        )
        debug(f"___________________________________________________________________")

        num_of_unserved_user_h2h_high_per_frame.append(len(unserved_h2h_user_high_list))
        num_of_unserved_user_h2h_medium_per_frame.append(
            len(unserved_h2h_user_medium_list)
        )
        num_of_unserved_user_h2h_low_per_frame.append(len(unserved_h2h_user_low_list))
        num_of_unserved_user_m2m_high_per_frame.append(len(unserved_m2m_user_high_list))
        num_of_unserved_user_m2m_medium_per_frame.append(
            len(unserved_m2m_user_medium_list)
        )
        num_of_unserved_user_m2m_low_per_frame.append(len(unserved_m2m_user_low_list))
        debug(
            f"num_of_unserved_user_h2h_high_per_frame is {num_of_unserved_user_h2h_high_per_frame}"
        )
        debug(
            f"num_of_unserved_user_h2h_medium_per_frame is {num_of_unserved_user_h2h_medium_per_frame}"
        )
        debug(
            f"num_of_unserved_user_h2h_low_per_frame is {num_of_unserved_user_h2h_low_per_frame}"
        )
        debug(
            f"num_of_unserved_user_m2m_high_per_frame is {num_of_unserved_user_m2m_high_per_frame}"
        )
        debug(
            f"num_of_unserved_user_m2m_medium_per_frame is {num_of_unserved_user_m2m_medium_per_frame}"
        )
        debug(
            f"num_of_unserved_user_m2m_low_per_frame is {num_of_unserved_user_m2m_low_per_frame}"
        )

        debug(f"++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")

        unserved_user_type_count = Counter(
            user.user_type for user in unserved_user_list
        )
        debug(
            f"unserved_user_type_count is {unserved_user_type_count}"
        )  # 四种用户未被服务的用户数量
        debug(
            f"unserved_user_type_count[H2H_HIGH] is {unserved_user_type_count[H2H_HIGH]}"
        )
        debug(
            f"unserved_user_type_count[H2H_MEDIUM] is {unserved_user_type_count[H2H_MEDIUM]}"
        )
        debug(
            f"unserved_user_type_count[H2H_LOW] is {unserved_user_type_count[H2H_LOW]}"
        )
        debug(
            f"unserved_user_type_count[M2M_HIGH] is {unserved_user_type_count[M2M_HIGH]}"
        )
        debug(
            f"unserved_user_type_count[M2M_MEDIUM] is {unserved_user_type_count[M2M_MEDIUM]}"
        )
        debug(
            f"unserved_user_type_count[M2M_LOW] is {unserved_user_type_count[M2M_LOW]}"
        )
        debug(f"+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        num_of_use_preamble_h2h_high_per_frame = (
            num1_of_h2h_user_high_per_frame[test_frame - 1]
            - num_of_unserved_user_h2h_high_per_frame[test_frame - 1]
        )
        num_of_use_preamble_h2h_medium_per_frame = (
            num1_of_h2h_user_medium_per_frame[test_frame - 1]
            - num_of_unserved_user_h2h_medium_per_frame[test_frame - 1]
        )
        num_of_use_preamble_h2h_low_per_frame = (
            num1_of_h2h_user_low_per_frame[test_frame - 1]
            - num_of_unserved_user_h2h_low_per_frame[test_frame - 1]
        )
        num_of_use_preamble_m2m_high_per_frame = (
            num1_of_m2m_user_high_per_frame[test_frame - 1]
            - num_of_unserved_user_m2m_high_per_frame[test_frame - 1]
        )
        num_of_use_preamble_m2m_medium_per_frame = (
            num1_of_m2m_user_medium_per_frame[test_frame - 1]
            - num_of_unserved_user_m2m_medium_per_frame[test_frame - 1]
        )
        num_of_use_preamble_m2m_low_per_frame = (
            num1_of_m2m_user_low_per_frame[test_frame - 1]
            - num_of_unserved_user_m2m_low_per_frame[test_frame - 1]
        )
        num_of_use_preamble_h2h_high_per_frame_list.append(
            num_of_use_preamble_h2h_high_per_frame
        )
        num_of_use_preamble_h2h_medium_per_frame_list.append(
            num_of_use_preamble_h2h_medium_per_frame
        )
        num_of_use_preamble_h2h_low_per_frame_list.append(
            num_of_use_preamble_h2h_low_per_frame
        )
        num_of_use_preamble_m2m_high_per_frame_list.append(
            num_of_use_preamble_m2m_high_per_frame
        )
        num_of_use_preamble_m2m_medium_per_frame_list.append(
            num_of_use_preamble_m2m_medium_per_frame
        )
        num_of_use_preamble_m2m_low_per_frame_frame_list.append(
            num_of_use_preamble_m2m_low_per_frame
        )
        debug(
            f"num_of_use_preamble_h2h_high_per_frame_list is{num_of_use_preamble_h2h_high_per_frame_list}"
        )
        debug(
            f"num_of_use_preamble_h2h_medium_per_frame_list is {num_of_use_preamble_h2h_medium_per_frame_list}"
        )
        debug(
            f"num_of_use_preamble_h2h_low_per_frame_list is {num_of_use_preamble_h2h_low_per_frame_list}"
        )
        debug(
            f"num_of_use_preamble_m2m_high_per_frame_list is {num_of_use_preamble_m2m_high_per_frame_list}"
        )
        debug(
            f"num_of_use_preamble_m2m_medium_per_frame_list is {num_of_use_preamble_m2m_medium_per_frame_list}"
        )
        debug(
            f"num_of_use_preamble_m2m_low_per_frame_frame_list is {num_of_use_preamble_m2m_low_per_frame_frame_list}"
        )
        debug(f"___________________________________________________________________")
        unserved_user_joined_frame_count = Counter(
            user.joined_frame for user in unserved_user_list
        )  # 对加入的帧数进行计数
        debug(
            f"unserved_user_joined_frame_count is {unserved_user_joined_frame_count}"
        )  # 例如({3: 8, 4: 8, 2: 5, 1: 4})
        debug(
            f"unserved_user_joined_frame_count[2] is {unserved_user_joined_frame_count[2]}"
        )

        debug(f"+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")

        if num1_of_h2h_user_high_per_frame[test_frame - 1] == 0:
            Access_Success_Probility_h2h_user_high = 0
        else:
            Access_Success_Probility_h2h_user_high = (
                num1_of_h2h_user_high_per_frame[test_frame - 1]
                - num_of_unserved_user_h2h_high_per_frame[test_frame - 1]
            ) / num1_of_h2h_user_high_per_frame[test_frame - 1]
            # Access_Success_Probility_h2h_user_high =  h2h_user_high_count / (h2h_user_high_count + unserved_user_counts_per_test_frame_dict["unserved_h2h_user_high_list"])
            debug(
                f"Access_Success_Probility_h2h_user_high is {Access_Success_Probility_h2h_user_high}"
            )
            debug(
                f"len(unserved_h2h_user_high_list) is {num1_of_h2h_user_high_per_frame[test_frame - 1]}"
            )

        if num1_of_h2h_user_medium_per_frame[test_frame - 1] == 0:
            Access_Success_Probility_h2h_user_medium = 0
        else:
            Access_Success_Probility_h2h_user_medium = (
                num1_of_h2h_user_medium_per_frame[test_frame - 1]
                - num_of_unserved_user_h2h_medium_per_frame[test_frame - 1]
            ) / num1_of_h2h_user_medium_per_frame[test_frame - 1]
            debug(
                f"Access_Success_Probility_h2h_user_medium is {Access_Success_Probility_h2h_user_medium}"
            )
            debug(
                f"len(unserved_h2h_user_medium_list) is {num1_of_h2h_user_medium_per_frame[test_frame - 1]}"
            )

        if num1_of_h2h_user_low_per_frame[test_frame - 1] == 0:
            Access_Success_Probility_h2h_user_low = 0
        else:
            Access_Success_Probility_h2h_user_low = (
                num1_of_h2h_user_low_per_frame[test_frame - 1]
                - num_of_unserved_user_h2h_low_per_frame[test_frame - 1]
            ) / num1_of_h2h_user_low_per_frame[test_frame - 1]
            debug(
                f"Access_Success_Probility_h2h_user_low is {Access_Success_Probility_h2h_user_low}"
            )
            debug(
                f"len(unserved_h2h_user_low_list) is {num1_of_h2h_user_low_per_frame[test_frame - 1]}"
            )
        if num1_of_m2m_user_high_per_frame[test_frame - 1] == 0:
            Access_Success_Probility_m2m_user_high = 0
        else:
            Access_Success_Probility_m2m_user_high = (
                num1_of_m2m_user_high_per_frame[test_frame - 1]
                - num_of_unserved_user_m2m_high_per_frame[test_frame - 1]
            ) / num1_of_m2m_user_high_per_frame[test_frame - 1]
            debug(
                f"Access_Success_Probility_m2m_user_high is {Access_Success_Probility_m2m_user_high}"
            )
            debug(
                f"len(unserved_m2m_user_high_list) is {num1_of_m2m_user_high_per_frame[test_frame - 1]}"
            )

        if num1_of_m2m_user_medium_per_frame[test_frame - 1] == 0:
            Access_Success_Probility_m2m_user_medium = 0
        else:
            Access_Success_Probility_m2m_user_medium = (
                num1_of_m2m_user_medium_per_frame[test_frame - 1]
                - num_of_unserved_user_m2m_medium_per_frame[test_frame - 1]
            ) / num1_of_m2m_user_medium_per_frame[test_frame - 1]
            debug(
                f"Access_Success_Probility_m2m_user_medium is {Access_Success_Probility_m2m_user_medium}"
            )
            debug(
                f"len(unserved_m2m_user_medium_list) is {num1_of_m2m_user_medium_per_frame[test_frame - 1]}"
            )

        if num1_of_m2m_user_low_per_frame[test_frame - 1] == 0:
            Access_Success_Probility_m2m_user_low = 0
        else:
            Access_Success_Probility_m2m_user_low = (
                num1_of_m2m_user_low_per_frame[test_frame - 1]
                - num_of_unserved_user_m2m_low_per_frame[test_frame - 1]
            ) / num1_of_m2m_user_low_per_frame[test_frame - 1]
            debug(
                f"Access_Success_Probility_m2m_user_low is {Access_Success_Probility_m2m_user_low}"
            )
            debug(
                f"len(unserved_m2m_user_low_list) is {num1_of_m2m_user_low_per_frame[test_frame - 1]}"
            )

        Access_Success_Probility_frame_data_h2h_high_for_plot.append(
            Access_Success_Probility_h2h_user_high
        )
        # np.savetxt("Access_Success_Probility_frame_data_h2h_high_for_plot.txt",Access_Success_Probility_frame_data_h2h_high_for_plot, delimiter=",")
        Access_Success_Probility_frame_data_h2h_medium_for_plot.append(
            Access_Success_Probility_h2h_user_medium
        )
        # np.savetxt("Access_Success_Probility_frame_data_h2h_medium_for_plot.txt",Access_Success_Probility_frame_data_h2h_medium_for_plot, delimiter=",")
        Access_Success_Probility_frame_data_h2h_low_for_plot.append(
            Access_Success_Probility_h2h_user_low
        )
        # np.savetxt("Access_Success_Probility_frame_data_h2h_low_for_plot.txt",Access_Success_Probility_frame_data_h2h_low_for_plot, delimiter=",")
        Access_Success_Probility_frame_data_m2m_high_for_plot.append(
            Access_Success_Probility_m2m_user_high
        )
        # np.savetxt("Access_Success_Probility_frame_data_m2m_high_for_plot.txt",Access_Success_Probility_frame_data_m2m_high_for_plot, delimiter=",")
        Access_Success_Probility_frame_data_m2m_medium_for_plot.append(
            Access_Success_Probility_m2m_user_medium
        )
        # np.savetxt("Access_Success_Probility_frame_data_m2m_medium_for_plot.txt",
        #            Access_Success_Probility_frame_data_m2m_medium_for_plot, delimiter=",")
        Access_Success_Probility_frame_data_m2m_low_for_plot.append(
            Access_Success_Probility_m2m_user_low
        )
        # np.savetxt("Access_Success_Probility_frame_data_m2m_low_for_plot.txt",Access_Success_Probility_frame_data_m2m_low_for_plot, delimiter=",")
        debug(
            f"Access_Success_Probility_frame_data_h2h_high_for_plot is {Access_Success_Probility_frame_data_h2h_high_for_plot}"
        )
        debug(
            f"Access_Success_Probility_frame_data_h2h_medium_for_plot is {Access_Success_Probility_frame_data_h2h_medium_for_plot}"
        )
        debug(
            f"Access_Success_Probility_frame_data_h2h_low_for_plot is {Access_Success_Probility_frame_data_h2h_low_for_plot}"
        )
        debug(
            f"Access_Success_Probility_frame_data_m2m_high_for_plot is {Access_Success_Probility_frame_data_m2m_high_for_plot}"
        )
        debug(
            f"Access_Success_Probility_frame_data_m2m_medium_for_plot is {Access_Success_Probility_frame_data_m2m_medium_for_plot}"
        )
        debug(
            f"Access_Success_Probility_frame_data_m2m_low_for_plot is {Access_Success_Probility_frame_data_m2m_low_for_plot}"
        )
        debug(
            f"++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
        )
        Average_Success_Probility_h2h_user_high = sum(
            Access_Success_Probility_frame_data_h2h_high_for_plot
        ) / len(Access_Success_Probility_frame_data_h2h_high_for_plot)
        Average_Success_Probability_h2h_user_high_list.append(
            Average_Success_Probility_h2h_user_high
        )

        Average_Success_Probility_h2h_user_medium = sum(
            Access_Success_Probility_frame_data_h2h_medium_for_plot
        ) / len(Access_Success_Probility_frame_data_h2h_medium_for_plot)
        Average_Success_Probability_h2h_user_medium_list.append(
            Average_Success_Probility_h2h_user_medium
        )
        Average_Success_Probility_h2h_user_low = sum(
            Access_Success_Probility_frame_data_h2h_low_for_plot
        ) / len(Access_Success_Probility_frame_data_h2h_low_for_plot)
        Average_Success_Probability_h2h_user_low_list.append(
            Average_Success_Probility_h2h_user_low
        )
        Average_Success_Probility_m2m_user_high = sum(
            Access_Success_Probility_frame_data_m2m_high_for_plot
        ) / len(Access_Success_Probility_frame_data_m2m_high_for_plot)
        Average_Success_Probability_m2m_user_high_list.append(
            Average_Success_Probility_m2m_user_high
        )
        Average_Success_Probility_m2m_user_medium = sum(
            Access_Success_Probility_frame_data_m2m_medium_for_plot
        ) / len(Access_Success_Probility_frame_data_m2m_medium_for_plot)
        Average_Success_Probability_m2m_user_medium_list.append(
            Average_Success_Probility_m2m_user_medium
        )
        Average_Success_Probility_m2m_user_low = sum(
            Access_Success_Probility_frame_data_m2m_low_for_plot
        ) / len(Access_Success_Probility_frame_data_m2m_low_for_plot)
        Average_Success_Probability_m2m_user_low_list.append(
            Average_Success_Probility_m2m_user_low
        )
        debug(
            f"Average_Success_Probility_h2h_user_high is {Average_Success_Probility_h2h_user_high}"
        )
        debug(
            f"Average_Success_Probility_h2h_user_medium is {Average_Success_Probility_h2h_user_medium}"
        )
        debug(
            f"Average_Success_Probility_h2h_user_low is {Average_Success_Probility_h2h_user_low}"
        )
        debug(
            f"Average_Success_Probility_m2m_user_high is {Average_Success_Probility_m2m_user_high}"
        )
        debug(
            f"Average_Success_Probility_m2m_user_medium is {Average_Success_Probility_m2m_user_medium}"
        )
        debug(
            f"Average_Success_Probility_m2m_user_low is {Average_Success_Probility_m2m_user_low}"
        )
        debug(
            f"Average_Success_Probility_h2h_user_high_list is {Average_Success_Probability_h2h_user_high_list}"
        )
        debug(
            f"Average_Success_Probility_h2h_user_medium_list is {Average_Success_Probability_h2h_user_medium_list}"
        )
        debug(
            f"Average_Success_Probility_h2h_user_low_list is {Average_Success_Probability_h2h_user_low_list}"
        )
        debug(
            f"Average_Success_Probility_m2m_user_high_list is {Average_Success_Probability_m2m_user_high_list}"
        )
        debug(
            f"Average_Success_Probility_m2m_user_medium_list is {Average_Success_Probability_m2m_user_medium_list}"
        )
        debug(
            f"Average_Success_Probility_m2m_user_low_list is {Average_Success_Probability_m2m_user_low_list}"
        )

        debug(
            f"++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
        )

        if (
            num1_of_h2h_user_high_per_frame[test_frame - 1]
            + num1_of_h2h_user_medium_per_frame[test_frame - 1]
            + num1_of_h2h_user_low_per_frame[test_frame - 1]
        ) == 0:
            Access_Success_Probility_H2H = 0
        else:
            Access_Success_Probility_H2H = (
                num1_of_h2h_user_high_per_frame[test_frame - 1]
                + num1_of_h2h_user_medium_per_frame[test_frame - 1]
                + num1_of_h2h_user_low_per_frame[test_frame - 1]
                - num_of_unserved_user_h2h_high_per_frame[test_frame - 1]
                - num_of_unserved_user_h2h_medium_per_frame[test_frame - 1]
                - num_of_unserved_user_h2h_low_per_frame[test_frame - 1]
            ) / (
                num1_of_h2h_user_high_per_frame[test_frame - 1]
                + num1_of_h2h_user_medium_per_frame[test_frame - 1]
                + num1_of_h2h_user_low_per_frame[test_frame - 1]
            )
        debug(f"Access_Success_Probility_H2H is {Access_Success_Probility_H2H}")
        Access_Success_Probility_H2H_for_plot.append(Access_Success_Probility_H2H)
        debug(
            f"Access_Success_Probility_H2H_for_plot is {Access_Success_Probility_H2H_for_plot}"
        )
        # np.savetxt("Access_Success_Probility_H2H_for_plot.txt", Access_Success_Probility_H2H_for_plot, delimiter=",")
        debug(
            f"++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
        )
        Average_Success_H2H_Probility = sum(
            Access_Success_Probility_H2H_for_plot
        ) / len(Access_Success_Probility_H2H_for_plot)
        debug(f"Average_Success_H2H_Probility is {Average_Success_H2H_Probility}")
        Average_Success_H2H_Probability_list.append(Average_Success_H2H_Probility)
        debug(
            f"Average_Success_H2H_Probility_list is {Average_Success_H2H_Probability_list}"
        )
        debug(
            f"++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
        )
        debug(
            f"Number_of_H2H is {num1_of_h2h_user_high_per_frame[test_frame - 1] + num1_of_h2h_user_medium_per_frame[test_frame - 1] + num1_of_h2h_user_low_per_frame[test_frame - 1]}"
        )

        if (
            num1_of_m2m_user_high_per_frame[test_frame - 1]
            + num1_of_m2m_user_medium_per_frame[test_frame - 1]
            + num1_of_m2m_user_low_per_frame[test_frame - 1]
        ) == 0:
            Access_Success_Probility_M2M = 0
        else:
            Access_Success_Probility_M2M = (
                num1_of_m2m_user_high_per_frame[test_frame - 1]
                + num1_of_m2m_user_medium_per_frame[test_frame - 1]
                + num1_of_m2m_user_low_per_frame[test_frame - 1]
                - num_of_unserved_user_m2m_high_per_frame[test_frame - 1]
                - num_of_unserved_user_m2m_medium_per_frame[test_frame - 1]
                - num_of_unserved_user_m2m_low_per_frame[test_frame - 1]
            ) / (
                num1_of_m2m_user_high_per_frame[test_frame - 1]
                + num1_of_m2m_user_medium_per_frame[test_frame - 1]
                + num1_of_m2m_user_low_per_frame[test_frame - 1]
            )
        debug(f"Access_Success_Probility_M2M is {Access_Success_Probility_M2M}")
        debug(
            f"Number_of_M2M is {num1_of_m2m_user_high_per_frame[test_frame - 1] + num1_of_m2m_user_medium_per_frame[test_frame - 1] + num1_of_m2m_user_low_per_frame[test_frame - 1]}"
        )
        Access_Success_Probility_M2M_for_plot.append(Access_Success_Probility_M2M)
        debug(
            f"Access_Success_Probility_M2M_for_plot is {Access_Success_Probility_M2M_for_plot}"
        )
        # np.savetxt("Access_Success_Probility_M2M_for_plot.txt", Access_Success_Probility_M2M_for_plot, delimiter=",")
        debug(
            f"++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
        )
        Average_Success_M2M_Probility = sum(
            Access_Success_Probility_M2M_for_plot
        ) / len(Access_Success_Probility_M2M_for_plot)
        debug(f"Average_Success_M2M_Probility is {Average_Success_M2M_Probility}")
        Average_Success_M2M_Probability_list.append(Average_Success_M2M_Probility)
        debug(
            f"Average_Success_M2M_Probility_list is {Average_Success_M2M_Probability_list}"
        )

        debug(
            f"++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
        )

        total_num_of_user_input = (
            num1_of_h2h_user_high_per_frame[test_frame - 1]
            + num1_of_h2h_user_medium_per_frame[test_frame - 1]
            + num1_of_h2h_user_low_per_frame[test_frame - 1]
            + num1_of_m2m_user_high_per_frame[test_frame - 1]
            + num1_of_m2m_user_medium_per_frame[test_frame - 1]
            + num1_of_m2m_user_low_per_frame[test_frame - 1]
        )
        total_unserved_user = (
            num_of_unserved_user_h2h_high_per_frame[test_frame - 1]
            + num_of_unserved_user_h2h_medium_per_frame[test_frame - 1]
            + num_of_unserved_user_h2h_low_per_frame[test_frame - 1]
            + num_of_unserved_user_m2m_high_per_frame[test_frame - 1]
            + num_of_unserved_user_m2m_medium_per_frame[test_frame - 1]
            + num_of_unserved_user_m2m_low_per_frame[test_frame - 1]
        )

        debug(f"total_num_of_user_input is {total_num_of_user_input}")
        Access_Success_Probility = (
            total_num_of_user_input - total_unserved_user
        ) / total_num_of_user_input  # 使用的前导数/总的用户数  （在当前帧中的用户成功接入概率）

        debug(f"Access_Success_Probility is {Access_Success_Probility}")
        Access_Success_Probility_test_frame_data_for_plot.append(test_frame)  #
        Access_Success_Probility_frame_data_for_plot.append(Access_Success_Probility)

        debug(
            f"Access_Success_Probility_test_frame_data_for_plot is {Access_Success_Probility_test_frame_data_for_plot}"
        )
        debug(
            f"Access_Success_Probility_frame_data_for_plot is {Access_Success_Probility_frame_data_for_plot}"
        )

        Access_Success_Probility_average = sum(
            Access_Success_Probility_frame_data_for_plot
        ) / len(Access_Success_Probility_frame_data_for_plot)
        debug(f"Access_Success_Probility_average is {Access_Success_Probility_average}")
        Access_Success_Probability_average_list.append(Access_Success_Probility_average)
        debug(
            f"Access_Success_Probility_average_list is {Access_Success_Probability_average_list}"
        )
        debug(
            f"++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
        )

        # user_type_count = update_blocked_user_list()

        debug(
            f"======================================================================================="
        )

        if sum(curr_state) == 0:
            Preamble_utilization_rate = 0
        else:
            Preamble_utilization_rate = (
                total_num_of_user_input - total_unserved_user
            ) / sum(
                curr_state
            )  # 当前帧使用的前导码数/当前帧总的前导码数
            debug(
                f"Preamble_utilization_rate is {Preamble_utilization_rate} and sum(curr_state) is {sum(curr_state)} "
                f"and total_num_of_user_input - len(unserved_user_list) is {(total_num_of_user_input - total_unserved_user)} "
            )

        Preamble_utilization_rate_test_frame_data_for_plot.append(test_frame)
        Preamble_utilization_rate_frame_data_for_plot.append(Preamble_utilization_rate)

        debug(
            f"Preamble_utilization_rate_test_frame_data_for_plot is {Preamble_utilization_rate_test_frame_data_for_plot}"
        )
        debug(
            f"Preamble_utilization_rate_frame_data_for_plot is {Preamble_utilization_rate_frame_data_for_plot}"
        )
        debug(
            f"++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
        )
        Preamble_utilization_rate_average = sum(
            Preamble_utilization_rate_frame_data_for_plot
        ) / len(Preamble_utilization_rate_frame_data_for_plot)
        debug(
            f" Preamble_utilization_rate_average  is {Preamble_utilization_rate_average}"
        )
        Preamble_utilization_rate_average_data_for_plot.append(
            Preamble_utilization_rate_average
        )
        debug(
            f"Preamble_utilization_rate_average_data_for_plot is {Preamble_utilization_rate_average_data_for_plot}"
        )

        if curr_state[0] == 0:
            Preamble_utilization_rate_h2h_high = 0
        else:
            Preamble_utilization_rate_h2h_high = (
                num1_of_h2h_user_high_per_frame[test_frame - 1]
                - num_of_unserved_user_h2h_high_per_frame[test_frame - 1]
            ) / curr_state[0]
            debug(
                f"Preamble_utilization_rate_h2h_high is {Preamble_utilization_rate_h2h_high}"
            )
        Preamble_utilization_rate_h2h_high_data_for_plot.append(
            Preamble_utilization_rate_h2h_high
        )
        Average_Preamble_utilization_rate_h2h_high = sum(
            Preamble_utilization_rate_h2h_high_data_for_plot
        ) / len(Preamble_utilization_rate_h2h_high_data_for_plot)
        debug(
            f"Average_Preamble_utilization_rate_h2h_high is {Average_Preamble_utilization_rate_h2h_high}"
        )
        Average_Preamble_utilization_rate_h2h_high_list.append(
            Average_Preamble_utilization_rate_h2h_high
        )
        debug(
            f"Average_Preamble_utilization_rate_h2h_high_list is {Average_Preamble_utilization_rate_h2h_high_list}"
        )
        debug(f"++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        if curr_state[1] == 0:
            Preamble_utilization_rate_h2h_medium = 0
        else:
            Preamble_utilization_rate_h2h_medium = (
                num1_of_h2h_user_medium_per_frame[test_frame - 1]
                - num_of_unserved_user_h2h_medium_per_frame[test_frame - 1]
            ) / curr_state[1]
            debug(
                f" Preamble_utilization_rate_h2h_medium is {Preamble_utilization_rate_h2h_medium}"
            )
        Preamble_utilization_rate_h2h_medium_data_for_plot.append(
            Preamble_utilization_rate_h2h_medium
        )
        Average_Preamble_utilization_rate_h2h_medium = sum(
            Preamble_utilization_rate_h2h_medium_data_for_plot
        ) / len(Preamble_utilization_rate_h2h_medium_data_for_plot)
        debug(
            f"Average_Preamble_utilization_rate_h2h_medium is {Average_Preamble_utilization_rate_h2h_medium}"
        )
        Average_Preamble_utilization_rate_h2h_medium_list.append(
            Average_Preamble_utilization_rate_h2h_medium
        )
        debug(
            f"Average_Preamble_utilization_rate_h2h_medium_list is {Average_Preamble_utilization_rate_h2h_medium_list}"
        )
        debug(f"++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")

        if curr_state[2] == 0:
            Preamble_utilization_rate_h2h_low = 0
        else:
            Preamble_utilization_rate_h2h_low = (
                num1_of_h2h_user_low_per_frame[test_frame - 1]
                - num_of_unserved_user_h2h_low_per_frame[test_frame - 1]
            ) / curr_state[2]
            debug(
                f" Preamble_utilization_rate_h2h_low  is {Preamble_utilization_rate_h2h_low}"
            )

        Preamble_utilization_rate_h2h_low_data_for_plot.append(
            Preamble_utilization_rate_h2h_low
        )
        Average_Preamble_utilization_rate_h2h_low = sum(
            Preamble_utilization_rate_h2h_low_data_for_plot
        ) / len(Preamble_utilization_rate_h2h_low_data_for_plot)
        debug(
            f"Average_Preamble_utilization_rate_h2h_low is {Average_Preamble_utilization_rate_h2h_low}"
        )
        Average_Preamble_utilization_rate_h2h_low_list.append(
            Average_Preamble_utilization_rate_h2h_low
        )
        debug(
            f"Average_Preamble_utilization_rate_h2h_low_list is {Average_Preamble_utilization_rate_h2h_low_list}"
        )
        debug(f"++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        if curr_state[3] == 0:
            Preamble_utilization_rate_m2m_high = 0
        else:
            Preamble_utilization_rate_m2m_high = (
                num1_of_m2m_user_high_per_frame[test_frame - 1]
                - num_of_unserved_user_m2m_high_per_frame[test_frame - 1]
            ) / curr_state[3]
            debug(
                f"Preamble_utilization_rate_m2m_high is {Preamble_utilization_rate_m2m_high}"
            )
        Preamble_utilization_rate_m2m_high_data_for_plot.append(
            Preamble_utilization_rate_m2m_high
        )
        Average_Preamble_utilization_rate_m2m_high = sum(
            Preamble_utilization_rate_m2m_high_data_for_plot
        ) / len(Preamble_utilization_rate_m2m_high_data_for_plot)
        debug(
            f"Average_Preamble_utilization_rate_m2m_high is {Average_Preamble_utilization_rate_m2m_high}"
        )
        Average_Preamble_utilization_rate_m2m_high_list.append(
            Average_Preamble_utilization_rate_m2m_high
        )
        debug(
            f" Average_Preamble_utilization_rate_m2m_high_list is { Average_Preamble_utilization_rate_m2m_high_list}"
        )
        debug(f"++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        if curr_state[4] == 0:
            Preamble_utilization_rate_m2m_medium = 0
        else:
            Preamble_utilization_rate_m2m_medium = (
                num1_of_m2m_user_medium_per_frame[test_frame - 1]
                - num_of_unserved_user_m2m_medium_per_frame[test_frame - 1]
            ) / curr_state[4]
            debug(
                f"Preamble_utilization_rate_m2m_medium is {Preamble_utilization_rate_m2m_medium}"
            )
        Preamble_utilization_rate_m2m_medium_data_for_plot.append(
            Preamble_utilization_rate_m2m_medium
        )
        Average_Preamble_utilization_rate_m2m_medium = sum(
            Preamble_utilization_rate_m2m_medium_data_for_plot
        ) / len(Preamble_utilization_rate_m2m_medium_data_for_plot)
        debug(
            f"Average_Preamble_utilization_rate_m2m_medium is {Average_Preamble_utilization_rate_m2m_medium}"
        )
        Average_Preamble_utilization_rate_m2m_medium_list.append(
            Average_Preamble_utilization_rate_m2m_medium
        )
        debug(
            f"Average_Preamble_utilization_rate_m2m_medium_list is {Average_Preamble_utilization_rate_m2m_medium_list}"
        )
        debug(f"++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")

        if curr_state[5] == 0:
            Preamble_utilization_rate_m2m_low = 0
        else:
            Preamble_utilization_rate_m2m_low = (
                num1_of_m2m_user_low_per_frame[test_frame - 1]
                - num_of_unserved_user_m2m_low_per_frame[test_frame - 1]
            ) / curr_state[5]
            debug(
                f"Preamble_utilization_rate_m2m_low is {Preamble_utilization_rate_m2m_low}"
            )
        Preamble_utilization_rate_m2m_low_data_for_plot.append(
            Preamble_utilization_rate_m2m_low
        )
        Average_Preamble_utilization_rate_m2m_low = sum(
            Preamble_utilization_rate_m2m_low_data_for_plot
        ) / len(Preamble_utilization_rate_m2m_low_data_for_plot)
        debug(
            f"Average_Preamble_utilization_rate_m2m_low is {Average_Preamble_utilization_rate_m2m_low}"
        )
        Average_Preamble_utilization_rate_m2m_low_list.append(
            Average_Preamble_utilization_rate_m2m_low
        )
        debug(
            f"Average_Preamble_utilization_rate_m2m_low_list is {Average_Preamble_utilization_rate_m2m_low_list}"
        )
        debug(f"++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        if (curr_state[0] + curr_state[1] + curr_state[2]) == 0:
            Preamble_utilization_rate_H2H_use = 0
        else:
            Preamble_utilization_rate_H2H_user = (
                num_of_use_preamble_h2h_high_per_frame_list[test_frame - 1]
                + num_of_use_preamble_h2h_medium_per_frame_list[test_frame - 1]
                + num_of_use_preamble_h2h_low_per_frame_list[test_frame - 1]
            ) / (curr_state[0] + curr_state[1] + curr_state[2])
            debug(
                f"Preamble_utilization_rate_H2H_user is {Preamble_utilization_rate_H2H_user}"
            )
            Preamble_utilization_rate_H2H_user_for_plot.append(
                Preamble_utilization_rate_H2H_user
            )
            Average_Preamble_utilization_rate_H2H = sum(
                Preamble_utilization_rate_H2H_user_for_plot
            ) / len(Preamble_utilization_rate_H2H_user_for_plot)
            debug(
                f"Average_Preamble_utilization_rate_H2H is {Average_Preamble_utilization_rate_H2H}"
            )
            Average_Preamble_utilization_rate_H2H_list.append(
                Average_Preamble_utilization_rate_H2H
            )
            debug(
                f"Average_Preamble_utilization_rate_H2H_list is {Average_Preamble_utilization_rate_H2H_list}"
            )
            debug(f"++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")

        if (curr_state[3] + curr_state[4] + curr_state[5]) == 0:
            Preamble_utilization_rate_M2M_user = 0
        else:
            Preamble_utilization_rate_M2M_user = (
                num_of_use_preamble_m2m_high_per_frame_list[test_frame - 1]
                + num_of_use_preamble_m2m_medium_per_frame_list[test_frame - 1]
                + num_of_use_preamble_m2m_low_per_frame_frame_list[test_frame - 1]
            ) / (curr_state[3] + curr_state[4] + curr_state[5])
            debug(
                f"Preamble_utilization_rate_M2M_user is {Preamble_utilization_rate_M2M_user}"
            )
            Preamble_utilization_rate_M2M_user_for_plot.append(
                Preamble_utilization_rate_M2M_user
            )
            Average_Preamble_utilization_rate_M2M = sum(
                Preamble_utilization_rate_M2M_user_for_plot
            ) / len(Preamble_utilization_rate_M2M_user_for_plot)

            debug(
                f"Average_Preamble_utilization_rate_M2M is {Average_Preamble_utilization_rate_M2M}"
            )
            Average_Preamble_utilization_rate_M2M_list.append(
                Average_Preamble_utilization_rate_M2M
            )
            debug(
                f"Average_Preamble_utilization_rate_M2M_list is {Average_Preamble_utilization_rate_M2M_list}"
            )
            debug(f"++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")

        debug(f"++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")

        update_blocked_user_list()
        debug(f"{sys._getframe().f_lineno}")
        user_preamble_assign_delay_calculation(test_frame)
        # 20230925 用于计算所有用户的平均前导码分配时延 Start
        debug(
            f"averageAllUserPreambleAssignDelayList is {averageAllUserPreambleAssignDelayList}"
        )
        debug(
            f"averageH2hHighPreambleAssignDelayList is {averageH2hHighPreambleAssignDelayList}"
        )
        debug(
            f"averageH2hMediumPreambleAssignDelayList is {averageH2hMediumPreambleAssignDelayList}"
        )
        debug(
            f"averageH2hLowPreambleAssignDelayList is {averageH2hLowPreambleAssignDelayList}"
        )
        debug(
            f"averageM2mHighPreambleAssignDelayList is {averageM2mHighPreambleAssignDelayList}"
        )
        debug(
            f"averageM2mMediumPreambleAssignDelayList is {averageM2mMediumPreambleAssignDelayList}"
        )
        debug(
            f"averageM2mLowPreambleAssignDelayList is {averageM2mLowPreambleAssignDelayList}"
        )
        debug(
            f"averageH2hPreambleAssignDelayList is {averageH2hPreambleAssignDelayList}"
        )
        debug(
            f"averageM2mPreambleAssignDelayList is {averageM2mPreambleAssignDelayList}"
        )
        debug(f"+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")

        debug(f"+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        averageAllUserPreambleAssignDelay = sum(
            averageAllUserPreambleAssignDelayList
        ) / len(averageAllUserPreambleAssignDelayList)
        averageH2hHighPreambleAssignDelay = sum(
            averageH2hHighPreambleAssignDelayList
        ) / len(averageH2hHighPreambleAssignDelayList)
        averageH2hMediumPreambleAssignDelay = sum(
            averageH2hMediumPreambleAssignDelayList
        ) / len(averageH2hMediumPreambleAssignDelayList)
        averageH2hLowPreambleAssignDelay = sum(
            averageH2hLowPreambleAssignDelayList
        ) / len(averageH2hLowPreambleAssignDelayList)
        averageM2mHighPreambleAssignDelay = sum(
            averageM2mHighPreambleAssignDelayList
        ) / len(averageM2mHighPreambleAssignDelayList)
        averageM2mMediumPreambleAssignDelay = sum(
            averageM2mMediumPreambleAssignDelayList
        ) / len(averageM2mMediumPreambleAssignDelayList)
        averageM2mLowPreambleAssignDelay = sum(
            averageM2mLowPreambleAssignDelayList
        ) / len(averageM2mLowPreambleAssignDelayList)
        averageH2hPreambleAssignDelay = sum(averageH2hPreambleAssignDelayList) / len(
            averageH2hPreambleAssignDelayList
        )
        averageM2mPreambleAssignDelay = sum(averageM2mPreambleAssignDelayList) / len(
            averageM2mPreambleAssignDelayList
        )
        debug(
            f"averageAllUserPreambleAssignDelay is {averageAllUserPreambleAssignDelay}"
        )
        debug(
            f"averageH2hHighPreambleAssignDelay is {averageH2hHighPreambleAssignDelay}"
        )
        debug(
            f"averageH2hMediumPreambleAssignDelay is {averageH2hMediumPreambleAssignDelay}"
        )
        debug(f"averageH2hLowPreambleAssignDelay is {averageH2hLowPreambleAssignDelay}")
        debug(
            f"averageM2mHighPreambleAssignDelay is {averageM2mHighPreambleAssignDelay}"
        )
        debug(
            f"averageM2mMediumPreambleAssignDelay is {averageM2mMediumPreambleAssignDelay}"
        )
        debug(f"averageM2mLowPreambleAssignDelay is {averageM2mLowPreambleAssignDelay}")
        debug(f"averageH2hPreambleAssignDelay is {averageH2hPreambleAssignDelay}")
        debug(f"averageM2mPreambleAssignDelay is {averageM2mPreambleAssignDelay}")
        debug(
            f"+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++"
        )  # 20230925 用于计算所有用户的平均前导码分配时延 End
        debug(f"_____________________________________________________________________")

        if flag_initial_run:
            # 如果不存在被阻塞用户，直接更新当前状态
            if len(blocked_user_list) == 0:
                action = random.choice(action_feature)
                debug(f"random action is {action}")
                next_state = choose_next_state(curr_state, action, 0.0)
                curr_state = next_state
                continue
            else:
                flag_initial_run = False

        # user_block_rate_calculation(test_frame)
        user_block_rate_calculation(test_frame)
        debug(f"{sys._getframe().f_lineno}")
        debug(f"block_probs is {block_probs}")

        block_probs_list.append(
            (
                test_frame,
                block_probs["h2h_high"],
                block_probs["h2h_medium"],
                block_probs["h2h_low"],
                block_probs["m2m_high"],
                block_probs["m2m_medium"],
                block_probs["m2m_low"],
                block_probs["H2H"],
                block_probs["M2M"],
                block_probs["ALL"],
            )
        )
        debug(f"block_probs_list is {block_probs_list}")

        clean_blocked_user_list(test_frame)
        debug(f"{sys._getframe().f_lineno}")
        action = choose_action(curr_state, q_net)
        debug(f"action is {action}")
        next_state = choose_next_state(curr_state, action, 0.0)
        debug(f"next_state is {next_state}")
        debug(f"==========================================")
        curr_state = next_state
        debug(f"curr_state is {curr_state}")

        debug(f"==========================================")

        # for unserved_user in unserved_user_list:  # 遍历未服务用户列表
        debug(f"len(unserved_user_list) is {len(unserved_user_list)}")
        debug(f"len(unserved_h2h_user_high_list) is {len(unserved_h2h_user_high_list)}")
        debug(
            f"len(unserved_h2h_user_medium_list) is {len(unserved_h2h_user_medium_list)}"
        )
        debug(f"len(unserved_h2h_user_low_list) is {len(unserved_h2h_user_low_list)}")
        debug(f"len(unserved_m2m_user_high_list) is {len(unserved_m2m_user_high_list)}")
        debug(
            f"len(unserved_m2m_user_medium_list) is {len(unserved_m2m_user_medium_list)}"
        )
        debug(f"len(unserved_m2m_user_low_list) is {len(unserved_m2m_user_low_list)}")

        num_of_remove_blocked_user_unserved_user_h2h_high_per_frame.append(
            len(unserved_h2h_user_high_list)
        )
        num_of_remove_blocked_user_unserved_user_h2h_medium_per_frame.append(
            len(unserved_h2h_user_medium_list)
        )
        num_of_remove_blocked_user_unserved_user_h2h_low_per_frame.append(
            len(unserved_h2h_user_low_list)
        )
        num_of_remove_blocked_user_unserved_user_m2m_high_per_frame.append(
            len(unserved_m2m_user_high_list)
        )
        num_of_remove_blocked_user_unserved_user_m2m_medium_per_frame.append(
            len(unserved_m2m_user_medium_list)
        )
        num_of_remove_blocked_user_unserved_user_m2m_low_per_frame.append(
            len(unserved_m2m_user_low_list)
        )
        debug(
            f"num_of_remove_blocked_user_unserved_user_h2h_high_per_frame is {num_of_remove_blocked_user_unserved_user_h2h_high_per_frame}"
        )
        debug(
            f"num_of_remove_blocked_user_unserved_user_h2h_medium_per_frame is {num_of_remove_blocked_user_unserved_user_h2h_medium_per_frame}"
        )
        debug(
            f"num_of_remove_blocked_user_unserved_user_h2h_low_per_frame is {num_of_remove_blocked_user_unserved_user_h2h_low_per_frame}"
        )
        debug(
            f"num_of_remove_blocked_user_unserved_user_m2m_high_per_frame is {num_of_remove_blocked_user_unserved_user_m2m_high_per_frame}"
        )
        debug(
            f"num_of_remove_blocked_user_unserved_user_m2m_medium_per_frame is {num_of_remove_blocked_user_unserved_user_m2m_medium_per_frame}"
        )
        debug(
            f"num_of_remove_blocked_user_unserved_user_m2m_low_per_frame is {num_of_remove_blocked_user_unserved_user_m2m_low_per_frame}"
        )

        debug(f"=============================================================")

    plot_test_block_probs(block_probs_list)  # 绘制测试阶段的阻塞概率图
    plt_Access_Success_Probility(
        test_frame,
        Access_Success_Probility_test_frame_data_for_plot,
        Access_Success_Probility_frame_data_for_plot,
        Access_Success_Probility_frame_data_h2h_high_for_plot,
        Access_Success_Probility_frame_data_h2h_medium_for_plot,
        Access_Success_Probility_frame_data_h2h_low_for_plot,
        Access_Success_Probility_frame_data_m2m_high_for_plot,
        Access_Success_Probility_frame_data_m2m_medium_for_plot,
        Access_Success_Probility_frame_data_m2m_low_for_plot,
    )  # 绘制测试阶段的接入成功率图

    plt_Preamble_utilization_rate(
        test_frame,
        Preamble_utilization_rate_test_frame_data_for_plot,
        Preamble_utilization_rate_h2h_high_data_for_plot,
        Preamble_utilization_rate_h2h_medium_data_for_plot,
        Preamble_utilization_rate_h2h_low_data_for_plot,
        Preamble_utilization_rate_m2m_high_data_for_plot,
        Preamble_utilization_rate_m2m_medium_data_for_plot,
        Preamble_utilization_rate_m2m_low_data_for_plot,
        Preamble_utilization_rate_H2H_user_for_plot,
        Preamble_utilization_rate_M2M_user_for_plot,
    )  # 绘制测试阶段的前导码利用率图
    # plt_delay_time(test_frame, delay_time_test_frame_data_for_plot, delay_time_frame_data_for_plot)  # 绘制测试阶段的时延图


def save_data_to_excel(list_to_save, list_name: str):
    file = list_name + ".xlsx"
    if not os.path.exists(file):
        openpyxl.Workbook().save(file)

    # print(list_name)
    wb = openpyxl.load_workbook(list_name + ".xlsx")
    sheet_names = wb.sheetnames
    table = wb[sheet_names[0]]
    n_rows = table.max_row  # 获得行数
    print(f"n_rows of {list_name} is {n_rows}")
    n_columns = 1
    for value in list_to_save:
        table.cell(n_rows + 1, n_columns).value = value
        n_columns += 1
    wb.save(list_name + ".xlsx")


if __name__ == "__main__":
    print(f"EPSILON is {EPSILON} N_HIDDEN is {N_HIDDEN} ALPHA is {ALPHA}")
    convergenceCheckMode = int(
        input(
            "Please select the mode: 1 for regular convergence check (RMS only), 2 for double convergence check (RMS and LOSS): "
        )
    )
    if convergenceCheckMode == 1:
        print("Regular convergence check (RMS only) mode")
    elif convergenceCheckMode == 2:
        print("Double convergence check (RMS and loss) mode")
        listLossPrevHundredFrames = []

    loopModeSelector = int(
        input("Please select the mode: 1 for regular run, 2 for loop run: ")
    )
    if loopModeSelector == 1:
        print("Regular run mode")

        globalModeSelector = int(
            input(
                "Please select the mode: 1 for fixed user number per frame, 2 for random user number per frame: "
            )
        )

        collisionDetectionMode = int(
            input("Collision Detection Mode: 0 for Disable, 1 for Enable: ")
        )

        if globalModeSelector == 1:
            num_of_h2h_user_high_per_frame = int(
                input(
                    "Please input the number of H2H user with high priority per frame: "
                )
            )
            num_of_h2h_user_medium_per_frame = int(
                input(
                    "Please input the number of H2H user with medium priority per frame: "
                )
            )

            num_of_h2h_user_low_per_frame = int(
                input(
                    "Please input the number of H2H user with low priority per frame: "
                )
            )

            num_of_m2m_user_high_per_frame = int(
                input(
                    "Please input the number of M2M user with high priority per frame: "
                )
            )
            num_of_m2m_user_medium_per_frame = int(
                input(
                    "Please input the number of M2M user with medium priority per frame: "
                )
            )

            num_of_m2m_user_low_per_frame = int(
                input(
                    "Please input the number of M2M user with low priority per frame: "
                )
            )

            print(
                f"num_of_h2h_user_high_per_frame is {num_of_h2h_user_high_per_frame},\n "
                f"num_of_h2h_user_medium_per_frame is {num_of_h2h_user_medium_per_frame},\n "
                f"num_of_h2h_user_low_per_frame is {num_of_h2h_user_low_per_frame},\n "
                f"num_of_m2m_user_high_per_frame is {num_of_m2m_user_high_per_frame},\n "
                f"num_of_m2m_user_medium_per_frame is {num_of_m2m_user_medium_per_frame},\n "
                f"num_of_m2m_user_low_per_frame is {num_of_m2m_user_low_per_frame},\n "
            )
            listForH2hUserHigh = []
            listForH2hUserMedium = []
            listForH2hUserLow = []
            listForM2mUserHigh = []
            listForM2mUserMedium = []
            listForM2mUserLow = []

            listForM2mUser = []
            listForH2hUser = []
            listForAllUser = []

        elif globalModeSelector == 2:
            numOfRandomUserPerFrame = int(
                input("Please input the number of random user per frame: ")
            )
            print(f"numOfRandomUserPerFrame is {numOfRandomUserPerFrame}\n")

            listForH2hUserHigh = []
            listForH2hUserMedium = []
            listForH2hUserLow = []
            listForM2mUserHigh = []
            listForM2mUserMedium = []
            listForM2mUserLow = []

            listForM2mUser = []
            listForH2hUser = []
            listForAllUser = []

        else:
            print("Wrong mode selector!")
            sys.exit(1)

        num_of_frame_for_calc = int(
            input("Please input the number of frame for blocking rate calculation: ")
        )

        print(f"num_of_frame_for_calc is {num_of_frame_for_calc}\n")

        total_served_user_list = []  # 总服务用户列表
        averageAllUserPreambleAssignDelayList = []  # 平均用户前导码分配时延列表
        averageH2hHighPreambleAssignDelayList = []  # 平均H2H高优先级用户前导码分配时延列表
        averageH2hMediumPreambleAssignDelayList = []  # 平均H2H中优先级用户前导码分配时延列表
        averageH2hLowPreambleAssignDelayList = []  # 平均H2H低优先级用户前导码分配时延列表
        averageM2mHighPreambleAssignDelayList = []  # 平均M2M高优先级用户前导码分配时延列表
        averageM2mMediumPreambleAssignDelayList = []
        averageM2mLowPreambleAssignDelayList = []
        averageH2hPreambleAssignDelayList = []
        averageM2mPreambleAssignDelayList = []

        avgAllUserPerFrameCollisionRateList = []  # 平均每帧用户碰撞率列表
        avgH2hHighPerFrameCollisionRateList = []
        avgH2hMediumPerFrameCollisionRateList = []
        avgH2hLowPerFrameCollisionRateList = []
        avgM2mHighPerFrameCollisionRateList = []
        avgM2mMediumPerFrameCollisionRateList = []
        avgM2mLowPerFrameCollisionRateList = []
        avgH2hPerFrameCollisionRateList = []
        avgM2mPerFrameCollisionRateList = []

        Average_Success_Probability_h2h_user_high_list = []
        Average_Success_Probability_h2h_user_medium_list = []
        Average_Success_Probability_h2h_user_low_list = []
        Average_Success_Probability_m2m_user_high_list = []
        Average_Success_Probability_m2m_user_medium_list = []
        Average_Success_Probability_m2m_user_low_list = []
        Average_Success_H2H_Probability_list = []
        Average_Success_M2M_Probability_list = []
        Access_Success_Probability_average_list = []

        Average_Preamble_utilization_rate_h2h_high_list = []
        Average_Preamble_utilization_rate_h2h_medium_list = []
        Average_Preamble_utilization_rate_h2h_low_list = []
        Average_Preamble_utilization_rate_m2m_high_list = []
        Average_Preamble_utilization_rate_m2m_medium_list = []
        Average_Preamble_utilization_rate_m2m_low_list = []
        Average_Preamble_utilization_rate_H2H_list = []
        Average_Preamble_utilization_rate_M2M_list = []

        # 初始化4个前导码列表
        preambleListForM2mLow = PreambleList()
        PreambleListForM2mMedium = PreambleList()
        preambleListForM2mHigh = PreambleList()
        preambleListForH2hLow = PreambleList()
        PreambleListForH2hMedium = PreambleList()
        preambleListForH2hHigh = PreambleList()

        # 初始化阻塞率字典
        block_probs = {
            "h2h_high": 0.0,
            "h2h_medium": 0.0,
            "h2h_low": 0.0,
            "m2m_high": 0.0,
            "m2m_medium": 0.0,
            "m2m_low": 0.0,
            "H2H": 0.0,
            "M2M": 0.0,
            "ALL": 0.0,
        }

        # 初始化权重字典
        user_priority_weights = {
            "h2h": {
                "high": 0.5,
                "medium": 0.3,
                "low": 0.2,
            },
            "m2m": {
                "high": 0.5,
                "medium": 0.3,
                "low": 0.2,
            },
        }

        user_type_weights = {"h2h": 0.7, "m2m": 0.3}
        num1_of_h2h_user_high_per_frame = []
        num1_of_h2h_user_medium_per_frame = []
        num1_of_h2h_user_low_per_frame = []
        num1_of_m2m_user_high_per_frame = []
        num1_of_m2m_user_medium_per_frame = []
        num1_of_m2m_user_low_per_frame = []

        num_of_unserved_user_h2h_high_per_frame = []
        num_of_unserved_user_h2h_medium_per_frame = []
        num_of_unserved_user_h2h_low_per_frame = []
        num_of_unserved_user_m2m_high_per_frame = []
        num_of_unserved_user_m2m_medium_per_frame = []
        num_of_unserved_user_m2m_low_per_frame = []

        num_test_frame_new_input_m2m_high_list = []
        num_test_frame_new_input_m2m_medium_list = []
        num_test_frame_new_input_m2m_low_list = []
        num_test_frame_new_input_h2h_high_list = []
        num_tset_frame_new_input_h2h_medium_list = []
        num_test_frame_new_input_h2h_low_list = []

        num_of_remove_blocked_user_unserved_user_h2h_high_per_frame = []
        num_of_remove_blocked_user_unserved_user_h2h_medium_per_frame = []
        num_of_remove_blocked_user_unserved_user_h2h_low_per_frame = []
        num_of_remove_blocked_user_unserved_user_m2m_high_per_frame = []
        num_of_remove_blocked_user_unserved_user_m2m_medium_per_frame = []
        num_of_remove_blocked_user_unserved_user_m2m_low_per_frame = []

        rl()

    elif loopModeSelector == 2:
        print("Loop run mode")

        globalModeSelector = int(
            input(
                "Please select the mode: 1 for fixed user number per frame, 2 for random user number per frame: "
            )
        )

        num_of_frame_for_calc = int(
            input("Please input the number of frame for blocking rate calculation: ")
        )

        print(f"num_of_frame_for_calc is {num_of_frame_for_calc}\n")

        collisionDetectionMode = int(
            input("Collision Detection Mode: 0 for Disable, 1 for Enable: ")
        )

        # for collisionDetectionMode in range(0, 2):
        if collisionDetectionMode == 1:
            if globalModeSelector == 1:
                for (
                    num_of_h2h_user_high_per_frame,
                    num_of_h2h_user_medium_per_frame,
                    num_of_h2h_user_low_per_frame,
                    num_of_m2m_user_high_per_frame,
                    num_of_m2m_user_medium_per_frame,
                    num_of_m2m_user_low_per_frame,
                ) in [
                    [3, 1, 1, 3, 1, 1],
                    [4, 2, 1, 4, 2, 1],
                    [5, 3, 2, 5, 3, 2],
                    [6, 4, 2, 6, 4, 2],
                    [8, 4, 3, 8, 4, 3],
                    [9, 5, 3, 9, 5, 3],
                    [10, 6, 4, 10, 6, 4],
                    [11, 7, 4, 11, 7, 4],
                    [13, 8, 4, 13, 8, 4],
                    [14, 8, 5, 14, 8, 5],
                ]:
                    print(
                        f"num_of_h2h_user_high_per_frame is {num_of_h2h_user_high_per_frame},\n "
                        f"num_of_h2h_user_medium_per_frame is {num_of_h2h_user_medium_per_frame},\n "
                        f"num_of_h2h_user_low_per_frame is {num_of_h2h_user_low_per_frame},\n "
                        f"num_of_m2m_user_high_per_frame is {num_of_m2m_user_high_per_frame},\n "
                        f"num_of_m2m_user_medium_per_frame is {num_of_m2m_user_medium_per_frame},\n "
                        f"num_of_m2m_user_low_per_frame is {num_of_m2m_user_low_per_frame},\n "
                    )

                    listForH2hUserHigh = []
                    listForH2hUserMedium = []
                    listForH2hUserLow = []
                    listForM2mUserHigh = []
                    listForM2mUserMedium = []
                    listForM2mUserLow = []
                    listForM2mUser = []
                    listForH2hUser = []
                    listForAllUser = []

                    total_served_user_list = []  # 总服务用户列表
                    averageAllUserPreambleAssignDelayList = []  # 平均用户前导码分配时延列表
                    averageH2hHighPreambleAssignDelayList = []  # 平均H2H高优先级用户前导码分配时延列表
                    averageH2hMediumPreambleAssignDelayList = []  # 平均H2H中优先级用户前导码分配时延列表
                    averageH2hLowPreambleAssignDelayList = []  # 平均H2H低优先级用户前导码分配时延列表
                    averageM2mHighPreambleAssignDelayList = []  # 平均M2M高优先级用户前导码分配时延列表
                    averageM2mMediumPreambleAssignDelayList = []
                    averageM2mLowPreambleAssignDelayList = []
                    averageH2hPreambleAssignDelayList = []
                    averageM2mPreambleAssignDelayList = []

                    Average_Success_Probability_h2h_user_high_list = []
                    Average_Success_Probability_h2h_user_medium_list = []
                    Average_Success_Probability_h2h_user_low_list = []
                    Average_Success_Probability_m2m_user_high_list = []
                    Average_Success_Probability_m2m_user_medium_list = []
                    Average_Success_Probability_m2m_user_low_list = []
                    Average_Success_H2H_Probability_list = []
                    Average_Success_M2M_Probability_list = []
                    Access_Success_Probability_average_list = []

                    avgAllUserPerFrameCollisionRateList = []  # 平均每帧用户碰撞率列表
                    avgH2hHighPerFrameCollisionRateList = []
                    avgH2hMediumPerFrameCollisionRateList = []
                    avgH2hLowPerFrameCollisionRateList = []
                    avgM2mHighPerFrameCollisionRateList = []
                    avgM2mMediumPerFrameCollisionRateList = []
                    avgM2mLowPerFrameCollisionRateList = []
                    avgH2hPerFrameCollisionRateList = []
                    avgM2mPerFrameCollisionRateList = []

                    block_rate_h2h_high_list = []
                    block_rate_h2h_medium_list = []
                    block_rate_h2h_low_list = []
                    block_rate_m2m_high_list = []
                    block_rate_m2m_medium_list = []
                    block_rate_m2m_low_list = []
                    block_rate_H2H_list = []
                    block_rate_M2M_list = []
                    block_rate_ALL_list = []

                    loss_list_for_plot = []
                    convergence_frame_list = []
                    convergence_value_list = []

                    Average_Preamble_utilization_rate_h2h_high_list = []
                    Average_Preamble_utilization_rate_h2h_medium_list = []
                    Average_Preamble_utilization_rate_h2h_low_list = []
                    Average_Preamble_utilization_rate_m2m_high_list = []
                    Average_Preamble_utilization_rate_m2m_medium_list = []
                    Average_Preamble_utilization_rate_m2m_low_list = []
                    Average_Preamble_utilization_rate_H2H_list = []
                    Average_Preamble_utilization_rate_M2M_list = []

                    frames = []
                    values_list = []
                    loss_values_list = []

                    # 初始化4个前导码列表
                    preambleListForM2mLow = PreambleList()
                    PreambleListForM2mMedium = PreambleList()
                    preambleListForM2mHigh = PreambleList()
                    preambleListForH2hLow = PreambleList()
                    PreambleListForH2hMedium = PreambleList()
                    preambleListForH2hHigh = PreambleList()

                    # 初始化阻塞率字典
                    block_probs = {
                        "h2h_high": 0.0,
                        "h2h_medium": 0.0,
                        "h2h_low": 0.0,
                        "m2m_high": 0.0,
                        "m2m_medium": 0.0,
                        "m2m_low": 0.0,
                        "H2H": 0.0,
                        "M2M": 0.0,
                        "ALL": 0.0,
                    }

                    # 初始化权重字典
                    user_priority_weights = {
                        "h2h": {
                            "high": 0.5,
                            "medium": 0.3,
                            "low": 0.2,
                        },
                        "m2m": {
                            "high": 0.5,
                            "medium": 0.3,
                            "low": 0.2,
                        },
                    }

                    user_type_weights = {"h2h": 0.7, "m2m": 0.3}
                    num1_of_h2h_user_high_per_frame = []
                    num1_of_h2h_user_medium_per_frame = []
                    num1_of_h2h_user_low_per_frame = []
                    num1_of_m2m_user_high_per_frame = []
                    num1_of_m2m_user_medium_per_frame = []
                    num1_of_m2m_user_low_per_frame = []

                    num_of_unserved_user_h2h_high_per_frame = []
                    num_of_unserved_user_h2h_medium_per_frame = []
                    num_of_unserved_user_h2h_low_per_frame = []
                    num_of_unserved_user_m2m_high_per_frame = []
                    num_of_unserved_user_m2m_medium_per_frame = []
                    num_of_unserved_user_m2m_low_per_frame = []

                    num_test_frame_new_input_m2m_high_list = []
                    num_test_frame_new_input_m2m_medium_list = []
                    num_test_frame_new_input_m2m_low_list = []
                    num_test_frame_new_input_h2h_high_list = []
                    num_tset_frame_new_input_h2h_medium_list = []
                    num_test_frame_new_input_h2h_low_list = []

                    num_of_remove_blocked_user_unserved_user_h2h_high_per_frame = []
                    num_of_remove_blocked_user_unserved_user_h2h_medium_per_frame = []
                    num_of_remove_blocked_user_unserved_user_h2h_low_per_frame = []
                    num_of_remove_blocked_user_unserved_user_m2m_high_per_frame = []
                    num_of_remove_blocked_user_unserved_user_m2m_medium_per_frame = []
                    num_of_remove_blocked_user_unserved_user_m2m_low_per_frame = []

                    rl()

                    save_data_to_excel(
                        averageAllUserPreambleAssignDelayList,
                        "averageAllUserPreambleAssignDelayList",
                    )
                    save_data_to_excel(
                        averageH2hHighPreambleAssignDelayList,
                        "averageH2hHighPreambleAssignDelayList",
                    )
                    save_data_to_excel(
                        averageH2hMediumPreambleAssignDelayList,
                        "averageH2hMediumPreambleAssignDelayList",
                    )
                    save_data_to_excel(
                        averageH2hLowPreambleAssignDelayList,
                        "averageH2hLowPreambleAssignDelayList",
                    )
                    save_data_to_excel(
                        averageM2mHighPreambleAssignDelayList,
                        "averageM2mHighPreambleAssignDelayList",
                    )
                    save_data_to_excel(
                        averageM2mMediumPreambleAssignDelayList,
                        "averageM2mMediumPreambleAssignDelayList",
                    )
                    save_data_to_excel(
                        averageM2mLowPreambleAssignDelayList,
                        "averageM2mLowPreambleAssignDelayList",
                    )
                    save_data_to_excel(
                        averageH2hPreambleAssignDelayList,
                        "averageH2hPreambleAssignDelayList",
                    )
                    save_data_to_excel(
                        averageM2mPreambleAssignDelayList,
                        "averageM2mPreambleAssignDelayList",
                    )

                    save_data_to_excel(
                        avgAllUserPerFrameCollisionRateList,
                        "avgAllUserPerFrameCollisionRateList",
                    )
                    save_data_to_excel(
                        avgH2hHighPerFrameCollisionRateList,
                        "avgH2hHighPerFrameCollisionRateList",
                    )
                    save_data_to_excel(
                        avgH2hMediumPerFrameCollisionRateList,
                        "avgH2hMediumPerFrameCollisionRateList",
                    )
                    save_data_to_excel(
                        avgH2hLowPerFrameCollisionRateList,
                        "avgH2hLowPerFrameCollisionRateList",
                    )
                    save_data_to_excel(
                        avgM2mHighPerFrameCollisionRateList,
                        "avgM2mHighPerFrameCollisionRateList",
                    )
                    save_data_to_excel(
                        avgM2mMediumPerFrameCollisionRateList,
                        "avgM2mMediumPerFrameCollisionRateList",
                    )
                    save_data_to_excel(
                        avgM2mLowPerFrameCollisionRateList,
                        "avgM2mLowPerFrameCollisionRateList",
                    )
                    save_data_to_excel(
                        avgH2hPerFrameCollisionRateList,
                        "avgH2hPerFrameCollisionRateList",
                    )
                    save_data_to_excel(
                        avgM2mPerFrameCollisionRateList,
                        "avgM2mPerFrameCollisionRateList",
                    )

                    save_data_to_excel(
                        Access_Success_Probility_frame_data_h2h_high_for_plot,
                        "Access_Success_Probility_frame_data_h2h_high_for_plot",
                    )
                    save_data_to_excel(
                        Access_Success_Probility_frame_data_h2h_medium_for_plot,
                        "Access_Success_Probility_frame_data_h2h_medium_for_plot",
                    )
                    save_data_to_excel(
                        Access_Success_Probility_frame_data_h2h_low_for_plot,
                        "Access_Success_Probility_frame_data_h2h_low_for_plot",
                    )
                    save_data_to_excel(
                        Access_Success_Probility_frame_data_m2m_high_for_plot,
                        "Access_Success_Probility_frame_data_m2m_high_for_plot",
                    )
                    save_data_to_excel(
                        Access_Success_Probility_frame_data_m2m_medium_for_plot,
                        "Access_Success_Probility_frame_data_m2m_medium_for_plot",
                    )
                    save_data_to_excel(
                        Access_Success_Probility_frame_data_m2m_low_for_plot,
                        "Access_Success_Probility_frame_data_m2m_low_for_plot",
                    )
                    save_data_to_excel(
                        Access_Success_Probility_frame_data_for_plot,
                        "Access_Success_Probility_frame_data_for_plot",
                    )
                    save_data_to_excel(
                        Access_Success_Probility_H2H_for_plot,
                        "Access_Success_Probility_H2H_for_plot",
                    )
                    save_data_to_excel(
                        Access_Success_Probility_M2M_for_plot,
                        "Access_Success_Probility_M2M_for_plot",
                    )

                    save_data_to_excel(
                        Average_Success_Probability_h2h_user_high_list,
                        "Average_Success_Probability_h2h_user_high_list",
                    )
                    save_data_to_excel(
                        Average_Success_Probability_h2h_user_medium_list,
                        "Average_Success_Probability_h2h_user_medium_list",
                    )
                    save_data_to_excel(
                        Average_Success_Probability_h2h_user_low_list,
                        "Average_Success_Probability_h2h_user_low_list",
                    )
                    save_data_to_excel(
                        Average_Success_Probability_m2m_user_high_list,
                        "Average_Success_Probability_m2m_user_high_list",
                    )
                    save_data_to_excel(
                        Average_Success_Probability_m2m_user_medium_list,
                        "Average_Success_Probability_m2m_user_medium_list",
                    )
                    save_data_to_excel(
                        Average_Success_Probability_m2m_user_low_list,
                        "Average_Success_Probability_m2m_user_low_list",
                    )
                    save_data_to_excel(
                        Average_Success_H2H_Probability_list,
                        "Average_Success_H2H_Probability_list",
                    )
                    save_data_to_excel(
                        Average_Success_M2M_Probability_list,
                        "Average_Success_M2M_Probability_list",
                    )
                    save_data_to_excel(
                        Access_Success_Probability_average_list,
                        "Access_Success_Probability_average_list",
                    )

                    save_data_to_excel(
                        block_rate_h2h_high_list, "block_rate_h2h_high_list"
                    )
                    save_data_to_excel(
                        block_rate_h2h_medium_list, "block_rate_h2h_medium_list"
                    )
                    save_data_to_excel(
                        block_rate_h2h_low_list, "block_rate_h2h_low_list"
                    )
                    save_data_to_excel(
                        block_rate_m2m_high_list, "block_rate_m2m_high_list"
                    )
                    save_data_to_excel(
                        block_rate_m2m_medium_list, "block_rate_m2m_medium_list"
                    )
                    save_data_to_excel(
                        block_rate_m2m_low_list, "block_rate_m2m_low_list"
                    )
                    save_data_to_excel(block_rate_H2H_list, "block_rate_H2H_list")
                    save_data_to_excel(block_rate_M2M_list, "block_rate_M2M_list")
                    save_data_to_excel(block_rate_ALL_list, "block_rate_ALL_list")

                    save_data_to_excel(
                        Preamble_utilization_rate_test_frame_data_for_plot,
                        "Preamble_utilization_rate_test_frame_data_for_plot",
                    )
                    save_data_to_excel(
                        Preamble_utilization_rate_frame_data_for_plot,
                        "Preamble_utilization_rate_frame_data_for_plot",
                    )
                    save_data_to_excel(
                        Preamble_utilization_rate_h2h_high_data_for_plot,
                        "Preamble_utilization_rate_h2h_high_data_for_plot",
                    )
                    save_data_to_excel(
                        Preamble_utilization_rate_h2h_medium_data_for_plot,
                        "Preamble_utilization_rate_h2h_medium_data_for_plot",
                    )
                    save_data_to_excel(
                        Preamble_utilization_rate_h2h_low_data_for_plot,
                        "Preamble_utilization_rate_h2h_low_data_for_plot",
                    )
                    save_data_to_excel(
                        Preamble_utilization_rate_m2m_high_data_for_plot,
                        "Preamble_utilization_rate_m2m_high_data_for_plot",
                    )
                    save_data_to_excel(
                        Preamble_utilization_rate_m2m_medium_data_for_plot,
                        "Preamble_utilization_rate_m2m_medium_data_for_plot",
                    )
                    save_data_to_excel(
                        Preamble_utilization_rate_m2m_low_data_for_plot,
                        "Preamble_utilization_rate_m2m_low_data_for_plot",
                    )
                    save_data_to_excel(
                        Preamble_utilization_rate_H2H_user_for_plot,
                        "Preamble_utilization_rate_H2H_user_for_plot",
                    )
                    save_data_to_excel(
                        Preamble_utilization_rate_M2M_user_for_plot,
                        "Preamble_utilization_rate_M2M_user_for_plot",
                    )
                    save_data_to_excel(
                        Preamble_utilization_rate_average_data_for_plot,
                        "Preamble_utilization_rate_average_data_for_plot",
                    )
                    save_data_to_excel(loss_values_list, "loss_values_list")
                    save_data_to_excel(frames, "frames")
                    save_data_to_excel(values_list, "values_list")

            elif globalModeSelector == 2:
                for numOfRandomUserPerFrame in range(5, 101, 5):
                    print(f"numOfRandomUserPerFrame is {numOfRandomUserPerFrame}\n")

                    listForH2hUserHigh = []
                    listForH2hUserMedium = []
                    listForH2hUserLow = []
                    listForM2mUserHigh = []
                    listForM2mUserMedium = []
                    listForM2mUserLow = []
                    listForM2mUser = []
                    listForH2hUser = []
                    listForAllUser = []

                    total_served_user_list = []  # 总服务用户列表
                    averageAllUserPreambleAssignDelayList = []  # 平均用户前导码分配时延列表
                    averageH2hHighPreambleAssignDelayList = []  # 平均H2H高优先级用户前导码分配时延列表
                    averageH2hMediumPreambleAssignDelayList = []  # 平均H2H中优先级用户前导码分配时延列表
                    averageH2hLowPreambleAssignDelayList = []  # 平均H2H低优先级用户前导码分配时延列表
                    averageM2mHighPreambleAssignDelayList = []  # 平均M2M高优先级用户前导码分配时延列表
                    averageM2mMediumPreambleAssignDelayList = []
                    averageM2mLowPreambleAssignDelayList = []
                    averageH2hPreambleAssignDelayList = []
                    averageM2mPreambleAssignDelayList = []

                    Average_Success_Probability_h2h_user_high_list = []
                    Average_Success_Probability_h2h_user_medium_list = []
                    Average_Success_Probability_h2h_user_low_list = []
                    Average_Success_Probability_m2m_user_high_list = []
                    Average_Success_Probability_m2m_user_medium_list = []
                    Average_Success_Probability_m2m_user_low_list = []
                    Average_Success_H2H_Probability_list = []
                    Average_Success_M2M_Probability_list = []
                    Access_Success_Probability_average_list = []

                    avgAllUserPerFrameCollisionRateList = []  # 平均每帧用户碰撞率列表
                    avgH2hHighPerFrameCollisionRateList = []
                    avgH2hMediumPerFrameCollisionRateList = []
                    avgH2hLowPerFrameCollisionRateList = []
                    avgM2mHighPerFrameCollisionRateList = []
                    avgM2mMediumPerFrameCollisionRateList = []
                    avgM2mLowPerFrameCollisionRateList = []
                    avgH2hPerFrameCollisionRateList = []
                    avgM2mPerFrameCollisionRateList = []

                    block_rate_h2h_high_list = []
                    block_rate_h2h_medium_list = []
                    block_rate_h2h_low_list = []
                    block_rate_m2m_high_list = []
                    block_rate_m2m_medium_list = []
                    block_rate_m2m_low_list = []
                    block_rate_H2H_list = []
                    block_rate_M2M_list = []
                    block_rate_ALL_list = []

                    loss_list_for_plot = []
                    convergence_frame_list = []
                    convergence_value_list = []

                    Average_Preamble_utilization_rate_h2h_high_list = []
                    Average_Preamble_utilization_rate_h2h_medium_list = []
                    Average_Preamble_utilization_rate_h2h_low_list = []
                    Average_Preamble_utilization_rate_m2m_high_list = []
                    Average_Preamble_utilization_rate_m2m_medium_list = []
                    Average_Preamble_utilization_rate_m2m_low_list = []
                    Average_Preamble_utilization_rate_H2H_list = []
                    Average_Preamble_utilization_rate_M2M_list = []

                    frames = []
                    values_list = []
                    loss_values_list = []

                    # 初始化4个前导码列表
                    preambleListForM2mLow = PreambleList()
                    PreambleListForM2mMedium = PreambleList()
                    preambleListForM2mHigh = PreambleList()
                    preambleListForH2hLow = PreambleList()
                    PreambleListForH2hMedium = PreambleList()
                    preambleListForH2hHigh = PreambleList()

                    # 初始化阻塞率字典
                    block_probs = {
                        "h2h_high": 0.0,
                        "h2h_medium": 0.0,
                        "h2h_low": 0.0,
                        "m2m_high": 0.0,
                        "m2m_medium": 0.0,
                        "m2m_low": 0.0,
                        "H2H": 0.0,
                        "M2M": 0.0,
                        "ALL": 0.0,
                    }

                    # 初始化权重字典
                    user_priority_weights = {
                        "h2h": {
                            "high": 0.5,
                            "medium": 0.3,
                            "low": 0.2,
                        },
                        "m2m": {
                            "high": 0.5,
                            "medium": 0.3,
                            "low": 0.2,
                        },
                    }

                    user_type_weights = {"h2h": 0.7, "m2m": 0.3}
                    num1_of_h2h_user_high_per_frame = []
                    num1_of_h2h_user_medium_per_frame = []
                    num1_of_h2h_user_low_per_frame = []
                    num1_of_m2m_user_high_per_frame = []
                    num1_of_m2m_user_medium_per_frame = []
                    num1_of_m2m_user_low_per_frame = []

                    num_of_unserved_user_h2h_high_per_frame = []
                    num_of_unserved_user_h2h_medium_per_frame = []
                    num_of_unserved_user_h2h_low_per_frame = []
                    num_of_unserved_user_m2m_high_per_frame = []
                    num_of_unserved_user_m2m_medium_per_frame = []
                    num_of_unserved_user_m2m_low_per_frame = []

                    num_test_frame_new_input_m2m_high_list = []
                    num_test_frame_new_input_m2m_medium_list = []
                    num_test_frame_new_input_m2m_low_list = []
                    num_test_frame_new_input_h2h_high_list = []
                    num_tset_frame_new_input_h2h_medium_list = []
                    num_test_frame_new_input_h2h_low_list = []

                    num_of_remove_blocked_user_unserved_user_h2h_high_per_frame = []
                    num_of_remove_blocked_user_unserved_user_h2h_medium_per_frame = []
                    num_of_remove_blocked_user_unserved_user_h2h_low_per_frame = []
                    num_of_remove_blocked_user_unserved_user_m2m_high_per_frame = []
                    num_of_remove_blocked_user_unserved_user_m2m_medium_per_frame = []
                    num_of_remove_blocked_user_unserved_user_m2m_low_per_frame = []

                    rl()

                    save_data_to_excel(
                        averageAllUserPreambleAssignDelayList,
                        "averageAllUserPreambleAssignDelayList",
                    )
                    save_data_to_excel(
                        averageH2hHighPreambleAssignDelayList,
                        "averageH2hHighPreambleAssignDelayList",
                    )
                    save_data_to_excel(
                        averageH2hMediumPreambleAssignDelayList,
                        "averageH2hMediumPreambleAssignDelayList",
                    )
                    save_data_to_excel(
                        averageH2hLowPreambleAssignDelayList,
                        "averageH2hLowPreambleAssignDelayList",
                    )
                    save_data_to_excel(
                        averageM2mHighPreambleAssignDelayList,
                        "averageM2mHighPreambleAssignDelayList",
                    )
                    save_data_to_excel(
                        averageM2mMediumPreambleAssignDelayList,
                        "averageM2mMediumPreambleAssignDelayList",
                    )
                    save_data_to_excel(
                        averageM2mLowPreambleAssignDelayList,
                        "averageM2mLowPreambleAssignDelayList",
                    )
                    save_data_to_excel(
                        averageH2hPreambleAssignDelayList,
                        "averageH2hPreambleAssignDelayList",
                    )
                    save_data_to_excel(
                        averageM2mPreambleAssignDelayList,
                        "averageM2mPreambleAssignDelayList",
                    )

                    save_data_to_excel(
                        avgAllUserPerFrameCollisionRateList,
                        "avgAllUserPerFrameCollisionRateList",
                    )
                    save_data_to_excel(
                        avgH2hHighPerFrameCollisionRateList,
                        "avgH2hHighPerFrameCollisionRateList",
                    )
                    save_data_to_excel(
                        avgH2hMediumPerFrameCollisionRateList,
                        "avgH2hMediumPerFrameCollisionRateList",
                    )
                    save_data_to_excel(
                        avgH2hLowPerFrameCollisionRateList,
                        "avgH2hLowPerFrameCollisionRateList",
                    )
                    save_data_to_excel(
                        avgM2mHighPerFrameCollisionRateList,
                        "avgM2mHighPerFrameCollisionRateList",
                    )
                    save_data_to_excel(
                        avgM2mMediumPerFrameCollisionRateList,
                        "avgM2mMediumPerFrameCollisionRateList",
                    )
                    save_data_to_excel(
                        avgM2mLowPerFrameCollisionRateList,
                        "avgM2mLowPerFrameCollisionRateList",
                    )
                    save_data_to_excel(
                        avgH2hPerFrameCollisionRateList,
                        "avgH2hPerFrameCollisionRateList",
                    )
                    save_data_to_excel(
                        avgM2mPerFrameCollisionRateList,
                        "avgM2mPerFrameCollisionRateList",
                    )

                    save_data_to_excel(
                        Access_Success_Probility_frame_data_h2h_high_for_plot,
                        "Access_Success_Probility_frame_data_h2h_high_for_plot",
                    )
                    save_data_to_excel(
                        Access_Success_Probility_frame_data_h2h_medium_for_plot,
                        "Access_Success_Probility_frame_data_h2h_medium_for_plot",
                    )
                    save_data_to_excel(
                        Access_Success_Probility_frame_data_h2h_low_for_plot,
                        "Access_Success_Probility_frame_data_h2h_low_for_plot",
                    )
                    save_data_to_excel(
                        Access_Success_Probility_frame_data_m2m_high_for_plot,
                        "Access_Success_Probility_frame_data_m2m_high_for_plot",
                    )
                    save_data_to_excel(
                        Access_Success_Probility_frame_data_m2m_medium_for_plot,
                        "Access_Success_Probility_frame_data_m2m_medium_for_plot",
                    )
                    save_data_to_excel(
                        Access_Success_Probility_frame_data_m2m_low_for_plot,
                        "Access_Success_Probility_frame_data_m2m_low_for_plot",
                    )
                    save_data_to_excel(
                        Access_Success_Probility_frame_data_for_plot,
                        "Access_Success_Probility_frame_data_for_plot",
                    )
                    save_data_to_excel(
                        Access_Success_Probility_H2H_for_plot,
                        "Access_Success_Probility_H2H_for_plot",
                    )
                    save_data_to_excel(
                        Access_Success_Probility_M2M_for_plot,
                        "Access_Success_Probility_M2M_for_plot",
                    )

                    save_data_to_excel(
                        Average_Success_Probability_h2h_user_high_list,
                        "Average_Success_Probability_h2h_user_high_list",
                    )
                    save_data_to_excel(
                        Average_Success_Probability_h2h_user_medium_list,
                        "Average_Success_Probability_h2h_user_medium_list",
                    )
                    save_data_to_excel(
                        Average_Success_Probability_h2h_user_low_list,
                        "Average_Success_Probability_h2h_user_low_list",
                    )
                    save_data_to_excel(
                        Average_Success_Probability_m2m_user_high_list,
                        "Average_Success_Probability_m2m_user_high_list",
                    )
                    save_data_to_excel(
                        Average_Success_Probability_m2m_user_medium_list,
                        "Average_Success_Probability_m2m_user_medium_list",
                    )
                    save_data_to_excel(
                        Average_Success_Probability_m2m_user_low_list,
                        "Average_Success_Probability_m2m_user_low_list",
                    )
                    save_data_to_excel(
                        Average_Success_H2H_Probability_list,
                        "Average_Success_H2H_Probability_list",
                    )
                    save_data_to_excel(
                        Average_Success_M2M_Probability_list,
                        "Average_Success_M2M_Probability_list",
                    )
                    save_data_to_excel(
                        Access_Success_Probability_average_list,
                        "Access_Success_Probability_average_list",
                    )

                    save_data_to_excel(
                        block_rate_h2h_high_list, "block_rate_h2h_high_list"
                    )
                    save_data_to_excel(
                        block_rate_h2h_medium_list, "block_rate_h2h_medium_list"
                    )
                    save_data_to_excel(
                        block_rate_h2h_low_list, "block_rate_h2h_low_list"
                    )
                    save_data_to_excel(
                        block_rate_m2m_high_list, "block_rate_m2m_high_list"
                    )
                    save_data_to_excel(
                        block_rate_m2m_medium_list, "block_rate_m2m_medium_list"
                    )
                    save_data_to_excel(
                        block_rate_m2m_low_list, "block_rate_m2m_low_list"
                    )
                    save_data_to_excel(block_rate_H2H_list, "block_rate_H2H_list")
                    save_data_to_excel(block_rate_M2M_list, "block_rate_M2M_list")
                    save_data_to_excel(block_rate_ALL_list, "block_rate_ALL_list")

                    save_data_to_excel(
                        Preamble_utilization_rate_test_frame_data_for_plot,
                        "Preamble_utilization_rate_test_frame_data_for_plot",
                    )
                    save_data_to_excel(
                        Preamble_utilization_rate_frame_data_for_plot,
                        "Preamble_utilization_rate_frame_data_for_plot",
                    )
                    save_data_to_excel(
                        Preamble_utilization_rate_h2h_high_data_for_plot,
                        "Preamble_utilization_rate_h2h_high_data_for_plot",
                    )
                    save_data_to_excel(
                        Preamble_utilization_rate_h2h_medium_data_for_plot,
                        "Preamble_utilization_rate_h2h_medium_data_for_plot",
                    )
                    save_data_to_excel(
                        Preamble_utilization_rate_h2h_low_data_for_plot,
                        "Preamble_utilization_rate_h2h_low_data_for_plot",
                    )
                    save_data_to_excel(
                        Preamble_utilization_rate_m2m_high_data_for_plot,
                        "Preamble_utilization_rate_m2m_high_data_for_plot",
                    )
                    save_data_to_excel(
                        Preamble_utilization_rate_m2m_medium_data_for_plot,
                        "Preamble_utilization_rate_m2m_medium_data_for_plot",
                    )
                    save_data_to_excel(
                        Preamble_utilization_rate_m2m_low_data_for_plot,
                        "Preamble_utilization_rate_m2m_low_data_for_plot",
                    )
                    save_data_to_excel(
                        Preamble_utilization_rate_H2H_user_for_plot,
                        "Preamble_utilization_rate_H2H_user_for_plot",
                    )
                    save_data_to_excel(
                        Preamble_utilization_rate_M2M_user_for_plot,
                        "Preamble_utilization_rate_M2M_user_for_plot",
                    )
                    save_data_to_excel(
                        Preamble_utilization_rate_average_data_for_plot,
                        "Preamble_utilization_rate_average_data_for_plot",
                    )

                    save_data_to_excel(loss_values_list, "loss_values_list")
                    save_data_to_excel(frames, "frames")
                    save_data_to_excel(values_list, "values_list")

    print("Congratulations! All processes has finished!")
